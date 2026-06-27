from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_STARTS = {
    "numeric": 101,
    "open": 1000,
    "multi": 1101,
}

# The marker may be changed later if a questionnaire uses a convention other
# than vQuestionoOption for its conditional open-text variables.
DEFAULT_OPEN_OPTION_MARKER = "o"


@dataclass(frozen=True)
class AllVar:
    row: int
    name: str
    kind: str
    width: int


@dataclass(frozen=True)
class PrecheckIssue:
    severity: str
    sheet: str
    row: str
    column: str
    code: str
    message: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, cell in enumerate(ws[1], start=1):
        text = cell_text(cell.value)
        if text:
            result[text] = index
    return result


def sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        if name.startswith(prefix) and name not in {"複選題內_互斥", "複選題變項清單"}:
            return wb[name]
    raise ValueError(f"missing sheet with prefix {prefix!r}")


def issue(severity: str, sheet: str, row: Any, column: str, code: str, message: str) -> PrecheckIssue:
    return PrecheckIssue(severity, sheet, str(row or ""), column, code, message)


def all_sheet_names(wb: openpyxl.Workbook) -> list[str]:
    return [name for name in wb.sheetnames if name.strip().lower() == "all"]


def precheck_all_sheet(wb: openpyxl.Workbook) -> list[PrecheckIssue]:
    issues: list[PrecheckIssue] = []
    matches = all_sheet_names(wb)
    if len(matches) != 1:
        issues.append(issue("error", "", "", "", "all_sheet_count", f"Expected exactly one all sheet, found {len(matches)}: {', '.join(matches) or '(none)'}"))
        return issues
    sheet_name = matches[0]
    if sheet_name != "all":
        issues.append(issue("error", sheet_name, "", "", "all_sheet_name", "The variable metadata sheet must be named exactly all."))
        return issues
    ws = wb[sheet_name]
    headers = read_headers(ws)
    required = ["變項名稱", "變項屬性", "寬度"]
    for header in required:
        if header not in headers:
            issues.append(issue("error", sheet_name, 1, header, "missing_header", f"all sheet missing required header {header}."))
    if any(item.severity == "error" for item in issues):
        return issues
    data_count = 0
    seen: dict[str, int] = {}
    allowed_kinds = {"數值", "字串", "文字"}
    for row_idx, values in data_rows(ws):
        name = cell_text(values[headers["變項名稱"] - 1])
        kind = cell_text(values[headers["變項屬性"] - 1])
        width = cell_text(values[headers["寬度"] - 1])
        if name.lower() == "last":
            continue
        data_count += 1
        if not name:
            issues.append(issue("error", sheet_name, row_idx, "變項名稱", "blank_var_name", "Variable name is blank."))
        else:
            key = name.lower()
            if key in seen:
                issues.append(issue("error", sheet_name, row_idx, "變項名稱", "duplicate_var", f"Variable {name} duplicates row {seen[key]}."))
            else:
                seen[key] = row_idx
        if not kind:
            issues.append(issue("error", sheet_name, row_idx, "變項屬性", "blank_kind", f"{name or '(blank)'} has blank variable type."))
        elif kind not in allowed_kinds:
            issues.append(issue("warning", sheet_name, row_idx, "變項屬性", "unknown_kind", f"{name} has unrecognized variable type {kind}; it will not be auto-filled into numeric/open sections."))
        if not width:
            issues.append(issue("error", sheet_name, row_idx, "寬度", "blank_width", f"{name or '(blank)'} has blank width."))
        elif not width.isdigit() or int(width) <= 0:
            issues.append(issue("error", sheet_name, row_idx, "寬度", "invalid_width", f"{name} width must be a positive integer, got {width}."))
    if data_count == 0:
        issues.append(issue("error", sheet_name, "", "", "no_variables", "all sheet has no variable rows."))
    return issues


def issue_counts(issues: list[PrecheckIssue]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in issues:
        counts[item.severity] = counts.get(item.severity, 0) + 1
    return counts


def data_rows(ws: openpyxl.worksheet.worksheet.Worksheet):
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(cell_text(value) for value in values):
            yield row_idx, values


def read_existing_by_key(ws: openpyxl.worksheet.worksheet.Worksheet, key_header: str) -> dict[str, dict[str, Any]]:
    headers = read_headers(ws)
    if key_header not in headers:
        return {}
    existing: dict[str, dict[str, Any]] = {}
    for _row_idx, values in data_rows(ws):
        key = cell_text(values[headers[key_header] - 1])
        if not key:
            continue
        existing[key] = {
            name: values[col_idx - 1] if col_idx - 1 < len(values) else None
            for name, col_idx in headers.items()
        }
    return existing


def read_all_vars(wb: openpyxl.Workbook) -> tuple[list[AllVar], int | None]:
    ws = wb["all"]
    headers = read_headers(ws)
    for required in ("變項名稱", "變項屬性", "寬度"):
        if required not in headers:
            raise ValueError(f"all sheet missing {required}")
    last_row = None
    for row_idx in range(2, ws.max_row + 1):
        name = cell_text(ws.cell(row_idx, headers["變項名稱"]).value)
        if name.lower() == "last":
            last_row = row_idx
            break
    # A new project may not yet have a technical-variable boundary.  In that
    # case, treat all variables as formal questionnaire variables.
    start_row = last_row + 1 if last_row is not None else 2

    variables: list[AllVar] = []
    for row_idx in range(start_row, ws.max_row + 1):
        name = cell_text(ws.cell(row_idx, headers["變項名稱"]).value)
        if not name:
            continue
        kind = cell_text(ws.cell(row_idx, headers["變項屬性"]).value)
        try:
            width = int(cell_text(ws.cell(row_idx, headers["寬度"]).value) or "2")
        except ValueError:
            width = 2
        variables.append(AllVar(row_idx, name, kind, width))
    return variables, last_row


def is_numeric(var: AllVar) -> bool:
    return var.kind == "數值"


def is_text(var: AllVar) -> bool:
    return var.kind in {"字串", "文字"}


def multi_match(var_name: str) -> re.Match[str] | None:
    return re.match(r"^(?P<group>v.+?)m(?P<option>\d+)$", var_name, flags=re.IGNORECASE)


def display_qid(group: str) -> str:
    return group[1:] if group.startswith("v") and len(group) > 1 else group


def option_number(var_name: str) -> int:
    match = multi_match(var_name)
    return int(match.group("option")) if match else 0


def option_width(var_name: str) -> int:
    match = multi_match(var_name)
    return len(match.group("option")) if match else 0


def clear_data(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def first_existing_m(ws: openpyxl.worksheet.worksheet.Worksheet, fallback: int) -> int:
    headers = read_headers(ws)
    if "m" not in headers:
        return fallback
    for row_idx, values in data_rows(ws):
        text = cell_text(values[headers["m"] - 1])
        if text.isdigit():
            return int(text)
    return fallback


def parse_open_option(var_name: str, marker: str = DEFAULT_OPEN_OPTION_MARKER) -> tuple[str, str]:
    """Return the original question variable and its selected option.

    The marker is deliberately case-sensitive: vO1o88 means vO1 + 88, while
    the capital O in the question identifier must never be treated as a marker.
    """
    if not marker:
        return "", ""
    marker_index = var_name.rfind(marker)
    if marker_index <= 0:
        return "", ""
    option = var_name[marker_index + len(marker) :]
    if not option.isdigit():
        return "", ""
    return var_name[:marker_index], str(int(option))


def infer_open_spss_fields(var_name: str, multi_var_names: set[str], marker: str) -> dict[str, str]:
    base_var, option = parse_open_option(var_name, marker)
    if "city" in var_name.lower():
        city_var = re.sub(r"_oth$", "", var_name, flags=re.IGNORECASE)
        return {
            "題號": city_var,
            "題號變項名稱": city_var,
            "選項數值": "29",
            "var2_new": city_var,
            "range_new": "29",
            "n": "3",
        }
    if base_var and option:
        multi_parent = f"{base_var}m{option}"
        is_multi = multi_parent in multi_var_names
        parent = multi_parent if is_multi else base_var
        return {
            "題號": base_var,
            "題號變項名稱": parent,
            "選項數值": option,
            "var2_new": parent,
            "range_new": "1" if is_multi else option,
            "n": "2" if len(option) == 1 else str(len(option)),
        }
    # General open-text fields have no controlling option.  They are still
    # listed in SPSS for review, but do not receive skip-logic conditions.
    return {
        "題號": var_name,
        "題號變項名稱": "",
        "選項數值": "",
        "var2_new": "",
        "range_new": "",
        "n": "",
    }


def write_row(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, values: dict[str, Any]) -> None:
    headers = read_headers(ws)
    for header, col_idx in headers.items():
        if header in values:
            ws.cell(row_idx, col_idx).value = values[header]


def build_multi_groups(variables: list[AllVar]) -> dict[str, list[AllVar]]:
    groups: dict[str, list[AllVar]] = {}
    for var in variables:
        if not is_numeric(var):
            continue
        match = multi_match(var.name)
        if not match:
            continue
        groups.setdefault(match.group("group"), []).append(var)
    for group_vars in groups.values():
        group_vars.sort(key=lambda item: item.row)
    return groups


def fill_numeric_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    variables: list[AllVar],
    existing: dict[str, dict[str, Any]],
) -> int:
    start_m = first_existing_m(ws, DEFAULT_STARTS["numeric"])
    clear_data(ws)
    numeric_vars = [var for var in variables if is_numeric(var)]
    for offset, var in enumerate(numeric_vars):
        m = start_m + offset
        old = existing.get(var.name, {})
        row_values = dict(old)
        row_values.update(
            {
                "m": m,
                "p": f"=A{offset + 2}",
                "var": var.name,
                "變項屬性": "數值",
                "寬度": var.width,
                "in_or_not": old.get("in_or_not") or "range",
            }
        )
        write_row(ws, offset + 2, row_values)
    return len(numeric_vars)


def fill_open_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    variables: list[AllVar],
    existing: dict[str, dict[str, Any]],
    multi_var_names: set[str],
    marker: str,
) -> int:
    start_m = first_existing_m(ws, DEFAULT_STARTS["open"])
    clear_data(ws)
    open_vars = [var for var in variables if is_text(var)]
    for offset, var in enumerate(open_vars):
        m = start_m + offset
        old = existing.get(var.name, {})
        inferred = infer_open_spss_fields(var.name, multi_var_names, marker)
        is_multi = inferred["var2_new"] in multi_var_names
        row_values = dict(old)
        row_values.update(
            {
                "m": m,
                "p": f"=A{offset + 2}",
                "var": var.name,
                "變項屬性": var.kind,
                "寬度": var.width,
                "是否複選": 1 if is_multi else 0,
                **inferred,
            }
        )
        write_row(ws, offset + 2, row_values)
    return len(open_vars)


def fill_multi_var_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, groups: dict[str, list[AllVar]]) -> int:
    clear_data(ws)
    row_idx = 2
    for group, vars_ in groups.items():
        for var in vars_:
            write_row(ws, row_idx, {"變項名稱": var.name, "分組": group})
            row_idx += 1
    return row_idx - 2


def read_multi_var_names(ws: openpyxl.worksheet.worksheet.Worksheet) -> set[str]:
    headers = read_headers(ws)
    var_col = headers.get("變項名稱")
    if not var_col:
        raise ValueError(f"{ws.title} missing 變項名稱 column")
    return {
        cell_text(ws.cell(row_idx, var_col).value)
        for row_idx in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row_idx, var_col).value)
    }


def preserve_main_multi_defaults(old: dict[str, Any], vars_: list[AllVar]) -> dict[str, Any]:
    options = [option_number(var.name) for var in vars_ if option_number(var.name)]
    start = min(options) if options else 1
    end = max(options) if options else len(vars_)
    width = max([option_width(var.name) for var in vars_] or [2])
    inferred_no_response = ""
    for candidate in (90, 96, 99):
        if candidate in options:
            inferred_no_response = str(candidate)
            break
    row_values = dict(old)
    row_values.update(
        {
            "寬度": width,
            "選項編號_起": start,
            "選項編號_迄": end,
            "選項數量": len(vars_),
            "迴圈進行次數": max(len(vars_) - 1, 0),
        }
    )
    if not cell_text(row_values.get("不可選擇數量")):
        row_values["不可選擇數量"] = 1 if inferred_no_response else 0
    if inferred_no_response and not cell_text(row_values.get("全部無反應選項")):
        row_values["全部無反應選項"] = inferred_no_response
    if inferred_no_response and not cell_text(row_values.get("無反應選項1")):
        row_values["無反應選項1"] = inferred_no_response
    return row_values


def fill_multi_main_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    groups: dict[str, list[AllVar]],
    existing: dict[str, dict[str, Any]],
) -> int:
    start_m = first_existing_m(ws, DEFAULT_STARTS["multi"])
    clear_data(ws)
    for offset, (group, vars_) in enumerate(groups.items()):
        m = start_m + offset
        qid = display_qid(group)
        old = existing.get(qid) or existing.get(group) or {}
        row_values = preserve_main_multi_defaults(old, vars_)
        row_values.update(
            {
                "m": m,
                "p": f"=A{offset + 2}",
                "題號": qid,
            }
        )
        write_row(ws, offset + 2, row_values)
    return len(groups)


def fill_sections(workbook_path: Path, apply: bool, open_option_marker: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path)
    precheck_issues = precheck_all_sheet(wb)
    precheck_report = {
        "issue_counts": issue_counts(precheck_issues),
        "issues": [item.__dict__ for item in precheck_issues],
    }
    if precheck_report["issue_counts"].get("error", 0):
        return {
            "workbook": str(workbook_path),
            "applied": False,
            "skipped_due_to_precheck": True,
            "all_precheck": precheck_report,
            "warnings": [],
        }
    variables, last_row = read_all_vars(wb)
    numeric_ws = sheet_by_prefix(wb, "數值題")
    open_ws = sheet_by_prefix(wb, "開放欄位")
    multi_ws = sheet_by_prefix(wb, "複選題")
    multi_var_ws = wb["複選題變項清單"]

    numeric_existing = read_existing_by_key(numeric_ws, "var")
    open_existing = read_existing_by_key(open_ws, "var")
    multi_existing = read_existing_by_key(multi_ws, "題號")
    multi_groups = build_multi_groups(variables)

    report: dict[str, Any] = {
        "workbook": str(workbook_path),
        "last_row": last_row,
        "all_start_row": (last_row + 1) if last_row is not None else 2,
        "last_found": last_row is not None,
        "open_option_marker": open_option_marker,
        "all_vars_after_last": len(variables),
        "numeric_vars": sum(1 for var in variables if is_numeric(var)),
        "text_vars": sum(1 for var in variables if is_text(var)),
        "multi_groups": len(multi_groups),
        "multi_vars": sum(len(vars_) for vars_ in multi_groups.values()),
        "applied": apply,
        "warnings": [],
        "all_precheck": precheck_report,
    }

    if apply:
        backup_path = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_fill_from_all.xlsx"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(workbook_path, backup_path)
        report["backup"] = str(backup_path)

        report["numeric_rows_written"] = fill_numeric_sheet(numeric_ws, variables, numeric_existing)
        report["multi_var_rows_written"] = fill_multi_var_sheet(multi_var_ws, multi_groups)
        multi_var_names = read_multi_var_names(multi_var_ws)
        report["open_rows_written"] = fill_open_sheet(
            open_ws, variables, open_existing, multi_var_names, open_option_marker
        )
        report["multi_main_rows_written"] = fill_multi_main_sheet(multi_ws, multi_groups, multi_existing)
        wb.save(workbook_path)
    else:
        report["numeric_rows_to_write"] = report["numeric_vars"]
        report["open_rows_to_write"] = report["text_vars"]
        report["multi_var_rows_to_write"] = report["multi_vars"]
        report["multi_main_rows_to_write"] = report["multi_groups"]
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill numeric/open/multi Excel sections from all sheet after variable 'last'.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument(
        "--open-option-marker",
        default=DEFAULT_OPEN_OPTION_MARKER,
        help="Case-sensitive marker before an open-field option number (default: o).",
    )
    args = parser.parse_args()

    report = fill_sections(args.workbook, args.apply, args.open_option_marker)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: v for k, v in report.items() if k != "warnings"}, ensure_ascii=False))
    if report.get("warnings"):
        print(f"warnings: {len(report['warnings'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
