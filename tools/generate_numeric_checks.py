from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class NumericRule:
    row: int
    m: str
    p: str
    s: str
    s_value: str
    var: str
    width: int
    in_or_not: str
    ranges: list[str]
    decimal_rule: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return wb[name]
    raise ValueError(f"missing sheet prefix {prefix}")


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(value): idx
        for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if cell_text(value)
    }


def resolved_mp(value: Any, row_values: dict[str, Any]) -> str:
    text = cell_text(value)
    if text.startswith("="):
        return cell_text(row_values.get("m"))
    return text


def resolved_optional_id(value: Any, row_values: dict[str, Any], fallback_key: str = "") -> str:
    text = cell_text(value)
    if text.startswith("="):
        return cell_text(row_values.get(fallback_key)) if fallback_key else ""
    return text


def load_rules(workbook_path: Path) -> tuple[list[NumericRule], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    ws = sheet_by_prefix(wb, "數值題")
    h = headers(ws)
    required = ["m", "p", "var", "變項屬性", "寬度", "in_or_not", "r1", "r2", "r3", "r4"]
    missing = [name for name in required if name not in h]
    if missing:
        raise ValueError(f"numeric sheet missing headers: {missing}")
    rules: list[NumericRule] = []
    skipped: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = {name: row[idx - 1] for name, idx in h.items()}
        var = cell_text(values.get("var"))
        if not var:
            continue
        if cell_text(values.get("變項屬性")) != "數值":
            continue
        ranges = [cell_text(values.get(col)) for col in ("r1", "r2", "r3", "r4") if cell_text(values.get(col))]
        if not ranges:
            skipped.append({"row": row_idx, "var": var, "reason": "blank r1-r4"})
            continue
        m = cell_text(values.get("m"))
        p = resolved_mp(values.get("p"), values)
        s = resolved_optional_id(values.get("s"), values, "m") if "s" in h else ""
        s_value = resolved_optional_id(values.get("s="), values) if "s=" in h else ""
        if not m or not p:
            skipped.append({"row": row_idx, "var": var, "reason": "blank m or p"})
            continue
        try:
            width = int(cell_text(values.get("寬度")) or "2")
        except ValueError:
            width = 2
        rules.append(
            NumericRule(
                row=row_idx,
                m=m,
                p=p,
                s=s,
                s_value=s_value,
                var=var,
                width=width,
                in_or_not=cell_text(values.get("in_or_not")) or "range",
                ranges=ranges,
                decimal_rule=cell_text(values.get("可輸入小數點")),
            )
        )
    return rules, skipped


def expand_range_args(values: list[str]) -> list[str]:
    args: list[str] = []
    for value in values:
        if "," in value:
            args.extend(part.strip() for part in value.split(",") if part.strip())
        else:
            args.extend([value, value])
    return args


def wrap_args(parts: list[str], indent: str = "  ", max_len: int = 160) -> list[str]:
    lines: list[str] = []
    current = indent
    for index, part in enumerate(parts):
        token = part + ("," if index < len(parts) - 1 else "")
        if current.strip() and len(current) + len(token) + 1 > max_len:
            lines.append(current.rstrip())
            current = indent + token
        else:
            current += ("" if current == indent else " ") + token
    if current.strip():
        lines.append(current.rstrip())
    return lines


def valid_expr(rule: NumericRule) -> str:
    args = expand_range_args(rule.ranges)
    fn = "any" if rule.in_or_not.lower() == "any" else "range"
    expr = f"{fn}({rule.var},{','.join(args)})"
    if rule.decimal_rule:
        expr = f"({expr}) & ({rule.decimal_rule})"
    return expr


def render_valid_expr(rule: NumericRule) -> list[str]:
    args = [rule.var, *expand_range_args(rule.ranges)]
    fn = "any" if rule.in_or_not.lower() == "any" else "range"
    expr = valid_expr(rule)
    if len(expr) <= 160:
        return [expr]
    if rule.decimal_rule:
        return ["(" + f"{fn}("] + wrap_args(args) + [f")) & ({rule.decimal_rule})"]
    return [f"{fn}("] + wrap_args(args) + [")"]


def render_do_if(rule: NumericRule) -> list[str]:
    valid_lines = render_valid_expr(rule)
    if len(valid_lines) == 1:
        return [f"do if not {valid_lines[0]} | sys({rule.var})."]
    lines = [f"do if not {valid_lines[0]}"]
    lines.extend(valid_lines[1:-1])
    lines.append(f"{valid_lines[-1]} | sys({rule.var}).")
    return lines


def render_comment(rule: NumericRule) -> list[str]:
    text = f"*{rule.var}={' '.join(rule.ranges)} ."
    if len(text) <= 160:
        return [text]
    return [f"*{rule.var}=<long valid range; see do if range arguments below>."]


def render_rule(rule: NumericRule) -> str:
    return "\n".join(
        [
            *render_comment(rule),
            *render_do_if(rule),
            f'compute m{rule.m}=concat("{rule.var}=",string({rule.var},f{rule.width})).',
            f'compute p{rule.p}="{rule.var}為不合理值或遺漏值".',
            *([f"compute s{rule.s}={rule.s_value}."] if rule.s and rule.s_value else []),
            "end if.",
            "Exec.",
            "",
        ]
    )


def render_spss(rules: list[NumericRule]) -> str:
    parts = [
        "* Encoding: UTF-8.",
        "**一、不合理值檢核.",
        "* SYNTAXWORK_BEGIN_NUMERIC.",
    ]
    parts.extend(render_rule(rule).rstrip() for rule in rules)
    parts.append("* SYNTAXWORK_END_NUMERIC.")
    return "\n\n".join(parts).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    rules, skipped = load_rules(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(render_spss(rules), encoding="utf-8-sig")
    report = {"rules": len(rules), "skipped": len(skipped), "skipped_rows": skipped, "output": str(args.output)}
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
