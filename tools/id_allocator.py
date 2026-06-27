from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(cell.value): col_idx
        for col_idx, cell in enumerate(ws[1], start=1)
        if cell_text(cell.value)
    }


def numeric_id(value: Any) -> int | None:
    text = cell_text(value)
    if text.startswith("="):
        return None
    if not text.isdigit():
        return None
    return int(text)


def max_m_in_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    h = headers(ws)
    if "m" not in h:
        return 0
    result = 0
    for row_idx in range(2, ws.max_row + 1):
        value = numeric_id(ws.cell(row_idx, h["m"]).value)
        if value:
            result = max(result, value)
    return result


def next_x01_after(value: int) -> int:
    return ((value // 100) + 1) * 100 + 1 if value else 101


def max_m_before_sheet(wb: openpyxl.Workbook, target_sheet: str) -> int:
    result = 0
    for sheet_name in wb.sheetnames:
        if sheet_name == target_sheet:
            break
        if sheet_name == "all" or sheet_name == "複選題變項清單":
            continue
        result = max(result, max_m_in_sheet(wb[sheet_name]))
    return result


def next_logic_start(wb: openpyxl.Workbook, logic_sheet: str = "邏輯組") -> int:
    return next_x01_after(max_m_before_sheet(wb, logic_sheet))


def next_external_start(workbook_path: Path, external_sheet: str = "檢核項目清單", logic_sheet: str = "邏輯組") -> int:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    logic_max = max_m_in_sheet(wb[logic_sheet]) if logic_sheet in wb.sheetnames else 0
    if logic_max:
        return next_x01_after(logic_max)
    if external_sheet in wb.sheetnames:
        return next_x01_after(max_m_before_sheet(wb, external_sheet))
    max_before = 0
    for sheet_name in wb.sheetnames:
        if sheet_name == "all" or sheet_name == "複選題變項清單":
            continue
        max_before = max(max_before, max_m_in_sheet(wb[sheet_name]))
    return next_x01_after(max_before)


def assign_ids(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_id: int,
    *,
    fill_s: bool = False,
    logic_columns: tuple[str, ...] = ("條件", "應答", "不應答", "限制", "互斥"),
    clear_blank_ids: bool = True,
) -> dict[str, int]:
    h = headers(ws)
    if "m" not in h or "p" not in h:
        return {"assigned": 0, "start_id": start_id, "end_id": 0, "filled_s": 0}
    next_id = start_id
    assigned = 0
    filled_s = 0
    for row_idx in range(2, ws.max_row + 1):
        if not any(name in h and cell_text(ws.cell(row_idx, h[name]).value) for name in logic_columns):
            if clear_blank_ids:
                ws.cell(row_idx, h["m"]).value = None
                ws.cell(row_idx, h["p"]).value = None
                if fill_s and "s" in h:
                    ws.cell(row_idx, h["s"]).value = None
            continue
        ws.cell(row_idx, h["m"]).value = next_id
        ws.cell(row_idx, h["p"]).value = f"=A{row_idx}"
        if fill_s and "s" in h:
            ws.cell(row_idx, h["s"]).value = next_id
            filled_s += 1
        if fill_s and "s=" in h and cell_text(ws.cell(row_idx, h["s="]).value).startswith("="):
            ws.cell(row_idx, h["s="]).value = None
        assigned += 1
        next_id += 1
    return {
        "assigned": assigned,
        "start_id": start_id,
        "end_id": next_id - 1 if assigned else 0,
        "filled_s": filled_s,
    }
