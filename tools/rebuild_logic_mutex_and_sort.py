from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from pathlib import Path

import openpyxl

from id_allocator import assign_ids, next_logic_start
from refresh_logic_group_from_docx import cell_text, load_all_vars
from refresh_logic_mutex_from_docx import LOGIC_SHEET, MUTEX_HEADER, collect_mutex_rules, ensure_mutex_header


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {cell_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}


def all_var_order(workbook_path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["all"]
    order: dict[str, int] = {}
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        name = cell_text(row[0])
        if name:
            order[name.lower()] = index
    return order


def condition_order(text: str, order: dict[str, int]) -> int:
    vars_found = re.findall(r"\bv[A-Za-z_][A-Za-z0-9_]*\b", text or "")
    values = [order[var_name.lower()] for var_name in vars_found if var_name.lower() in order]
    return min(values) if values else 10**9


def row_order(row_values: dict[str, str], order: dict[str, int], original_index: int) -> tuple[int, int]:
    fields = [
        row_values.get("條件", ""),
        row_values.get("限制", ""),
        row_values.get(MUTEX_HEADER, ""),
        row_values.get("應答", ""),
        row_values.get("不應答", ""),
    ]
    rank = min(condition_order(field, order) for field in fields)
    return rank, original_index


def remove_mutex_only_rows(ws: openpyxl.worksheet.worksheet.Worksheet, h: dict[str, int]) -> int:
    if MUTEX_HEADER not in h:
        return 0
    removed = 0
    for row_idx in range(ws.max_row, 1, -1):
        has_mutex = bool(cell_text(ws.cell(row_idx, h[MUTEX_HEADER]).value))
        has_other_logic = any(
            cell_text(ws.cell(row_idx, h[name]).value)
            for name in ("應答", "不應答", "限制")
            if name in h
        )
        if has_mutex and not has_other_logic:
            ws.delete_rows(row_idx)
            removed += 1
    return removed


def append_mutex_rules(ws: openpyxl.worksheet.worksheet.Worksheet, h: dict[str, int], rules) -> int:
    existing = {
        (cell_text(ws.cell(row_idx, h["條件"]).value), cell_text(ws.cell(row_idx, h[MUTEX_HEADER]).value))
        for row_idx in range(2, ws.max_row + 1)
    }
    appended = 0
    for rule in rules:
        key = (rule.condition, rule.mutex)
        if key in existing:
            continue
        row_idx = ws.max_row + 1
        ws.cell(row_idx, h["條件"]).value = rule.condition
        ws.cell(row_idx, h[MUTEX_HEADER]).value = rule.mutex
        existing.add(key)
        appended += 1
    return appended


def sort_logic_rows(ws: openpyxl.worksheet.worksheet.Worksheet, h: dict[str, int], order: dict[str, int]) -> None:
    max_col = ws.max_column
    rows: list[tuple[tuple[int, int], list[object]]] = []
    for offset, row_idx in enumerate(range(2, ws.max_row + 1), start=1):
        values = [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]
        by_header = {
            name: cell_text(values[col_idx - 1])
            for name, col_idx in h.items()
            if 1 <= col_idx <= len(values)
        }
        rows.append((row_order(by_header, order, offset), values))
    rows.sort(key=lambda item: item[0])
    for row_offset, (_key, values) in enumerate(rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_offset, col_idx).value = value
    fix_row_relative_p_formulas(ws, h)


def fix_row_relative_p_formulas(ws: openpyxl.worksheet.worksheet.Worksheet, h: dict[str, int]) -> int:
    if "p" not in h:
        return 0
    fixed = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, h["p"])
        value = cell_text(cell.value)
        if re.fullmatch(r"=A\d+", value, flags=re.IGNORECASE):
            new_value = f"=A{row_idx}"
            if value.upper() != new_value.upper():
                cell.value = new_value
                fixed += 1
    return fixed


def write_csv(path: Path, rules, review: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["status", "qid", "option_code", "條件", MUTEX_HEADER, "reason", "bracket", "source"]
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for rule in rules:
            writer.writerow(
                {
                    "status": "rebuilt",
                    "qid": rule.qid,
                    "option_code": rule.option_code,
                    "條件": rule.condition,
                    MUTEX_HEADER: rule.mutex,
                    "bracket": rule.bracket,
                    "source": rule.source_text,
                }
            )
        for item in review:
            writer.writerow({"status": "review", **item})


def rebuild(
    docx_path: Path,
    workbook_path: Path,
    *,
    output_csv: Path | None = None,
    report_path: Path | None = None,
    dry_run: bool = False,
    allocate_ids: bool = False,
    fill_s: bool = False,
) -> dict[str, object]:
    all_vars = load_all_vars(workbook_path)
    rules, review = collect_mutex_rules(docx_path, all_vars)
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[LOGIC_SHEET]
    h = ensure_mutex_header(ws)
    removed = remove_mutex_only_rows(ws, h)
    h = headers(ws)
    appended = append_mutex_rules(ws, h, rules)
    sort_logic_rows(ws, h, all_var_order(workbook_path))
    id_assignment = (
        assign_ids(ws, next_logic_start(wb, LOGIC_SHEET), fill_s=fill_s)
        if allocate_ids
        else {"assigned": 0, "start_id": 0, "end_id": 0, "filled_s": 0}
    )

    if not dry_run:
        backup = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_logic_mutex_rebuild_sort.xlsx"
        backup.parent.mkdir(parents=True, exist_ok=True)
        if not backup.exists():
            shutil.copy2(workbook_path, backup)
        wb.save(workbook_path)

    if output_csv:
        write_csv(output_csv, rules, review)
    report = {
        "docx": str(docx_path),
        "workbook": str(workbook_path),
        "dry_run": dry_run,
        "cross_question_mutex_rules": len(rules),
        "removed_mutex_only_rows": removed,
        "appended_mutex_rows": appended,
        "id_assignment": id_assignment,
        "review": len(review),
    }
    if report_path:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Rebuild cross-question logic mutex rows and sort 邏輯組 by condition variable order.")
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output-csv", type=Path)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--allocate-ids", action="store_true", help="Fill m/p ids for the final sorted 邏輯組 rows.")
    parser.add_argument("--fill-s", action="store_true", help="Also fill s with the same id when the 邏輯組 sheet has an s column.")
    args = parser.parse_args()
    print(
        json.dumps(
            rebuild(
                args.docx,
                args.workbook,
                output_csv=args.output_csv,
                report_path=args.report,
                dry_run=args.dry_run,
                allocate_ids=args.allocate_ids,
                fill_s=args.fill_s,
            ),
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
