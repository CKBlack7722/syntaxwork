from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl
from docx import Document


SHEET_NAME = "檢核項目清單"
HEADERS = ["m", "p", "s", "s=", "項目編號", "題號", "變項名稱", "檢核說明", "備註", "條件", "額外列出變項"]
SOURCE_HEADERS = {"題號", "變項名稱", "檢核說明", "備註"}


@dataclass
class VarSpec:
    name: str
    kind: str
    width: int

    @property
    def is_numeric(self) -> bool:
        return self.kind == "數值"


@dataclass
class ExternalCheckRow:
    source_table: int
    source_row: int
    number: str
    qid: str
    vars: str
    description: str
    note: str
    condition: str
    extra_vars: str
    status: str
    reason: str


def norm(value: str) -> str:
    return unicodedata.normalize("NFKC", value or "").strip()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_pipe(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s*(?:\||,)\s*", norm(value)) if part.strip()]


def normalize_var_token(token: str) -> str:
    token = norm(token)
    token = token.replace("~", "-")
    return token


def load_var_names(workbook_path: Path) -> set[str]:
    return set(load_var_specs(workbook_path))


def load_var_specs(workbook_path: Path) -> dict[str, VarSpec]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["all"]
    specs: dict[str, VarSpec] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not cell_text(row[0]):
            continue
        try:
            width = int(cell_text(row[2]) or "2")
        except ValueError:
            width = 2
        name = cell_text(row[0])
        specs[name] = VarSpec(name, cell_text(row[1]), width)
    return specs


def range_vars(start_var: str, end_var: str, var_names: set[str]) -> list[str]:
    start_var = normalize_var_token(start_var)
    end_var = normalize_var_token(end_var)
    match1 = re.match(r"^([A-Za-z_]+)(\d+)$", start_var)
    match2 = re.match(r"^([A-Za-z_]+)(\d+)$", end_var)
    if not match1 or not match2 or match1.group(1) != match2.group(1):
        return [start_var, end_var]
    prefix = match1.group(1)
    start = int(match1.group(2))
    end = int(match2.group(2))
    pad_width = max(len(match1.group(2)), len(match2.group(2))) if match1.group(2).startswith("0") or match2.group(2).startswith("0") else 0
    step = 1 if start <= end else -1
    result = []
    for value in range(start, end + step, step):
        var = f"{prefix}{value:0{pad_width}d}" if pad_width else f"{prefix}{value}"
        if var in var_names:
            result.append(var)
    return result or [start_var, end_var]


def expand_var_text(value: str, var_names: set[str]) -> list[str]:
    result: list[str] = []
    for part in split_pipe(value.replace("～", "~")):
        part = normalize_var_token(part)
        if "~" in part or "-" in part:
            sep = "~" if "~" in part else "-"
            left, right = [p.strip() for p in part.split(sep, 1)]
            expanded = range_vars(left, right, var_names)
        else:
            expanded = [part]
        for var in expanded:
            if var and var not in result:
                result.append(var)
    return result


def comma_join(values: list[str]) -> str:
    return ",".join(value for value in values if value)


def any_expr(var: str, values: list[int | str]) -> str:
    return f"any({var},{','.join(str(v) for v in values)})"


def not_special(var: str, values: list[int]) -> str:
    return f"not({any_expr(var, values)})"


def hhmm_minutes(var: str) -> str:
    return f"(trunc({var}/100)*60 + mod({var},100))"


def parse_threshold_hours(desc: str, vars_list: list[str]) -> tuple[str, str]:
    match = re.search(r"超過\s*(\d+)\s*小時", desc)
    if not match:
        return "", ""
    hours = int(match.group(1))
    if len(vars_list) == 1:
        var = vars_list[0]
        if re.search(r"g0?1$", var):
            specials = [997, 998]
            return f"{not_special(var, specials)} & {var}>{hours}", "single hour-part threshold"
        specials = [9797, 9898]
        return f"{not_special(var, specials)} & {var}>{hours * 100}", "single HHMM threshold"
    minutes_exprs = [hhmm_minutes(var) for var in vars_list]
    guards = [not_special(var, [9797, 9898]) for var in vars_list]
    return f"{' & '.join(guards)} & ({' + '.join(minutes_exprs)})>{hours * 60}", "summed HHMM threshold"


def parse_greater_than_people(desc: str, vars_list: list[str]) -> tuple[str, str]:
    match = re.search(r"超過\s*(\d+)\s*人", desc)
    if not match or not vars_list:
        return "", ""
    threshold = int(match.group(1))
    checks = [f"({not_special(var, [991, 997, 998])} & {var}>{threshold})" for var in vars_list]
    return " | ".join(checks), "people-count threshold"


def parse_age_education(desc: str) -> tuple[str, str]:
    if "出生年在57年次以後" in desc and "vA9" in desc:
        return "vA1>=57 & any(vA3,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,97,98) & any(vA9,1,2)", "education by birth year"
    if "出生年超過民國92年次" in desc and "vA9" in desc:
        return "vA1>92 & any(vA9,7,8)", "education too high by birth year"
    if "出生年超過民國74年次" in desc and "vO1" in desc:
        return "vA1>74 & vO1=9", "retirement by birth year"
    return "", ""


def parse_gender_work(desc: str) -> tuple[str, str]:
    if "女性" in desc and "vO1" in desc and ("義務役" in desc or "替代役" in desc):
        return "sel_gen=2 & any(vO1,12,13)", "female military-service option"
    return "", ""


def parse_h_contact_unknown(desc: str) -> tuple[str, str]:
    if "vH1" in desc and "vH4" in desc and "vH7" in desc and "997" in desc:
        return "range(vH1,0,990) & range(vH4,0,990) & vH7=997", "known H1/H4 but H7 unknown"
    return "", ""


def parse_k_tiktok(desc: str, var_names: set[str]) -> tuple[str, str]:
    if "K1有選(11)抖音" not in desc:
        return "", ""
    pieces = ["vK1m11=1"]
    for var in ("vKTT4m97", "vKTT4m98"):
        if var in var_names:
            pieces.append(f"{var}~=1")
    pieces.extend(["vKTT4m13~=1", "vKTT4m14~=1"])
    return " & ".join(pieces), "multiple-response missing related TikTok actions"


def parse_q_internet_all_never(desc: str) -> tuple[str, str]:
    if "Q5答有透過網路接觸到政治消息" in desc and "Q6-Q9皆回答(01)" in desc:
        return "vQ5=1 & vQ6=1 & vQ7=1 & vQ8=1 & vQ9=1", "internet political news contradiction"
    if "Q10答有透過網路轉貼或分享政治消息" in desc and "Q11-Q14皆回答(01)" in desc:
        return "vQ10=1 & vQ11=1 & vQ12=1 & vQ13=1 & vQ14=1", "internet political sharing contradiction"
    return "", ""


def parse_time_sum_exceeds_reference(desc: str) -> tuple[str, str]:
    match = re.search(
        r"(v[A-Za-z0-9_]+)、(v[A-Za-z0-9_]+).*?時間相加大於(v[A-Za-z0-9_]+)的時間(\d+)小時",
        desc,
    )
    if not match:
        return "", ""
    left, right, ref, hours = match.groups()
    delta = int(hours) * 100
    return f"({left}+{right})>({ref}+{delta})", "HHMM pair sum exceeds reference by hours"


def parse_score_option_confirmation(desc: str) -> tuple[str, str]:
    match = re.search(
        r"(Q\d+|P\d+_\d+)答(\d+)分者,在(Q\d+|P\d+_\d+)答\((\d+)\)",
        desc,
    )
    if match:
        score_qid, score, option_qid, option = match.groups()
        return f"v{score_qid}={int(score)} & v{option_qid}={int(option)}", "score and option confirmation"
    match = re.search(
        r"(Q\d+|P\d+_\d+)答\((\d+)\).*?,(Q\d+|P\d+_\d+)答(\d+)分者",
        desc,
    )
    if match:
        option_qid, option, score_qid, score = match.groups()
        return f"v{option_qid}={int(option)} & v{score_qid}={int(score)}", "option and score confirmation"
    return "", ""


def is_hhmm_time_var(var_name: str, specs: dict[str, VarSpec]) -> bool:
    spec = specs.get(var_name)
    return bool(spec and spec.is_numeric and spec.width == 5)


def aggregate_name(vars_list: list[str], compact_range: bool = False) -> str:
    labels = [var[1:] if var.startswith("v") else var for var in vars_list]
    if not labels:
        return ""
    if compact_range and len(labels) > 2:
        return f"sum{labels[0]}_{labels[-1]}_min"
    return f"sum{'_'.join(labels)}_min"


def next_external_start(workbook_path: Path) -> int:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    used: set[int] = set()
    logic_max = 0
    for sheet_name in wb.sheetnames:
        if sheet_name == SHEET_NAME:
            continue
        ws = wb[sheet_name]
        try:
            h = {cell_text(value): idx for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1) if cell_text(value)}
        except StopIteration:
            continue
        if "m" not in h:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            value = cell_text(row[h["m"] - 1] if len(row) >= h["m"] else "")
            if value.startswith("="):
                continue
            try:
                m_id = int(value)
            except ValueError:
                continue
            used.add(m_id)
            if sheet_name.startswith("邏輯組"):
                logic_max = max(logic_max, m_id)
    start = ((logic_max // 100) + 1) * 100 + 1 if logic_max else 101
    while start in used:
        start += 100
    return start


def infer_extra_vars(note: str, vars_text: str, specs: dict[str, VarSpec]) -> str:
    note = norm(note)
    var_names = set(specs)
    extras: list[str] = []
    if "E5" in note and "E12" in note:
        all_e_vars = [var for var in range_vars("vE5", "vE12", var_names) if is_hhmm_time_var(var, specs)]
        check_vars = [var for var in expand_var_text(vars_text, var_names) if is_hhmm_time_var(var, specs)]
        aggregates = [aggregate_name(all_e_vars, compact_range=True), aggregate_name(check_vars)]
        extras.extend([var for var in aggregates if var])
        extras.extend(all_e_vars)
        return comma_join(list(dict.fromkeys(extras)))
    for start, end in re.findall(r"(v?[A-Za-z]+\d+)\s*[~～-]\s*(v?[A-Za-z]+\d+)", note):
        start_var = start if start.startswith("v") else f"v{start}"
        end_var = end if end.startswith("v") else f"v{end}"
        extras.extend(range_vars(start_var, end_var, var_names))
    result: list[str] = []
    for var in extras:
        if var in var_names and var not in result:
            result.append(var)
    return comma_join(result)


def propose_condition(desc: str, vars_text: str, note: str, var_names: set[str]) -> tuple[str, str, str]:
    desc = norm(desc)
    note = norm(note)
    if "刪除" in note:
        return "", "exclude", "deleted item"
    vars_list = expand_var_text(vars_text, var_names)
    parsers = [
        lambda: parse_gender_work(desc),
        lambda: parse_age_education(desc),
        lambda: parse_h_contact_unknown(desc),
        lambda: parse_k_tiktok(desc, var_names),
        lambda: parse_q_internet_all_never(desc),
        lambda: parse_time_sum_exceeds_reference(desc),
        lambda: parse_score_option_confirmation(desc),
        lambda: parse_threshold_hours(desc, vars_list),
        lambda: parse_greater_than_people(desc, vars_list),
    ]
    for parser in parsers:
        condition, reason = parser()
        if condition:
            return condition, "auto", reason
    if "不可重複" in desc or "Email" in desc or "手機號碼" in desc or "市話號碼" in desc:
        return "", "review", "string or duplicate validation needs dedicated rule"
    if "收入" in desc:
        return "", "review", "income category mapping needed"
    return "", "review", "no stable automatic conversion rule"


def extract_rows(docx_path: Path, workbook_path: Path) -> list[ExternalCheckRow]:
    doc = Document(docx_path)
    specs = load_var_specs(workbook_path)
    var_names = set(specs)
    rows: list[ExternalCheckRow] = []
    matched_table_count = 0
    for table_index, table in enumerate(doc.tables):
        headers = [norm(cell.text) for cell in table.rows[0].cells]
        if not SOURCE_HEADERS.issubset(set(headers)):
            continue
        group_code = chr(ord("A") + matched_table_count)
        matched_table_count += 1
        group_row_count = 0
        header_index = {name: idx for idx, name in enumerate(headers)}
        for row_index, row in enumerate(table.rows[1:], start=1):
            values = [norm(cell.text).replace("\n", " | ") for cell in row.cells]
            note = values[header_index["備註"]]
            if "刪除" in note:
                continue
            group_row_count += 1
            number = f"【{group_code}{group_row_count:02d}】"
            qid = values[header_index["題號"]]
            vars_text = values[header_index["變項名稱"]]
            desc = values[header_index["檢核說明"]]
            condition, status, reason = propose_condition(desc, vars_text, note, var_names)
            extra_vars = infer_extra_vars(note, vars_text, specs)
            qid = comma_join(split_pipe(qid))
            vars_text = comma_join(expand_var_text(vars_text, var_names))
            rows.append(
                ExternalCheckRow(
                    source_table=table_index,
                    source_row=row_index,
                    number=number,
                    qid=qid,
                    vars=vars_text,
                    description=desc,
                    note=note,
                    condition=condition,
                    extra_vars=extra_vars,
                    status=status,
                    reason=reason,
                )
            )
    return rows


def ensure_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET_NAME)
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(1, col_idx).value = header
    ws.freeze_panes = "A2"
    widths = [10, 10, 10, 10, 14, 22, 36, 80, 44, 90, 44]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    return ws


def write_workbook(workbook_path: Path, rows: list[ExternalCheckRow], *, fill_s: bool = True) -> Path:
    backup = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_external_checks.xlsx"
    backup.parent.mkdir(parents=True, exist_ok=True)
    if not backup.exists():
        shutil.copy2(workbook_path, backup)
    wb = openpyxl.load_workbook(workbook_path)
    ws = ensure_sheet(wb)
    start_id = next_external_start(workbook_path)
    for row_idx, row in enumerate(rows, start=2):
        item_id = str(start_id + row_idx - 2)
        values = [item_id, item_id, item_id if fill_s else "", "", row.number, row.qid, row.vars, row.description, row.note, row.condition, row.extra_vars]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx).value = value
    wb.save(workbook_path)
    return backup


def write_csv(path: Path, rows: list[ExternalCheckRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = list(asdict(rows[0]).keys()) if rows else ["source_table"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--no-fill-s", action="store_true", help="Leave s blank even when the 檢核項目清單 sheet has an s column.")
    args = parser.parse_args()

    rows = extract_rows(args.docx, args.workbook)
    write_csv(args.output_csv, rows)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.status] = counts.get(row.status, 0) + 1
    report: dict[str, Any] = {
        "rows": len(rows),
        "counts": counts,
        "output_csv": str(args.output_csv),
        "sheet": SHEET_NAME,
        "headers": HEADERS,
    }
    if args.apply:
        report["backup"] = str(write_workbook(args.workbook, rows, fill_s=not args.no_fill_s))
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
