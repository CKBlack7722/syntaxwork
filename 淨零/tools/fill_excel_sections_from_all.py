from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_STARTS = {
    "numeric": 101,
    "open": 1000,
    "multi": 1101,
}


@dataclass(frozen=True)
class AllVar:
    row: int
    name: str
    kind: str
    width: int


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, cell in enumerate(ws[1], start=1):
        text = cell_text(cell.value)
        if text:
            result[text] = index
    return result


def sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        if name.startswith(prefix) and name not in {"複選題內_互斥", "複選題變項清單"}:
            return wb[name]
    raise ValueError(f"missing sheet with prefix {prefix!r}")


def data_rows(ws: openpyxl.worksheet.worksheet.Worksheet):
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(cell_text(value) for value in values):
            yield row_idx, values


def read_existing_by_key(ws: openpyxl.worksheet.worksheet.Worksheet, key_header: str) -> dict[str, dict[str, Any]]:
    headers = read_headers(ws)
    if key_header not in headers:
        return {}
    existing: dict[str, dict[str, Any]] = {}
    for _row_idx, values in data_rows(ws):
        key = cell_text(values[headers[key_header] - 1])
        if not key:
            continue
        existing[key] = {
            name: values[col_idx - 1] if col_idx - 1 < len(values) else None
            for name, col_idx in headers.items()
        }
    return existing


def read_all_vars(wb: openpyxl.Workbook) -> tuple[list[AllVar], int]:
    ws = wb["all"]
    headers = read_headers(ws)
    for required in ("變項名稱", "變項屬性", "寬度"):
        if required not in headers:
            raise ValueError(f"all sheet missing {required}")
    last_row = None
    for row_idx in range(2, ws.max_row + 1):
        name = cell_text(ws.cell(row_idx, headers["變項名稱"]).value)
        if name.lower() == "last":
            last_row = row_idx
            break
    if last_row is None:
        raise ValueError("cannot find variable named 'last' in all sheet")

    variables: list[AllVar] = []
    for row_idx in range(last_row + 1, ws.max_row + 1):
        name = cell_text(ws.cell(row_idx, headers["變項名稱"]).value)
        if not name:
            continue
        kind = cell_text(ws.cell(row_idx, headers["變項屬性"]).value)
        try:
            width = int(cell_text(ws.cell(row_idx, headers["寬度"]).value) or "2")
        except ValueError:
            width = 2
        variables.append(AllVar(row_idx, name, kind, width))
    return variables, last_row


def is_numeric(var: AllVar) -> bool:
    return var.kind == "數值"


def is_text(var: AllVar) -> bool:
    return var.kind in {"字串", "文字"}


def multi_match(var_name: str) -> re.Match[str] | None:
    return re.match(r"^(?P<group>v.+?)m(?P<option>\d+)$", var_name, flags=re.IGNORECASE)


def display_qid(group: str) -> str:
    return group[1:] if group.startswith("v") and len(group) > 1 else group


def option_number(var_name: str) -> int:
    match = multi_match(var_name)
    return int(match.group("option")) if match else 0


def option_width(var_name: str) -> int:
    match = multi_match(var_name)
    return len(match.group("option")) if match else 0


def clear_data(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def first_existing_m(ws: openpyxl.worksheet.worksheet.Worksheet, fallback: int) -> int:
    headers = read_headers(ws)
    if "m" not in headers:
        return fallback
    for row_idx, values in data_rows(ws):
        text = cell_text(values[headers["m"] - 1])
        if text.isdigit():
            return int(text)
    return fallback


def infer_open_qid(var_name: str) -> str:
    if var_name.lower().endswith("_oth"):
        return re.sub(r"_oth$", "", var_name, flags=re.IGNORECASE)
    return re.sub(r"o[0-9A-Za-z]+$", "", var_name, flags=re.IGNORECASE)


def infer_open_option(var_name: str, qid: str) -> str:
    if var_name.lower().endswith("_oth"):
        return ""
    match = re.search(r"o([0-9A-Za-z]+)$", var_name[len(qid) :], flags=re.IGNORECASE)
    if not match:
        return ""
    value = match.group(1)
    return str(int(value)) if value.isdigit() else value


def infer_open_spss_fields(var_name: str, is_multi: bool) -> dict[str, str]:
    qid = infer_open_qid(var_name)
    option = infer_open_option(var_name, qid)
    if var_name.lower().endswith("city_oth"):
        return {
            "題號": qid,
            "題號變項名稱": qid,
            "選項數值": "29",
            "var2_new": qid,
            "range_new": "29",
            "n": "3",
        }
    if is_multi and option:
        parent = f"{qid}m{option}"
        return {
            "題號": qid,
            "題號變項名稱": parent,
            "選項數值": option,
            "var2_new": parent,
            "range_new": "1",
            "n": "2" if len(option) == 1 else str(len(option)),
        }
    return {
        "題號": qid,
        "題號變項名稱": option,
        "選項數值": option,
        "var2_new": qid if option else "",
        "range_new": option,
        "n": "2" if len(option) == 1 else (str(len(option)) if option else ""),
    }


def write_row(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, values: dict[str, Any]) -> None:
    headers = read_headers(ws)
    for header, col_idx in headers.items():
        if header in values:
            ws.cell(row_idx, col_idx).value = values[header]


def build_multi_groups(variables: list[AllVar]) -> dict[str, list[AllVar]]:
    groups: dict[str, list[AllVar]] = {}
    for var in variables:
        if not is_numeric(var):
            continue
        match = multi_match(var.name)
        if not match:
            continue
        groups.setdefault(match.group("group"), []).append(var)
    for group_vars in groups.values():
        group_vars.sort(key=lambda item: item.row)
    return groups


def fill_numeric_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    variables: list[AllVar],
    existing: dict[str, dict[str, Any]],
) -> int:
    start_m = first_existing_m(ws, DEFAULT_STARTS["numeric"])
    clear_data(ws)
    numeric_vars = [var for var in variables if is_numeric(var)]
    for offset, var in enumerate(numeric_vars):
        m = start_m + offset
        old = existing.get(var.name, {})
        row_values = dict(old)
        row_values.update(
            {
                "m": m,
                "p": f"=A{offset + 2}",
                "var": var.name,
                "變項屬性": "數值",
                "寬度": var.width,
                "in_or_not": old.get("in_or_not") or "range",
            }
        )
        write_row(ws, offset + 2, row_values)
    return len(numeric_vars)


def fill_open_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    variables: list[AllVar],
    existing: dict[str, dict[str, Any]],
    multi_groups: dict[str, list[AllVar]],
) -> tuple[int, list[dict[str, Any]]]:
    start_m = first_existing_m(ws, DEFAULT_STARTS["open"])
    clear_data(ws)
    open_vars = [var for var in variables if is_text(var)]
    warnings: list[dict[str, Any]] = []
    for offset, var in enumerate(open_vars):
        m = start_m + offset
        old = existing.get(var.name, {})
        qid = infer_open_qid(var.name)
        is_multi = qid in multi_groups
        inferred = infer_open_spss_fields(var.name, is_multi)
        row_values = dict(old)
        row_values.update(
            {
                "m": m,
                "p": f"=A{offset + 2}",
                "var": var.name,
                "變項屬性": var.kind,
                "寬度": var.width,
                "是否複選": 1 if is_multi else 0,
                **inferred,
            }
        )
        if not inferred.get("var2_new") or not inferred.get("range_new"):
            warnings.append(
                {
                    "var": var.name,
                    "reason": "cannot infer complete parent/range fields for SPSS open-field check",
                    "inferred": inferred,
                }
            )
        write_row(ws, offset + 2, row_values)
    return len(open_vars), warnings


def fill_multi_var_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, groups: dict[str, list[AllVar]]) -> int:
    clear_data(ws)
    row_idx = 2
    for group, vars_ in groups.items():
        for var in vars_:
            write_row(ws, row_idx, {"變項名稱": var.name, "分組": group})
            row_idx += 1
    return row_idx - 2


def preserve_main_multi_defaults(old: dict[str, Any], vars_: list[AllVar]) -> dict[str, Any]:
    options = [option_number(var.name) for var in vars_ if option_number(var.name)]
    start = min(options) if options else 1
    end = max(options) if options else len(vars_)
    width = max([option_width(var.name) for var in vars_] or [2])
    inferred_no_response = ""
    for candidate in (90, 96, 99):
        if candidate in options:
            inferred_no_response = str(candidate)
            break
    row_values = dict(old)
    row_values.update(
        {
            "寬度": width,
            "選項編號_起": start,
            "選項編號_迄": end,
            "選項數量": len(vars_),
            "迴圈進行次數": max(len(vars_) - 1, 0),
        }
    )
    if not cell_text(row_values.get("不可選擇數量")):
        row_values["不可選擇數量"] = 1 if inferred_no_response else 0
    if inferred_no_response and not cell_text(row_values.get("全部無反應選項")):
        row_values["全部無反應選項"] = inferred_no_response
    if inferred_no_response and not cell_text(row_values.get("無反應選項1")):
        row_values["無反應選項1"] = inferred_no_response
    return row_values


def fill_multi_main_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    groups: dict[str, list[AllVar]],
    existing: dict[str, dict[str, Any]],
) -> int:
    start_m = first_existing_m(ws, DEFAULT_STARTS["multi"])
    clear_data(ws)
    for offset, (group, vars_) in enumerate(groups.items()):
        m = start_m + offset
        qid = display_qid(group)
        old = existing.get(qid) or existing.get(group) or {}
        row_values = preserve_main_multi_defaults(old, vars_)
        row_values.update(
            {
                "m": m,
                "p": f"=A{offset + 2}",
                "題號": qid,
            }
        )
        write_row(ws, offset + 2, row_values)
    return len(groups)


def fill_sections(workbook_path: Path, apply: bool) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path)
    variables, last_row = read_all_vars(wb)
    numeric_ws = sheet_by_prefix(wb, "數值題")
    open_ws = sheet_by_prefix(wb, "開放欄位")
    multi_ws = sheet_by_prefix(wb, "複選題")
    multi_var_ws = wb["複選題變項清單"]

    numeric_existing = read_existing_by_key(numeric_ws, "var")
    open_existing = read_existing_by_key(open_ws, "var")
    multi_existing = read_existing_by_key(multi_ws, "題號")
    multi_groups = build_multi_groups(variables)

    report: dict[str, Any] = {
        "workbook": str(workbook_path),
        "last_row": last_row,
        "all_vars_after_last": len(variables),
        "numeric_vars": sum(1 for var in variables if is_numeric(var)),
        "text_vars": sum(1 for var in variables if is_text(var)),
        "multi_groups": len(multi_groups),
        "multi_vars": sum(len(vars_) for vars_ in multi_groups.values()),
        "applied": apply,
        "warnings": [],
    }

    if apply:
        backup_path = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_fill_from_all.xlsx"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(workbook_path, backup_path)
        report["backup"] = str(backup_path)

        report["numeric_rows_written"] = fill_numeric_sheet(numeric_ws, variables, numeric_existing)
        open_count, open_warnings = fill_open_sheet(open_ws, variables, open_existing, multi_groups)
        report["open_rows_written"] = open_count
        report["warnings"].extend(open_warnings)
        report["multi_var_rows_written"] = fill_multi_var_sheet(multi_var_ws, multi_groups)
        report["multi_main_rows_written"] = fill_multi_main_sheet(multi_ws, multi_groups, multi_existing)
        wb.save(workbook_path)
    else:
        report["numeric_rows_to_write"] = report["numeric_vars"]
        report["open_rows_to_write"] = report["text_vars"]
        report["multi_var_rows_to_write"] = report["multi_vars"]
        report["multi_main_rows_to_write"] = report["multi_groups"]
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill numeric/open/multi Excel sections from all sheet after variable 'last'.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    report = fill_sections(args.workbook, args.apply)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: v for k, v in report.items() if k != "warnings"}, ensure_ascii=False))
    if report.get("warnings"):
        print(f"warnings: {len(report['warnings'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
