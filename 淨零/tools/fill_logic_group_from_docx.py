from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl

from analyze_logic_from_docx import extract_cues


LOGIC_HEADERS = ["m", "p", "條件", "應答", "不應答", "限制"]


@dataclass
class Proposal:
    status: str
    row_source: str
    qid: str
    condition_text: str
    condition: str
    required_vars: str
    forbidden_vars: str
    limit: str
    reason: str
    raw_text: str


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize(value: str) -> str:
    return unicodedata.normalize("NFKC", value or "").strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {text(value): idx for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1) if text(value)}


def ensure_logic_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    current = headers(ws)
    aliases = {
        "條件": ("條件", "汚ン"),
        "應答": ("應答", "棲紀"),
        "不應答": ("不應答", "ぃ棲紀"),
        "限制": ("限制",),
    }
    for wanted in ("m", "p"):
        if wanted not in current:
            raise ValueError(f"邏輯組 missing required header {wanted}")
    for canonical, names in aliases.items():
        found = next((name for name in names if name in current), "")
        if found and found != canonical:
            ws.cell(1, current[found]).value = canonical
            current[canonical] = current.pop(found)
        elif not found:
            col = ws.max_column + 1
            ws.cell(1, col).value = canonical
            current[canonical] = col
    return headers(ws)


def load_vars(workbook_path: Path) -> set[str]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["all"]
    return {text(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row and text(row[0])}


def normalize_qid(qid: str, var_names: set[str]) -> str:
    qid = normalize(qid).upper()
    match = re.match(r"^([A-Z]+)0+([1-9][0-9A-Z_]*)$", qid)
    if match and f"v{match.group(1)}{match.group(2)}" in var_names:
        return f"{match.group(1)}{match.group(2)}"
    return qid


def option_var(qid: str, code: str, var_names: set[str]) -> str:
    qid = normalize_qid(qid, var_names)
    bare = str(int(float(code))) if re.fullmatch(r"\d+(?:\.0)?", code) else code
    for padded in (bare.zfill(2), bare):
        candidate = f"v{qid}m{padded}"
        if candidate in var_names:
            return candidate
    return ""


def target_vars(qid: str, var_names: set[str]) -> list[str]:
    qid = normalize_qid(qid, var_names)
    base = f"v{qid}"
    if base in var_names:
        return [base]
    prefixes = (f"{base}m", f"{base}s", f"{base}city", f"{base}town")
    result = []
    for name in sorted(var_names):
        if name.startswith(prefixes) and not re.search(r"o\d+$", name, flags=re.IGNORECASE) and not name.endswith("_oth"):
            result.append(name)
    return result


def expand_codes(fragment: str) -> list[str]:
    fragment = normalize(fragment).replace("~", "-")
    values: list[str] = []
    for start, end in re.findall(r"\((\d+)\)\s*-\s*\((\d+)\)", fragment):
        width = max(len(start), len(end))
        for value in range(int(start), int(end) + 1):
            values.append(str(value).zfill(width))
    cleaned = re.sub(r"\(\d+\)\s*-\s*\(\d+\)", "", fragment)
    values.extend(re.findall(r"\((\d+)\)", cleaned))
    if not values:
        match = re.search(r"答\s*([0-9]+(?:\.[0-9]+)?)", fragment)
        if match:
            values.append(match.group(1))
    return values


def expr_for_ref(qid: str, fragment: str, var_names: set[str]) -> str:
    qid = normalize_qid(qid, var_names)
    range_match = re.search(r"答\s*([0-9]+(?:\.[0-9]+)?)\s*[-~]\s*([0-9]+(?:\.[0-9]+)?)", fragment)
    if range_match and not re.search(r"\([0-9]+\)", fragment):
        var = f"v{qid}"
        if var in var_names:
            return f"({var}>={range_match.group(1)} & {var}<={range_match.group(2)})"
    codes = expand_codes(fragment)
    if not codes:
        return ""
    multi_exprs = []
    for code in codes:
        var = option_var(qid, code, var_names)
        if var:
            multi_exprs.append(f"{var} in 1")
    if multi_exprs:
        return " | ".join(multi_exprs)
    var = f"v{qid}"
    if var in var_names:
        compact = ",".join(str(int(code)) if code.isdigit() else code for code in codes)
        return f"{var} in {compact}"
    return ""


def normalize_shared_answer_subjects(part: str) -> str:
    part = normalize(part)
    while True:
        updated = re.sub(r"([A-Z][A-Z0-9_]*)答\((\d+)\)\s*或\s*\((\d+)\)", r"\1答(\2)、(\3)", part)
        if updated == part:
            break
        part = updated
    match = re.match(r"^([A-Z][A-Z0-9_]*)(?:或([A-Z][A-Z0-9_]*))(?:或([A-Z][A-Z0-9_]*))?答(.+)$", part)
    if match:
        q1, q2, q3, suffix = match.groups()
        subjects = [q1, q2] + ([q3] if q3 else [])
        return "或".join(f"{subject}答{suffix}" for subject in subjects)
    return part


def parse_condition(condition_text: str, var_names: set[str]) -> tuple[str, str]:
    clean = normalize(condition_text)
    clean = re.split(r"才需回答此題|需回答此題|需答此題|不需回答此題|顯示此題|選項|只能|開放填答", clean)[0]
    clean = clean.replace("資料處理:", "")
    clean = clean.replace("有答", "答")
    clean = clean.replace("者", "")
    clean = clean.strip(" ,，;；")
    and_parts = [part for part in re.split(r"且", clean) if part.strip()]
    and_exprs: list[str] = []
    for part in and_parts:
        part = normalize_shared_answer_subjects(part)
        or_parts = [p for p in re.split(r"或", part) if p.strip()]
        or_exprs: list[str] = []
        for atom in or_parts:
            atom = normalize(atom)
            atom = atom.strip(" ,，;；")
            match = re.search(r"([A-Z][A-Z0-9_]*)\s*答(.+)", atom)
            if not match:
                return "", f"cannot parse condition atom: {atom}"
            expr = expr_for_ref(match.group(1), match.group(0), var_names)
            if not expr:
                return "", f"cannot map condition atom: {atom}"
            or_exprs.append(expr)
        if len(or_exprs) == 1:
            and_exprs.append(or_exprs[0])
        else:
            and_exprs.append("(" + " | ".join(or_exprs) + ")")
    if not and_exprs:
        return "", "blank condition"
    return " & ".join(and_exprs), ""


def proposal_for_cue(cue: Any, var_names: set[str]) -> Proposal:
    qid = normalize_qid(cue.qid, var_names)
    raw_condition = normalize(cue.condition_text)
    source = f"{cue.source}:{cue.index}"
    unsupported_tokens = ("互斥", "必選", "加總", "超過", "顯示", "跳至", "開放填答")
    targets = target_vars(qid, var_names)
    condition, error = parse_condition(raw_condition, var_names)
    if any(token in raw_condition for token in unsupported_tokens):
        return Proposal("review", source, qid, raw_condition, condition, "", "", "", "restriction/check/display cue needs rule-specific handling", cue.raw_text)
    if error:
        return Proposal("review", source, qid, raw_condition, "", "", "", "", error, cue.raw_text)
    if not targets:
        return Proposal("review", source, qid, raw_condition, condition, "", "", "", f"target qid {qid} not found in all sheet", cue.raw_text)
    target_text = ",".join(targets)
    if "不需回答此題" in raw_condition:
        return Proposal("apply", source, qid, raw_condition, condition, "", target_text, "", "", cue.raw_text)
    if "需回答此題" in raw_condition or "才需回答此題" in raw_condition or "需答此題" in raw_condition:
        return Proposal("apply", source, qid, raw_condition, condition, target_text, "", "", "", cue.raw_text)
    return Proposal("review", source, qid, raw_condition, condition, "", "", "", "not an answer-required/not-required cue", cue.raw_text)


def write_proposals(path: Path, proposals: list[Proposal]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(proposals[0]).keys()) if proposals else ["status"])
        writer.writeheader()
        for proposal in proposals:
            writer.writerow(asdict(proposal))


def apply_to_workbook(workbook_path: Path, proposals: list[Proposal], backup_path: Path) -> dict[str, Any]:
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    if not backup_path.exists():
        shutil.copy2(workbook_path, backup_path)
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["邏輯組"]
    h = ensure_logic_headers(ws)
    apply_rows = [p for p in proposals if p.status == "apply"]
    row_iter = (row for row in range(2, ws.max_row + 1) if not any(text(ws.cell(row, h[col]).value) for col in ("條件", "應答", "不應答", "限制")))
    written = 0
    for proposal in apply_rows:
        try:
            row = next(row_iter)
        except StopIteration:
            break
        ws.cell(row, h["條件"]).value = proposal.condition
        ws.cell(row, h["應答"]).value = proposal.required_vars
        ws.cell(row, h["不應答"]).value = proposal.forbidden_vars
        ws.cell(row, h["限制"]).value = proposal.limit
        written += 1
    wb.save(workbook_path)
    return {"apply_proposals": len(apply_rows), "written": written, "backup": str(backup_path)}


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill stable Word-derived logic rules into the Excel 邏輯組 sheet.")
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--proposal-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    var_names = load_vars(args.workbook)
    cues = extract_cues(args.docx)
    proposals = [proposal_for_cue(cue, var_names) for cue in cues if cue.status != "ignore"]
    write_proposals(args.proposal_csv, proposals)
    counts: dict[str, int] = {}
    for proposal in proposals:
        counts[proposal.status] = counts.get(proposal.status, 0) + 1
    result: dict[str, Any] = {"counts": counts, "proposal_csv": str(args.proposal_csv)}
    if args.apply:
        result["apply"] = apply_to_workbook(
            args.workbook,
            proposals,
            args.workbook.parent / "generated" / f"{args.workbook.stem}.before_logic_fill.xlsx",
        )
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
