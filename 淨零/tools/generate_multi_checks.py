from __future__ import annotations

import argparse
import difflib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


MULTI_START = "**3.複選題檢核."
MUTEX_START = "**3.1複選互斥邏輯."
NEXT_SECTION_PREFIX = "**4."


@dataclass(frozen=True)
class MultiQuestion:
    row: int
    m: str
    p: str
    raw_question: str
    display_name: str
    group_key: str
    vars: tuple[str, ...]
    no_response_codes: tuple[str, ...]
    check_at_least_one: bool
    check_code_consistency: bool


@dataclass(frozen=True)
class MutexRule:
    row: int
    m: str
    p: str
    raw_question: str
    display_name: str
    group_key: str
    exclusive_code: str
    exclusive_var: str
    other_vars: tuple[str, ...]
    recovered_from_main: bool = False


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def id_text(value: Any) -> str:
    text = cell_text(value)
    if text.endswith(".0"):
        return text[:-2]
    return text


def question_display(value: Any) -> str:
    text = id_text(value)
    if not text:
        return ""
    if text.isdigit():
        return f"v{text}"
    return text


def group_candidates(display_name: str) -> tuple[str, ...]:
    candidates = [display_name]
    if display_name.startswith("v") and len(display_name) > 1:
        candidates.append(display_name[1:])
    else:
        candidates.append(f"v{display_name}")
    return tuple(dict.fromkeys(candidates))


def infer_group_from_var(var_name: str) -> str:
    match = re.match(r"^(v.+?)m[0-9]+$", cell_text(var_name), flags=re.IGNORECASE)
    return match.group(1) if match else ""


def normalize_id(value: Any) -> str:
    text = id_text(value)
    return text


def get_sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    matches = [
        name
        for name in wb.sheetnames
        if name.startswith(prefix) and name not in {"複選題內_互斥", "複選題變項清單"}
    ]
    if not matches:
        raise ValueError(f"Workbook has no sheet starting with {prefix!r}")
    return wb[matches[0]]


def read_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if value not in (None, ""):
            headers[str(value).strip()] = idx
    return headers


def read_var_widths(wb: openpyxl.Workbook) -> dict[str, int]:
    widths: dict[str, int] = {}
    if "all" not in wb.sheetnames:
        return widths
    ws = wb["all"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = cell_text(row[0])
        try:
            widths[name] = int(row[2])
        except (TypeError, ValueError):
            widths[name] = 2
    return widths


def read_multi_groups(wb: openpyxl.Workbook) -> dict[str, list[str]]:
    ws = wb["複選題變項清單"]
    groups: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        var_name = cell_text(row[0])
        group = cell_text(row[1]) if len(row) > 1 else ""
        inferred = infer_group_from_var(var_name)
        if not group or group.startswith("=") or group.lower() == "v":
            group = inferred
        if group:
            groups.setdefault(group, []).append(var_name)
    return groups


def find_group(display_name: str, groups: dict[str, list[str]]) -> tuple[str, tuple[str, ...]] | None:
    for candidate in group_candidates(display_name):
        if candidate in groups:
            return candidate, tuple(groups[candidate])
    return None


def option_code(var_name: str) -> str | None:
    match = re.search(r"m0*([0-9]+)$", var_name, flags=re.IGNORECASE)
    if not match:
        return None
    return str(int(match.group(1)))


def option_map(vars_for_question: tuple[str, ...]) -> dict[str, str]:
    result: dict[str, str] = {}
    for var_name in vars_for_question:
        code = option_code(var_name)
        if code is not None:
            result[code] = var_name
    return result


def parse_codes(*values: Any) -> tuple[str, ...]:
    codes: list[str] = []
    for value in values:
        text = cell_text(value)
        if not text:
            continue
        for part in re.split(r"[,;、\s]+", text):
            part = part.strip()
            if part and part not in codes:
                codes.append(part)
    return tuple(codes)


def flag_from_cell(value: Any, default: bool = True) -> bool:
    text = cell_text(value).lower()
    if text in {"0", "n", "no", "false", "否", "不檢查"}:
        return False
    if text in {"1", "y", "yes", "true", "是", "檢查"}:
        return True
    return default


def load_questions(wb: openpyxl.Workbook, issues: list[dict[str, Any]]) -> list[MultiQuestion]:
    ws = get_sheet_by_prefix(wb, "複選題")
    headers = read_headers(ws)
    groups = read_multi_groups(wb)
    questions: list[MultiQuestion] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[headers["m"]]:
            continue
        display = question_display(row[headers["題號"]])
        if not display:
            continue
        found = find_group(display, groups)
        if not found:
            issues.append(
                {
                    "level": "error",
                    "sheet": ws.title,
                    "row": row_idx,
                    "message": f"題號 {cell_text(row[headers['題號']])} 找不到對應的複選題變項清單分組",
                }
            )
            continue
        group_key, vars_for_question = found
        no_response = parse_codes(
            row[headers.get("全部無反應選項", -1)] if "全部無反應選項" in headers else "",
            *(row[headers[f"無反應選項{i}"]] for i in range(1, 6) if f"無反應選項{i}" in headers),
        )
        if not no_response:
            no_response = ("96",)

        m_id = normalize_id(row[headers["m"]])
        p_id = normalize_id(row[headers.get("p", headers["m"])])
        if not p_id or p_id.startswith("="):
            p_id = m_id
        questions.append(
            MultiQuestion(
                row=row_idx,
                m=m_id,
                p=p_id,
                raw_question=cell_text(row[headers["題號"]]),
                display_name=display,
                group_key=group_key,
                vars=vars_for_question,
                no_response_codes=no_response,
                check_at_least_one=flag_from_cell(row[headers.get("是否檢查至少一項", -1)] if "是否檢查至少一項" in headers else "", True),
                check_code_consistency=flag_from_cell(row[headers.get("是否檢查96一致", -1)] if "是否檢查96一致" in headers else "", True),
            )
        )

        formula = cell_text(row[headers.get("加總程式", -1)] if "加總程式" in headers else "")
        if formula:
            first, last = vars_for_question[0], vars_for_question[-1]
            if first not in formula or last not in formula:
                issues.append(
                    {
                        "level": "warning",
                        "sheet": ws.title,
                        "row": row_idx,
                        "message": f"加總程式未使用實際變項清單範圍 {first} to {last}；產生器已忽略此公式欄",
                    }
                )
    return questions


def question_by_display(questions: list[MultiQuestion]) -> dict[str, MultiQuestion]:
    result: dict[str, MultiQuestion] = {}
    for question in questions:
        for candidate in group_candidates(question.display_name):
            result[candidate] = question
        result[question.raw_question] = question
    return result


def load_mutex_rules(wb: openpyxl.Workbook, questions: list[MultiQuestion], issues: list[dict[str, Any]]) -> list[MutexRule]:
    if "複選題內_互斥" not in wb.sheetnames:
        return []
    ws = wb["複選題內_互斥"]
    headers = read_headers(ws)
    by_display = question_by_display(questions)
    main_mutex_code = {q.display_name: "" for q in questions}
    multi_ws = get_sheet_by_prefix(wb, "複選題")
    multi_headers = read_headers(multi_ws)
    for row in multi_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[multi_headers["m"]]:
            continue
        display = question_display(row[multi_headers["題號"]])
        main_mutex_code[display] = cell_text(row[multi_headers.get("互斥選項", -1)] if "互斥選項" in multi_headers else "")

    rules: list[MutexRule] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[headers["m"]]:
            continue
        display = question_display(row[headers["題號"]])
        if not display:
            continue
        question = by_display.get(display)
        if question is None:
            issues.append(
                {
                    "level": "error",
                    "sheet": ws.title,
                    "row": row_idx,
                    "message": f"互斥題號 {cell_text(row[headers['題號']])} 找不到對應複選題分組；此列未產生",
                }
            )
            continue
        opt_map = option_map(question.vars)
        exclusive_code = cell_text(row[headers["互斥選項"]])
        recovered = False
        if exclusive_code not in opt_map and main_mutex_code.get(question.display_name) in opt_map:
            original = exclusive_code
            exclusive_code = main_mutex_code[question.display_name]
            recovered = True
            issues.append(
                {
                    "level": "warning",
                    "sheet": ws.title,
                    "row": row_idx,
                    "message": f"互斥選項 {original} 找不到變項；已用複選題主表的互斥選項 {exclusive_code} 產生",
                }
            )
        if exclusive_code not in opt_map:
            issues.append(
                {
                    "level": "error",
                    "sheet": ws.title,
                    "row": row_idx,
                    "message": f"互斥選項 {exclusive_code} 找不到對應變項；此列未產生",
                }
            )
            continue

        start = cell_text(row[headers.get("互斥選項編號_起", -1)] if "互斥選項編號_起" in headers else "")
        end = cell_text(row[headers.get("互斥選項編號_迄", -1)] if "互斥選項編號_迄" in headers else "")
        other_codes: list[str] = []
        if start.isdigit() and end.isdigit():
            for code in range(int(start), int(end) + 1):
                text_code = str(code)
                if text_code != exclusive_code and text_code in opt_map:
                    other_codes.append(text_code)
        if not other_codes:
            other_codes = [code for code in opt_map if code != exclusive_code]
        other_vars = tuple(opt_map[code] for code in other_codes)
        if not other_vars:
            issues.append(
                {
                    "level": "error",
                    "sheet": ws.title,
                    "row": row_idx,
                    "message": f"互斥選項 {exclusive_code} 沒有可比對的其他選項；此列未產生",
                }
            )
            continue

        m_id = normalize_id(row[headers["m"]])
        p_id = normalize_id(row[headers.get("p", headers["m"])])
        if not p_id or p_id.startswith("="):
            p_id = m_id
        rules.append(
            MutexRule(
                row=row_idx,
                m=m_id,
                p=p_id,
                raw_question=cell_text(row[headers["題號"]]),
                display_name=question.display_name,
                group_key=question.group_key,
                exclusive_code=exclusive_code,
                exclusive_var=opt_map[exclusive_code],
                other_vars=other_vars,
                recovered_from_main=recovered,
            )
        )
    return rules


def spss_range(vars_for_question: tuple[str, ...]) -> str:
    return f"{vars_for_question[0]} to {vars_for_question[-1]}"


def string_format(var_name: str, widths: dict[str, int]) -> str:
    width = widths.get(var_name, 2)
    return f"string({var_name},f{width})"


def render_concat(display_name: str, vars_for_question: tuple[str, ...], widths: dict[str, int]) -> str:
    parts: list[str] = []
    for index, var_name in enumerate(vars_for_question):
        if index > 0:
            parts.append('" , "')
        parts.append(f'"{var_name}="')
        parts.append(string_format(var_name, widths))
    return f"COMPUTE {display_name} = Rtrim(Ltrim(concat({','.join(parts)})))."


def render_multi_section(questions: list[MultiQuestion], mutex_rules: list[MutexRule], widths: dict[str, int]) -> str:
    lines: list[str] = [MULTI_START]
    if questions:
        lines.append(f"* {' '.join(q.display_name for q in questions)} 組合.")
        lines.append(f"STRING {' '.join(q.display_name for q in questions)} (A600).")
        for question in questions:
            lines.append(render_concat(question.display_name, question.vars, widths))
        lines.append("")

    for question in questions:
        first, last = question.vars[0], question.vars[-1]
        range_text = spss_range(question.vars)
        zero_name = re.sub(r"^v", "", question.display_name)
        code_list = ",".join(question.no_response_codes)
        conditions: list[str] = []
        if question.check_code_consistency:
            conditions.extend(
                [
                    f"(any(a(#i),0,1) & any(a(#i+1),{code_list}))",
                    f"(any(a(#i),{code_list}) and not any(a(#i+1),{code_list}))",
                ]
            )
        if question.check_at_least_one:
            conditions.append(f"#{zero_name}zero=1")

        lines.extend(
            [
                f"*{question.display_name}.",
                f"vector a={range_text}.",
                f"COMPUTE #{zero_name}zero = (SUM({first} TO {last}) = 0).",
                f"loop #i=1 to {len(question.vars) - 1}.",
                "do if " + "\n| ".join(conditions) + ".",
                f"compute m{question.m}=Rtrim(Ltrim({question.display_name})).",
                f'compute p{question.p}="{question.display_name}至少選1項或選特殊碼應一致".',
                "end if.",
                "end loop.",
                "exec.",
                "",
            ]
        )

    lines.append(MUTEX_START)
    for rule in mutex_rules:
        other_range = spss_range(rule.other_vars) if len(rule.other_vars) > 1 else rule.other_vars[0]
        lines.extend(
            [
                f"*{rule.display_name}={rule.exclusive_code} 複選題內互斥.",
                f"do if any({rule.exclusive_var},1) & any(1,{other_range}).",
                f"compute m{rule.m}=Rtrim(Ltrim({rule.display_name})).",
                f'compute p{rule.p}="{rule.display_name}({rule.exclusive_code})複選題選項應互斥".',
                "end if.",
                "Exec.",
                "",
            ]
        )
    return "\n".join(lines).rstrip() + "\n"


def extract_reference_multi_block(path: Path) -> str:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    start = text.find(MULTI_START)
    if start == -1:
        return ""
    end = text.find(NEXT_SECTION_PREFIX, start)
    if end == -1:
        end = len(text)
    return text[start:end].strip() + "\n"


def normalize_for_compare(text: str) -> list[str]:
    result: list[str] = []
    for line in text.splitlines():
        stripped = re.sub(r"\s+", " ", line.strip())
        if stripped:
            result.append(stripped.lower())
    return result


def compare_blocks(generated: str, reference: str) -> dict[str, Any]:
    generated_lines = normalize_for_compare(generated)
    reference_lines = normalize_for_compare(reference)
    diff = list(difflib.unified_diff(reference_lines, generated_lines, fromfile="reference", tofile="generated", lineterm=""))
    return {
        "reference_lines": len(reference_lines),
        "generated_lines": len(generated_lines),
        "normalized_equal": generated_lines == reference_lines,
        "diff": diff[:300],
        "diff_truncated": len(diff) > 300,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate SPSS syntax for expanded multiple-response checks.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--reference-syntax", type=Path)
    parser.add_argument("--compare-output", type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    issues: list[dict[str, Any]] = []
    widths = read_var_widths(wb)
    questions = load_questions(wb, issues)
    mutex_rules = load_mutex_rules(wb, questions, issues)
    generated = render_multi_section(questions, mutex_rules, widths)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(generated, encoding="utf-8-sig")

    report: dict[str, Any] = {
        "workbook": str(args.workbook),
        "output": str(args.output),
        "questions": [
            {
                "row": q.row,
                "m": q.m,
                "p": q.p,
                "display_name": q.display_name,
                "group_key": q.group_key,
                "vars": list(q.vars),
                "no_response_codes": list(q.no_response_codes),
            }
            for q in questions
        ],
        "mutex_rules": [
            {
                "row": r.row,
                "m": r.m,
                "p": r.p,
                "display_name": r.display_name,
                "exclusive_code": r.exclusive_code,
                "exclusive_var": r.exclusive_var,
                "other_vars": list(r.other_vars),
                "recovered_from_main": r.recovered_from_main,
            }
            for r in mutex_rules
        ],
        "issues": issues,
    }
    if args.reference_syntax:
        reference = extract_reference_multi_block(args.reference_syntax)
        report["compare"] = compare_blocks(generated, reference) if reference else {"error": "reference block not found"}

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.compare_output and "compare" in report:
        args.compare_output.parent.mkdir(parents=True, exist_ok=True)
        args.compare_output.write_text(json.dumps(report["compare"], ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"multi questions: {len(questions)}")
    print(f"mutex rules: {len(mutex_rules)}")
    print(f"issues: {len(issues)}")
    if "compare" in report:
        compare = report["compare"]
        print(f"compare normalized_equal: {compare.get('normalized_equal')}")
        print(f"compare reference_lines: {compare.get('reference_lines')}")
        print(f"compare generated_lines: {compare.get('generated_lines')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
