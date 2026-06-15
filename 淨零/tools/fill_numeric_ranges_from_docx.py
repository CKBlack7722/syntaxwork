from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl
from docx import Document


SYSTEM_VARS = {
    "ct",
    "ic",
    "r",
    "rt",
    "sendtime",
    "smst",
    "smstime1",
    "termination",
    "wendtime",
    "wlast",
    "wst1",
    "wst2",
    "wst3",
    "wst4",
    "wstarttime",
}


@dataclass
class NumericProposal:
    row: int
    var: str
    qid: str
    status: str
    reason: str
    existing_r1: str
    existing_r2: str
    existing_r3: str
    existing_r4: str
    proposed_r1: str
    proposed_r2: str
    proposed_r3: str
    proposed_r4: str
    source: str
    question_excerpt: str


def norm(value: str) -> str:
    return unicodedata.normalize("NFKC", value or "").strip()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {cell_text(v): i for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) if cell_text(v)}


def normalize_date(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    parts = re.findall(r"\d+", value)
    if len(parts) == 3:
        return f"{int(parts[0]):04d}{int(parts[1]):02d}{int(parts[2]):02d}000000"
    if len(digits) == 8:
        return digits + "000000"
    if len(digits) == 14:
        return digits
    raise ValueError(f"Invalid date: {value}")


def normalize_number(value: str) -> str:
    value = value.strip()
    if re.fullmatch(r"0+\d+", value):
        return str(int(value))
    return value


def qid_from_var(var: str) -> str:
    qid = var[1:] if var.lower().startswith("v") else var
    qid = re.sub(r"o\d+$", "", qid, flags=re.I)
    qid = re.sub(r"m\d+$", "", qid, flags=re.I)
    qid = re.sub(r"g\d+$", "", qid)
    qid = re.sub(r"(city|town|_oth)$", "", qid, flags=re.I)
    return qid.upper()


def qid_candidates(var: str) -> list[str]:
    qid = qid_from_var(var)
    candidates = [qid]
    matrix = re.match(r"^([A-Z]+[0-9]+)S[A-Z0-9_]+$", qid)
    if matrix:
        candidates.append(matrix.group(1))
    if qid.startswith("CK"):
        candidates.append(qid[2:])
    return list(dict.fromkeys(candidates))


def doc_blocks(docx_path: Path) -> dict[str, str]:
    doc = Document(docx_path)
    paras = [norm(p.text) for p in doc.paragraphs if norm(p.text)]
    starts: list[tuple[int, str]] = []
    for idx, para in enumerate(paras):
        compact = para.replace(" ", "")
        match = re.search(r"(?:^|】)([A-Z]{1,5}[0-9][A-Z0-9_]*)[.‧．、 ]?", compact)
        if match:
            starts.append((idx, match.group(1).upper()))
    blocks: dict[str, str] = {}
    for i, (start, qid) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else len(paras)
        blocks.setdefault(qid, "\n".join(paras[start:end]))
    return blocks


def find_block(var: str, blocks: dict[str, str]) -> tuple[str, str]:
    for qid in qid_candidates(var):
        if qid in blocks:
            return qid, blocks[qid]
    return "", ""


def read_special_ranges(wb: openpyxl.Workbook) -> dict[str, str]:
    if "特殊變項範圍" not in wb.sheetnames:
        return {}
    ws = wb["特殊變項範圍"]
    result: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and row[1]:
            result[cell_text(row[0])] = cell_text(row[1])
    return result


def read_multi_vars(wb: openpyxl.Workbook) -> set[str]:
    if "複選題變項清單" not in wb.sheetnames:
        return set()
    ws = wb["複選題變項清單"]
    return {cell_text(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row and cell_text(row[0])}


def range_from_explicit_text(text: str) -> tuple[list[str], str]:
    text = norm(text)
    match = re.search(r"範圍(?:為|:)?([^】\n]+)", text)
    if not match:
        return [], ""
    raw = match.group(1)
    raw = re.split(r"【|\]|;|；", raw)[0]
    raw = raw.replace("...", "、")
    values: list[str] = []
    for start, end in re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*[-~]\s*([0-9]+(?:\.[0-9]+)?)", raw):
        values.append(f"{normalize_number(start)},{normalize_number(end)}")
    consumed = re.sub(r"([0-9]+(?:\.[0-9]+)?)\s*[-~]\s*([0-9]+(?:\.[0-9]+)?)", "", raw)
    listed = [normalize_number(number) for number in re.findall(r"\d+(?:\.\d+)?", consumed)]
    if not values and len(listed) > 4:
        normal_numbers = [float(number) for number in listed if float(number) < 90]
        special_numbers = [number for number in listed if float(number) >= 90]
        if normal_numbers:
            low = min(normal_numbers)
            high = max(normal_numbers)
            low_text = str(int(low)) if low.is_integer() else str(low)
            high_text = str(int(high)) if high.is_integer() else str(high)
            values.append(f"{low_text},{high_text}" if low_text != high_text else low_text)
        values.extend(special_numbers)
    else:
        for number in listed:
            if number not in values:
                values.append(number)
    for code in option_codes(text):
        normalized = normalize_number(code)
        if float(normalized) >= 90 and normalized not in values:
            values.append(normalized)
    return values[:4], "explicit numeric range"


def option_codes(block: str) -> list[str]:
    codes = []
    for code in re.findall(r"[□\(（]\s*(\d{1,6})\s*[）)]", block):
        if code not in codes:
            codes.append(code)
    return codes


def split_option_range(codes: list[str]) -> list[str]:
    numeric = [int(c) for c in codes]
    normal = [n for n in numeric if n < 90]
    specials = [n for n in numeric if n >= 90]
    result: list[str] = []
    if normal:
        result.append(f"{min(normal)},{max(normal)}" if min(normal) != max(normal) else str(min(normal)))
    result.extend(str(n) for n in specials)
    return result[:4]


def range_sort_key(value: str) -> tuple[float, float]:
    parts = [part.strip() for part in value.split(",") if part.strip()]
    try:
        nums = [float(part) for part in parts]
    except ValueError:
        return (float("inf"), float("inf"))
    if not nums:
        return (float("inf"), float("inf"))
    return (min(nums), max(nums))


def normalize_range_cells(values: list[str]) -> list[str]:
    cleaned: list[str] = []
    for value in values:
        if value and value not in cleaned:
            cleaned.append(value)
    cleaned.sort(key=range_sort_key)
    return cleaned[:4]


def infer_values(var: str, width: int, block: str, special_ranges: dict[str, str], multi_vars: set[str], date_range: str) -> tuple[list[str], str, str]:
    lower = var.lower()
    if lower == "va3":
        return ["1,21", "97,98"], "force_apply", "project rule: vA3 normal range plus continuous special codes"
    if lower in SYSTEM_VARS:
        return [], "manual", "system variable"
    if width == 14:
        return [date_range], "apply", "configured date range"
    if "city" in lower:
        return ["1,29"], "apply", "city code rule"
    if lower.endswith("town") and "鄉鎮市區" in special_ranges:
        return [special_ranges["鄉鎮市區"]], "apply", "town code list"
    if var in multi_vars or re.search(r"m\d+$", var, flags=re.I):
        values = ["0,1"]
        if not re.search(r"m0\d+$", var, flags=re.I):
            values.append("96")
        return values, "apply", "multiple-response binary"
    if re.search(r"g1$", var, flags=re.I):
        qid = qid_from_var(var)
        if "小時範圍" in block or qid in {"M1", "Q24", "Q21", "Q26A"}:
            return ["0,140", "997", "998"], "apply", "hour-minute hour part"
    if re.search(r"g2$", var, flags=re.I):
        qid = qid_from_var(var)
        if "分鐘範圍" in block or qid in {"M1", "Q24", "Q21", "Q26A"}:
            return ["0,59", "97", "98"], "apply", "hour-minute minute part"
    values, source = range_from_explicit_text(block)
    if values:
        return values, "apply", source
    codes = option_codes(block)
    if codes:
        return normalize_range_cells(split_option_range(codes)), "apply", "option code span"
    return [], "manual", "no numeric range found"


def compare(existing: list[str], proposed: list[str], status: str) -> tuple[str, str]:
    if status == "force_apply":
        return "force_apply", "confirmed project rule overrides existing r1-r4"
    if status == "manual":
        return "manual", "no automatic proposal"
    padded = (proposed + ["", "", "", ""])[:4]
    if existing == padded:
        return "match", ""
    if existing[0] == padded[0] and existing[1:] != padded[1:]:
        return "range_match_special_diff", "r1 matches; r2-r4 differ"
    if any(existing):
        return "mismatch", "existing r1-r4 differ from proposal"
    return "apply", ""


def build_proposals(workbook_path: Path, docx_path: Path, date_start: str, date_end: str) -> list[NumericProposal]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["數值題"]
    h = headers(ws)
    blocks = doc_blocks(docx_path)
    special_ranges = read_special_ranges(wb)
    multi_vars = read_multi_vars(wb)
    date_range = f"{normalize_date(date_start)},{normalize_date(date_end)}"
    proposals: list[NumericProposal] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        var = cell_text(row[h["var"]])
        if not var:
            continue
        width = int(cell_text(row[h["寬度"]]) or "2")
        existing = [cell_text(row[h[col]]) for col in ("r1", "r2", "r3", "r4")]
        qid, block = find_block(var, blocks)
        values, base_status, source = infer_values(var, width, block, special_ranges, multi_vars, date_range)
        values = normalize_range_cells(values)
        status, reason = compare(existing, values, base_status)
        padded = (values + ["", "", "", ""])[:4]
        proposals.append(
            NumericProposal(
                row=row_idx,
                var=var,
                qid=qid,
                status=status,
                reason=reason,
                existing_r1=existing[0],
                existing_r2=existing[1],
                existing_r3=existing[2],
                existing_r4=existing[3],
                proposed_r1=padded[0],
                proposed_r2=padded[1],
                proposed_r3=padded[2],
                proposed_r4=padded[3],
                source=source,
                question_excerpt=re.sub(r"\s+", " ", block[:300]),
            )
        )
    return proposals


def write_csv(path: Path, proposals: list[NumericProposal]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(proposals[0]).keys()) if proposals else ["row"])
        writer.writeheader()
        for proposal in proposals:
            writer.writerow(asdict(proposal))


def apply(workbook_path: Path, proposals: list[NumericProposal]) -> dict[str, Any]:
    backup = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_numeric_fill.xlsx"
    backup.parent.mkdir(parents=True, exist_ok=True)
    if not backup.exists():
        shutil.copy2(workbook_path, backup)
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["數值題"]
    h = headers(ws)
    written = 0
    for proposal in proposals:
        if proposal.status not in {"apply", "match", "range_match_special_diff", "force_apply"}:
            continue
        values = [proposal.proposed_r1, proposal.proposed_r2, proposal.proposed_r3, proposal.proposed_r4]
        if not values[0]:
            continue
        for col, value in zip(("r1", "r2", "r3", "r4"), values):
            ws.cell(proposal.row, h[col] + 1).value = value or None
        written += 1
    wb.save(workbook_path)
    return {"written": written, "backup": str(backup)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--date-start", required=True)
    parser.add_argument("--date-end", required=True)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    proposals = build_proposals(args.workbook, args.docx, args.date_start, args.date_end)
    write_csv(args.output_csv, proposals)
    counts: dict[str, int] = {}
    sources: dict[str, int] = {}
    for proposal in proposals:
        counts[proposal.status] = counts.get(proposal.status, 0) + 1
        sources[proposal.source] = sources.get(proposal.source, 0) + 1
    report: dict[str, Any] = {"counts": counts, "sources": sources, "output_csv": str(args.output_csv)}
    if args.apply:
        report["apply"] = apply(args.workbook, proposals)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
