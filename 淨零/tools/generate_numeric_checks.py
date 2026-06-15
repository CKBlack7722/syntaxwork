from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class NumericRule:
    row: int
    m: str
    p: str
    var: str
    width: int
    in_or_not: str
    ranges: list[str]
    decimal_rule: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return wb[name]
    raise ValueError(f"missing sheet prefix {prefix}")


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(value): idx
        for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if cell_text(value)
    }


def resolved_mp(value: Any, row_values: dict[str, Any]) -> str:
    text = cell_text(value)
    if text.startswith("="):
        return cell_text(row_values.get("m"))
    return text


def load_rules(workbook_path: Path) -> tuple[list[NumericRule], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    ws = sheet_by_prefix(wb, "數值題")
    h = headers(ws)
    required = ["m", "p", "var", "變項屬性", "寬度", "in_or_not", "r1", "r2", "r3", "r4"]
    missing = [name for name in required if name not in h]
    if missing:
        raise ValueError(f"numeric sheet missing headers: {missing}")
    rules: list[NumericRule] = []
    skipped: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = {name: row[idx - 1] for name, idx in h.items()}
        var = cell_text(values.get("var"))
        if not var:
            continue
        if cell_text(values.get("變項屬性")) != "數值":
            continue
        ranges = [cell_text(values.get(col)) for col in ("r1", "r2", "r3", "r4") if cell_text(values.get(col))]
        if not ranges:
            skipped.append({"row": row_idx, "var": var, "reason": "blank r1-r4"})
            continue
        m = cell_text(values.get("m"))
        p = resolved_mp(values.get("p"), values)
        if not m or not p:
            skipped.append({"row": row_idx, "var": var, "reason": "blank m or p"})
            continue
        try:
            width = int(cell_text(values.get("寬度")) or "2")
        except ValueError:
            width = 2
        rules.append(
            NumericRule(
                row=row_idx,
                m=m,
                p=p,
                var=var,
                width=width,
                in_or_not=cell_text(values.get("in_or_not")) or "range",
                ranges=ranges,
                decimal_rule=cell_text(values.get("可輸入小數點")),
            )
        )
    return rules, skipped


def expand_range_args(values: list[str]) -> list[str]:
    args: list[str] = []
    for value in values:
        if "," in value:
            left, right = [part.strip() for part in value.split(",", 1)]
            args.extend([left, right])
        else:
            args.extend([value, value])
    return args


def valid_expr(rule: NumericRule) -> str:
    args = expand_range_args(rule.ranges)
    fn = "any" if rule.in_or_not.lower() == "any" else "range"
    expr = f"{fn}({rule.var},{','.join(args)})"
    if rule.decimal_rule:
        expr = f"({expr}) & ({rule.decimal_rule})"
    return expr


def render_rule(rule: NumericRule) -> str:
    range_comment = " ".join(rule.ranges)
    return "\n".join(
        [
            f"*{rule.var}={range_comment} .",
            f"do if not {valid_expr(rule)} | sys({rule.var}).",
            f'compute m{rule.m}=concat("{rule.var}=",string({rule.var},f{rule.width})).',
            f'compute p{rule.p}="{rule.var}為不合理值或遺漏值".',
            "end if.",
            "Exec.",
            "",
        ]
    )


def render_spss(rules: list[NumericRule]) -> str:
    parts = [
        "* Encoding: UTF-8.",
        "**NUMERIC CHECKS.",
        "* SYNTAXWORK_BEGIN_NUMERIC.",
    ]
    parts.extend(render_rule(rule).rstrip() for rule in rules)
    parts.append("* SYNTAXWORK_END_NUMERIC.")
    return "\n\n".join(parts).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    rules, skipped = load_rules(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(render_spss(rules), encoding="utf-8-sig")
    report = {"rules": len(rules), "skipped": len(skipped), "skipped_rows": skipped, "output": str(args.output)}
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
