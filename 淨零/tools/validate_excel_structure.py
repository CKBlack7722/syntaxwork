from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


NUMERIC_REQUIRED = ["m", "p", "var", "變項屬性", "寬度", "可輸入小數點", "in_or_not", "r1", "r2", "r3", "r4", "range"]
OPEN_REQUIRED = ["m", "p", "var", "變項屬性", "寬度", "是否複選", "題號", "題號變項名稱", "選項數值", "var2_new", "range_new", "n"]
MULTI_REQUIRED = ["m", "p", "題號", "寬度", "選項編號_起", "選項編號_迄", "選項數量", "互斥選項", "迴圈進行次數"]
MUTEX_REQUIRED = ["m", "p", "題號", "寬度", "互斥選項"]
MULTI_VAR_REQUIRED = ["變項名稱", "分組"]
LOGIC_REQUIRED = ["m", "p", "條件", "應答", "不應答"]
LOGIC_RECOMMENDED = ["限制"]
ALL_REQUIRED = ["變項名稱", "變項屬性", "寬度"]

ERROR_TOKENS = {"#VALUE!", "#REF!", "#NAME?", "#N/A", "#DIV/0!"}


@dataclass(frozen=True)
class Issue:
    project: str
    severity: str
    sheet: str
    row: str
    column: str
    code: str
    message: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if value not in (None, ""):
            result[str(value).strip()] = index
    return result


def sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> str | None:
    for name in wb.sheetnames:
        if name.startswith(prefix) and name not in {"複選題內_互斥", "複選題變項清單"}:
            return name
    return None


def infer_multi_group(var_name: str) -> str:
    match = re.match(r"^(v.+?)m[0-9]+$", cell_text(var_name), flags=re.IGNORECASE)
    return match.group(1) if match else ""


def issue(project: str, severity: str, sheet: str, row: Any, column: str, code: str, message: str) -> Issue:
    return Issue(project, severity, sheet, str(row or ""), column, code, message)


def resolved_mp_value(ws: openpyxl.worksheet.worksheet.Worksheet, h: dict[str, int], row_index: int, col: str) -> str:
    value = cell_text(ws.cell(row_index, h[col] + 1).value)
    match = re.fullmatch(r"=A(\d+)", value, flags=re.IGNORECASE)
    if match and "m" in h and int(match.group(1)) == row_index:
        return cell_text(ws.cell(row_index, h["m"] + 1).value)
    return value


def check_headers(project: str, wb: openpyxl.Workbook, sheet_name: str | None, required: list[str], recommended: list[str] | None = None) -> list[Issue]:
    if not sheet_name:
        return [issue(project, "error", "", "", "", "missing_sheet", f"Missing required sheet.")]
    ws = wb[sheet_name]
    found = headers(ws)
    issues: list[Issue] = []
    for name in required:
        if name not in found:
            issues.append(issue(project, "error", sheet_name, 1, name, "missing_header", f"Missing required header {name}."))
    for name in recommended or []:
        if name not in found:
            issues.append(issue(project, "warning", sheet_name, 1, name, "missing_recommended_header", f"Recommended header {name} is absent; current tools can continue, but related rules cannot be generated."))
    return issues


def rows(ws: openpyxl.worksheet.worksheet.Worksheet):
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and any(cell_text(value) for value in row):
            yield row_index, row


def check_error_tokens(project: str, wb: openpyxl.Workbook, sheet_names: list[str]) -> list[Issue]:
    issues: list[Issue] = []
    for sheet_name in sheet_names:
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_values = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        for row_index, row in rows(ws):
            for col_index, value in enumerate(row):
                text = cell_text(value)
                if text in ERROR_TOKENS:
                    column = cell_text(header_values[col_index]) if col_index < len(header_values) else f"col{col_index + 1}"
                    issues.append(issue(project, "error", sheet_name, row_index, column, "formula_error", f"Cell contains {text}."))
    return issues


def check_unique_vars(project: str, wb: openpyxl.Workbook, sheet_name: str, var_header: str) -> list[Issue]:
    ws = wb[sheet_name]
    h = headers(ws)
    if var_header not in h:
        return []
    seen: dict[str, int] = {}
    issues: list[Issue] = []
    for row_index, row in rows(ws):
        var = cell_text(row[h[var_header]])
        if not var:
            continue
        key = var.lower()
        if key in seen:
            issues.append(issue(project, "error", sheet_name, row_index, var_header, "duplicate_var", f"Variable {var} duplicates row {seen[key]}."))
        else:
            seen[key] = row_index
    return issues


def check_global_mp(project: str, wb: openpyxl.Workbook) -> list[Issue]:
    issues: list[Issue] = []
    seen: dict[tuple[str, str], tuple[str, int]] = {}
    target_prefixes = ("數值題", "開放欄位", "複選題")
    target_names = {"複選題內_互斥", "邏輯組"}
    for sheet_name in wb.sheetnames:
        if sheet_name == "複選題變項清單":
            continue
        if not (sheet_name.startswith(target_prefixes) or sheet_name in target_names):
            continue
        ws = wb[sheet_name]
        h = headers(ws)
        for col in ("m", "p"):
            if col not in h:
                continue
            for row_index, _row in rows(ws):
                value = resolved_mp_value(ws, h, row_index, col)
                if not value:
                    continue
                key = (col, value)
                if key in seen:
                    first_sheet, first_row = seen[key]
                    issues.append(
                        issue(
                            project,
                            "error",
                            sheet_name,
                            row_index,
                            col,
                            "duplicate_global_mp",
                            f"{col} {value} duplicates {first_sheet} row {first_row}.",
                        )
                    )
                else:
                    seen[key] = (sheet_name, row_index)
    return issues


def check_mp(project: str, wb: openpyxl.Workbook, sheet_name: str, required_columns: list[str]) -> list[Issue]:
    ws = wb[sheet_name]
    h = headers(ws)
    issues: list[Issue] = []
    seen_m: dict[str, int] = {}
    for row_index, row in rows(ws):
        for col in required_columns:
            if col in h and not cell_text(row[h[col]]):
                issues.append(issue(project, "error", sheet_name, row_index, col, "blank_required_cell", f"{col} is required."))
        m = resolved_mp_value(ws, h, row_index, "m") if "m" in h else ""
        if m:
            if not m.isdigit():
                issues.append(issue(project, "error", sheet_name, row_index, "m", "non_numeric_m", f"m should be numeric, got {m}."))
            elif m in seen_m:
                issues.append(issue(project, "error", sheet_name, row_index, "m", "duplicate_m", f"m {m} duplicates row {seen_m[m]}."))
            else:
                seen_m[m] = row_index
    return issues


def check_numeric(project: str, wb: openpyxl.Workbook, sheet_name: str) -> list[Issue]:
    ws = wb[sheet_name]
    h = headers(ws)
    issues = check_unique_vars(project, wb, sheet_name, "var")
    issues.extend(check_mp(project, wb, sheet_name, ["m", "p", "var", "變項屬性", "寬度"]))
    for row_index, row in rows(ws):
        var = cell_text(row[h["var"]])
        kind = cell_text(row[h["變項屬性"]])
        width = cell_text(row[h["寬度"]])
        mode = cell_text(row[h["in_or_not"]]).lower() if "in_or_not" in h else ""
        range_expr = cell_text(row[h["range"]]) if "range" in h else ""
        r_values = [cell_text(row[h[col]]) for col in ("r1", "r2", "r3", "r4") if col in h and cell_text(row[h[col]])]
        if kind != "數值":
            issues.append(issue(project, "error", sheet_name, row_index, "變項屬性", "invalid_numeric_kind", f"{var} should be 數值."))
        if not width.isdigit():
            issues.append(issue(project, "error", sheet_name, row_index, "寬度", "invalid_width", f"{var} width should be numeric."))
        if r_values and mode not in {"any", "range"}:
            issues.append(issue(project, "error", sheet_name, row_index, "in_or_not", "invalid_in_or_not", f"{var} in_or_not should be any or range when r1-r4 has values."))
        if r_values and not range_expr:
            issues.append(issue(project, "error", sheet_name, row_index, "range", "blank_range_expression", f"{var} has r1-r4 values but blank range expression."))
        if range_expr and not range_expr.startswith("=") and not re.search(r"\b(any|range)\s*\(", range_expr, flags=re.IGNORECASE):
            issues.append(issue(project, "warning", sheet_name, row_index, "range", "range_expression_review", f"{var} range expression does not contain any(...) or range(...)."))
        for col in ("r1", "r2", "r3", "r4"):
            value = cell_text(row[h[col]])
            if value and not re.fullmatch(r"-?\d+(?:\.\d+)?(?:,-?\d+(?:\.\d+)?)?", value):
                issues.append(issue(project, "warning", sheet_name, row_index, col, "range_format_review", f"{var} {col} has non-standard range/code format: {value}."))
    return issues


def check_open(project: str, wb: openpyxl.Workbook, sheet_name: str) -> list[Issue]:
    ws = wb[sheet_name]
    h = headers(ws)
    issues = check_unique_vars(project, wb, sheet_name, "var")
    issues.extend(check_mp(project, wb, sheet_name, ["m", "p", "var", "變項屬性", "寬度", "是否複選"]))
    for row_index, row in rows(ws):
        var = cell_text(row[h["var"]])
        if cell_text(row[h["變項屬性"]]) != "字串":
            issues.append(issue(project, "error", sheet_name, row_index, "變項屬性", "invalid_open_kind", f"{var} should be 字串."))
        if cell_text(row[h["寬度"]]) != "150":
            issues.append(issue(project, "warning", sheet_name, row_index, "寬度", "open_width_review", f"{var} open field width is expected to be 150."))
        if cell_text(row[h["是否複選"]]) not in {"0", "1"}:
            issues.append(issue(project, "error", sheet_name, row_index, "是否複選", "invalid_multi_flag", f"{var} 是否複選 should be 0 or 1."))
    return issues


def check_multi_vars(project: str, wb: openpyxl.Workbook, sheet_name: str) -> list[Issue]:
    ws = wb[sheet_name]
    h = headers(ws)
    issues = check_unique_vars(project, wb, sheet_name, "變項名稱")
    for row_index, row in rows(ws):
        var = cell_text(row[h["變項名稱"]])
        group = cell_text(row[h["分組"]])
        if not group or group.startswith("=") or group.lower() == "v":
            group = infer_multi_group(var)
        if not group:
            issues.append(issue(project, "error", sheet_name, row_index, "分組", "blank_group", f"{var} has blank group."))
        if not re.search(r"m\d+$", var, flags=re.IGNORECASE):
            issues.append(issue(project, "warning", sheet_name, row_index, "變項名稱", "multi_var_name_review", f"{var} does not look like a multiple-response variable."))
    return issues


def check_multi(project: str, wb: openpyxl.Workbook, sheet_name: str, multi_var_sheet: str) -> list[Issue]:
    ws = wb[sheet_name]
    h = headers(ws)
    issues = check_mp(project, wb, sheet_name, ["m", "p", "題號", "寬度"])
    groups = set()
    if multi_var_sheet in wb.sheetnames:
        vh = headers(wb[multi_var_sheet])
        for _, row in rows(wb[multi_var_sheet]):
            var = cell_text(row[vh["變項名稱"]]) if "變項名稱" in vh else ""
            group = cell_text(row[vh["分組"]])
            if not group or group.startswith("=") or group.lower() == "v":
                group = infer_multi_group(var)
            if group:
                groups.add(group.lower())
    for row_index, row in rows(ws):
        qid = cell_text(row[h["題號"]])
        if qid and groups and qid.lower() not in groups and f"v{qid}".lower() not in groups:
            issues.append(issue(project, "warning", sheet_name, row_index, "題號", "multi_group_review", f"題號 {qid} is not directly found in 複選題變項清單 分組."))
    return issues


def check_logic(project: str, wb: openpyxl.Workbook, sheet_name: str) -> list[Issue]:
    ws = wb[sheet_name]
    h = headers(ws)
    issues = check_mp(project, wb, sheet_name, ["m"])
    for row_index, row in rows(ws):
        values = {name: cell_text(row[h[name]]) for name in h if h[name] < len(row)}
        if not any(values.get(name, "") for name in ("條件", "應答", "不應答", "限制")):
            issues.append(issue(project, "warning", sheet_name, row_index, "", "blank_logic_rule", "Logic row has no condition/action fields."))
    return issues


def validate(project: str, workbook_path: Path) -> tuple[list[Issue], dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    numeric_sheet = sheet_by_prefix(wb, "數值題")
    open_sheet = sheet_by_prefix(wb, "開放欄位")
    multi_sheet = sheet_by_prefix(wb, "複選題")
    mutex_sheet = "複選題內_互斥" if "複選題內_互斥" in wb.sheetnames else None
    multi_var_sheet = "複選題變項清單" if "複選題變項清單" in wb.sheetnames else None
    logic_sheet = "邏輯組" if "邏輯組" in wb.sheetnames else None
    issues: list[Issue] = []

    for sheet_name, required, recommended in [
        ("all", ALL_REQUIRED, []),
        (numeric_sheet, NUMERIC_REQUIRED, []),
        (open_sheet, OPEN_REQUIRED, []),
        (multi_sheet, MULTI_REQUIRED, []),
        (mutex_sheet, MUTEX_REQUIRED, []),
        (multi_var_sheet, MULTI_VAR_REQUIRED, []),
    ]:
        issues.extend(check_headers(project, wb, sheet_name, required, recommended))
    if logic_sheet:
        issues.extend(check_headers(project, wb, logic_sheet, LOGIC_REQUIRED, LOGIC_RECOMMENDED))

    issues.extend(check_global_mp(project, wb))
    check_sheets = [name for name in [numeric_sheet, open_sheet, multi_sheet, mutex_sheet, multi_var_sheet, logic_sheet] if name]
    issues.extend(check_error_tokens(project, wb, check_sheets))
    if numeric_sheet:
        issues.extend(check_numeric(project, wb, numeric_sheet))
    if open_sheet:
        issues.extend(check_open(project, wb, open_sheet))
    if multi_var_sheet:
        issues.extend(check_multi_vars(project, wb, multi_var_sheet))
    if multi_sheet and multi_var_sheet:
        issues.extend(check_multi(project, wb, multi_sheet, multi_var_sheet))
    if logic_sheet:
        issues.extend(check_logic(project, wb, logic_sheet))

    counts: dict[str, int] = {}
    for item in issues:
        counts[item.severity] = counts.get(item.severity, 0) + 1
    return issues, {
        "workbook": str(workbook_path),
        "sheets": {
            "numeric": numeric_sheet,
            "open": open_sheet,
            "multi": multi_sheet,
            "mutex": mutex_sheet,
            "multi_vars": multi_var_sheet,
            "logic": logic_sheet,
        },
        "issue_counts": counts,
    }


def write_csv(path: Path, issues: list[Issue]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(Issue.__dataclass_fields__))
        writer.writeheader()
        for item in issues:
            writer.writerow(item.__dict__)


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate workbook structure before survey-to-SPSS conversion.")
    parser.add_argument("--project", required=True)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    issues, summary = validate(args.project, args.workbook)
    write_csv(args.output_csv, issues)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(
        json.dumps({"summary": summary, "issues": [item.__dict__ for item in issues]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(args.project, summary["issue_counts"])
    return 1 if summary["issue_counts"].get("error", 0) else 0


if __name__ == "__main__":
    raise SystemExit(main())
