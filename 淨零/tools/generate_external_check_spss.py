from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


SHEET_NAME = "檢核項目清單"
RESERVED_WORDS = {"and", "or", "not", "any", "range", "sysmis", "sys", "mod", "trunc"}


@dataclass(frozen=True)
class VarSpec:
    name: str
    kind: str
    width: int

    @property
    def is_numeric(self) -> bool:
        return self.kind == "數值"


@dataclass(frozen=True)
class ExternalCheckRule:
    row: int
    item_number: str
    m: str
    p: str
    s: str
    s_value: str
    qid: str
    vars_text: str
    description: str
    note: str
    condition: str
    extra_vars: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(value): idx
        for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if cell_text(value)
    }


def resolved_mp_value(ws: openpyxl.worksheet.worksheet.Worksheet, h: dict[str, int], row_idx: int, col_name: str) -> str:
    value = cell_text(ws.cell(row_idx, h[col_name]).value)
    if value.startswith("="):
        return cell_text(ws.cell(row_idx, h["m"]).value)
    return value


def resolved_optional_value(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    h: dict[str, int],
    row_idx: int,
    col_name: str,
    fallback_col: str = "",
) -> str:
    if col_name not in h:
        return ""
    value = cell_text(ws.cell(row_idx, h[col_name]).value)
    if value.startswith("="):
        return cell_text(ws.cell(row_idx, h[fallback_col]).value) if fallback_col and fallback_col in h else ""
    return value


def split_vars(value: object) -> list[str]:
    if value is None:
        return []
    result: list[str] = []
    for part in re.split(r"\s*(?:\||,)\s*", str(value)):
        clean = part.strip()
        if clean and clean not in result:
            result.append(clean)
    return result


def normalize_condition(condition: str, specs: dict[str, VarSpec]) -> str:
    def replace_token(match: re.Match[str]) -> str:
        token = match.group(0)
        if token in specs or token.lower() in RESERVED_WORDS:
            return token
        prefixed = f"v{token}"
        if prefixed in specs:
            return prefixed
        return token

    return re.sub(r"\b(?!v\b)[A-Z][A-Za-z0-9_]*\b", replace_token, condition)


def load_var_specs(workbook_path: Path) -> dict[str, VarSpec]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["all"]
    specs: dict[str, VarSpec] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, kind, width = row[:3]
        if not name:
            continue
        try:
            width_int = int(width)
        except (TypeError, ValueError):
            width_int = 2
        specs[str(name).strip()] = VarSpec(str(name).strip(), str(kind).strip(), width_int)
    return specs


def spec_for(var_name: str, specs: dict[str, VarSpec]) -> VarSpec:
    if is_aggregate_var(var_name):
        return VarSpec(var_name, "數值", 8)
    return specs.get(var_name, VarSpec(var_name, "數值", 2))


def display_expr(var_name: str, specs: dict[str, VarSpec]) -> str:
    spec = spec_for(var_name, specs)
    if spec.is_numeric:
        return f"string({var_name},n{spec.width})"
    return f"rtrim(ltrim({var_name}))"


def concat_parts(vars_to_show: list[str], specs: dict[str, VarSpec]) -> list[str]:
    parts: list[str] = []
    for index, var_name in enumerate(vars_to_show):
        parts.append(f'",{var_name}="' if index else f'"{var_name}="')
        parts.append(display_expr(var_name, specs))
    return parts


def render_compute_m(m_id: str, vars_to_show: list[str], specs: dict[str, VarSpec]) -> list[str]:
    parts = concat_parts(vars_to_show, specs)
    if len(parts) <= 4:
        return [f"compute m{m_id}=concat({','.join(parts)})."]
    lines = [f"compute m{m_id}=concat("]
    for index, part in enumerate(parts):
        suffix = "," if index < len(parts) - 1 else ""
        lines.append(f"  {part}{suffix}")
    lines.append(").")
    return lines


def split_outside_quotes(text: str, separators: set[str]) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    in_quote = False
    for char in text:
        if char == '"':
            in_quote = not in_quote
        if not in_quote and char in separators:
            piece = "".join(current).strip()
            if piece:
                parts.append(piece)
            parts.append(char)
            current = []
        else:
            current.append(char)
    piece = "".join(current).strip()
    if piece:
        parts.append(piece)
    return parts


def wrap_expression(expr: str, first_prefix: str, final_suffix: str = ".", max_len: int = 160) -> list[str]:
    single = f"{first_prefix}{expr}{final_suffix}"
    if len(single) <= max_len:
        return [single]
    tokens = split_outside_quotes(expr, {"&", "|", "+"})
    lines: list[str] = []
    current = first_prefix
    pending_op = ""
    for token in tokens:
        if token in {"&", "|", "+"}:
            pending_op = token
            continue
        piece = f"{pending_op} {token}" if pending_op else token
        if current.strip() and len(current) + len(piece) + 1 > max_len:
            lines.append(current.rstrip())
            current = f"{pending_op} {token}" if pending_op else token
        else:
            current += ("" if current == first_prefix else " ") + piece
        pending_op = ""
    if current.strip():
        lines.append(current.rstrip() + final_suffix)
    return lines


def string_chunks(text: str, chunk_size: int = 100) -> list[str]:
    return [text[index : index + chunk_size] for index in range(0, len(text), chunk_size)] or [""]


def render_compute_p(p_id: str, message: str, commented: bool = False) -> list[str]:
    escaped = message.replace('"', "'")
    prefix = "* " if commented else ""
    single = f'{prefix}compute p{p_id}="{escaped}".'
    if len(single) <= 160:
        return [single]
    lines = [f"{prefix}compute p{p_id}=concat("]
    chunks = string_chunks(escaped)
    for index, chunk in enumerate(chunks):
        suffix = "," if index < len(chunks) - 1 else ""
        lines.append(f'{prefix}  "{chunk}"{suffix}')
    lines.append(f"{prefix}).")
    return lines


def condition_vars(condition: str, specs: dict[str, VarSpec]) -> list[str]:
    names = re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", normalize_condition(condition, specs))
    result: list[str] = []
    for name in names:
        if name.lower() in RESERVED_WORDS or name.isdigit():
            continue
        if name in specs and name not in result:
            result.append(name)
    return result


def message_text(rule: ExternalCheckRule) -> str:
    text = rule.description.replace('"', "'")
    return text[:230]


def unique(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        if value and value not in result:
            result.append(value)
    return result


def is_aggregate_var(var_name: str) -> bool:
    return bool(re.fullmatch(r"sum[A-Za-z0-9_]+(?:_min)?", var_name))


def range_vars(start_var: str, end_var: str, specs: dict[str, VarSpec]) -> list[str]:
    match1 = re.match(r"^([A-Za-z_]+)(\d+)$", start_var)
    match2 = re.match(r"^([A-Za-z_]+)(\d+)$", end_var)
    if not match1 or not match2 or match1.group(1) != match2.group(1):
        return [start_var, end_var]
    prefix = match1.group(1)
    start = int(match1.group(2))
    end = int(match2.group(2))
    pad_width = max(len(match1.group(2)), len(match2.group(2))) if match1.group(2).startswith("0") or match2.group(2).startswith("0") else 0
    step = 1 if start <= end else -1
    result = []
    for value in range(start, end + step, step):
        var = f"{prefix}{value:0{pad_width}d}" if pad_width else f"{prefix}{value}"
        spec = specs.get(var)
        if spec and spec.is_numeric and spec.width == 5:
            result.append(var)
    return result


def aggregate_source_vars(var_name: str, specs: dict[str, VarSpec]) -> list[str]:
    if not is_aggregate_var(var_name):
        return []
    body = var_name.removeprefix("sum").removesuffix("_min")
    parts = [part for part in body.split("_") if part]
    if len(parts) == 2:
        ranged = range_vars(f"v{parts[0]}", f"v{parts[1]}", specs)
        if len(ranged) > 2:
            return ranged
    result = []
    for part in parts:
        var = part if part.startswith("v") else f"v{part}"
        if var in specs:
            result.append(var)
    return result


def hhmm_minutes(var_name: str) -> str:
    return f"(trunc({var_name}/100)*60 + mod({var_name},100))"


def hhmm_hours(var_name: str) -> str:
    return f"((trunc({var_name}/100)*60 + mod({var_name},100))/60)"


def aggregate_value_expr(var_name: str, source_var: str) -> str:
    if var_name.endswith("_min"):
        return hhmm_minutes(source_var)
    return hhmm_hours(source_var)


def render_aggregate_computes(aggregate_vars: list[str], specs: dict[str, VarSpec]) -> list[str]:
    lines: list[str] = []
    for var_name in aggregate_vars:
        source_vars = aggregate_source_vars(var_name, specs)
        if not source_vars:
            continue
        minute_var = var_name if var_name.endswith("_min") else f"{var_name}_min"
        lines.append(f"compute {minute_var}=0.")
        for source_var in source_vars:
            lines.extend(
                wrap_expression(
                    f"(not(any({source_var},9797,9898,99996))) {minute_var}={minute_var} + {hhmm_minutes(source_var)}",
                    "if ",
                )
            )
        if var_name != minute_var:
            lines.append(f"compute {var_name}={minute_var}/60.")
        lines.append("")
    return lines


def rule_listed_vars(rule: ExternalCheckRule, specs: dict[str, VarSpec]) -> list[str]:
    condition = normalize_condition(rule.condition, specs)
    return unique(split_vars(rule.vars_text) + split_vars(rule.extra_vars) + condition_vars(condition, specs))


def collect_aggregate_vars(rules: list[ExternalCheckRule], specs: dict[str, VarSpec]) -> list[str]:
    result: list[str] = []
    for rule in rules:
        for var_name in rule_listed_vars(rule, specs):
            if is_aggregate_var(var_name) and var_name not in result:
                result.append(var_name)
    return result


def load_rules(workbook_path: Path) -> tuple[list[ExternalCheckRule], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return [], [{"severity": "error", "row": "", "reason": f"missing sheet {SHEET_NAME}"}]
    ws = wb[SHEET_NAME]
    h = headers(ws)
    required = ["m", "p", "題號", "變項名稱", "檢核說明", "備註", "條件", "額外列出變項"]
    missing = [name for name in required if name not in h]
    if missing:
        return [], [{"severity": "error", "row": "", "reason": f"missing headers: {', '.join(missing)}"}]
    rules: list[ExternalCheckRule] = []
    skipped: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        m = resolved_mp_value(ws, h, row_idx, "m")
        p = resolved_mp_value(ws, h, row_idx, "p")
        s = resolved_optional_value(ws, h, row_idx, "s", "m")
        s_value = resolved_optional_value(ws, h, row_idx, "s=")
        item_number = cell_text(ws.cell(row_idx, h["項目編號"]).value) if "項目編號" in h else ""
        qid = cell_text(ws.cell(row_idx, h["題號"]).value)
        vars_text = cell_text(ws.cell(row_idx, h["變項名稱"]).value)
        desc = cell_text(ws.cell(row_idx, h["檢核說明"]).value)
        note = cell_text(ws.cell(row_idx, h["備註"]).value)
        condition = cell_text(ws.cell(row_idx, h["條件"]).value)
        extra_vars = cell_text(ws.cell(row_idx, h["額外列出變項"]).value)
        if not any([qid, vars_text, desc, note, condition]):
            continue
        if not m or not p:
            skipped.append({"severity": "warning", "row": row_idx, "qid": qid, "reason": "blank m or p"})
            continue
        rules.append(ExternalCheckRule(row_idx, item_number, m, p, s, s_value, qid, vars_text, desc, note, condition, extra_vars))
    return rules, skipped


def render_rule(rule: ExternalCheckRule, specs: dict[str, VarSpec]) -> str:
    condition = normalize_condition(rule.condition, specs)
    listed_vars = rule_listed_vars(rule, specs)
    lines = [
        f"* external check {rule.item_number or rule.row}: {rule.qid}.",
    ]
    if not condition:
        skeleton = [
            "* TODO: blank condition in Excel; fill the condition before enabling this rule.",
            "* do if .",
        ]
        skeleton.extend(f"* {line}" for line in render_compute_m(rule.m, listed_vars, specs))
        skeleton.extend(render_compute_p(rule.p, message_text(rule), commented=True))
        skeleton.extend(
            [
                *([f"* compute s{rule.s}={rule.s_value}."] if rule.s and rule.s_value else []),
                "* end if.",
                "* exec.",
                "",
            ]
        )
        lines.extend(skeleton)
        return "\n".join(lines)
    lines.extend(wrap_expression(f"({condition})", "do if "))
    lines.extend(render_compute_m(rule.m, listed_vars, specs))
    lines.extend(render_compute_p(rule.p, message_text(rule)))
    lines.extend(
        [
            *([f"compute s{rule.s}={rule.s_value}."] if rule.s and rule.s_value else []),
            "end if.",
            "exec.",
            "",
        ]
    )
    return "\n".join(lines)


def render_spss(rules: list[ExternalCheckRule], specs: dict[str, VarSpec]) -> str:
    blocks = [
        "* Encoding: UTF-8.",
        "**五、檢核項目清單.",
        "* SYNTAXWORK_BEGIN_EXTERNAL_CHECKS.",
    ]
    aggregate_lines = render_aggregate_computes(collect_aggregate_vars(rules, specs), specs)
    if aggregate_lines:
        blocks.append("* aggregate variables.")
        blocks.append("\n".join(aggregate_lines).rstrip())
    blocks.extend(render_rule(rule, specs) for rule in rules)
    blocks.append("* SYNTAXWORK_END_EXTERNAL_CHECKS.")
    return "\n".join(blocks).rstrip() + "\n"


def duplicate_mp_report(workbook_path: Path, rules: list[ExternalCheckRule]) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    seen: dict[tuple[str, str], str] = {}
    issues: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("generated"):
            continue
        ws = wb[sheet_name]
        h = headers(ws)
        if "m" not in h and "p" not in h and "s" not in h:
            continue
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_values = {name: row[idx - 1] for name, idx in h.items() if idx - 1 < len(row)}
            for col_name in ("m", "p", "s"):
                if col_name not in h:
                    continue
                value = cell_text(row_values.get(col_name))
                if value.startswith("="):
                    value = cell_text(row_values.get("m"))
                if not value:
                    continue
                key = (col_name, value)
                where = f"{sheet_name}!row{row_idx}"
                if key in seen:
                    issues.append({"severity": "error", "code": "duplicate_mp", "value": f"{col_name}{value}", "first": seen[key], "second": where})
                else:
                    seen[key] = where
    return issues


def pending_condition_rows(rules: list[ExternalCheckRule]) -> list[dict[str, Any]]:
    return [
        {"severity": "info", "row": rule.row, "qid": rule.qid, "reason": "blank condition; emitted commented SPSS skeleton"}
        for rule in rules
        if not rule.condition
    ]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    specs = load_var_specs(args.workbook)
    rules, skipped = load_rules(args.workbook)
    pending_conditions = pending_condition_rows(rules)
    duplicate_issues = duplicate_mp_report(args.workbook, rules)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(render_spss(rules, specs), encoding="utf-8-sig")
    report = {
        "rules": len(rules),
        "skipped": len(skipped),
        "pending_conditions": len(pending_conditions),
        "duplicate_mp_issues": len(duplicate_issues),
        "skipped_rows": skipped,
        "pending_condition_rows": pending_conditions,
        "duplicate_mp": duplicate_issues,
        "output": str(args.output),
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
