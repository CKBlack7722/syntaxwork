from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from docx import Document


SYSTEM_VAR_NAMES = {
    "ct",
    "ic",
    "r",
    "rt",
    "sendtime",
    "smst",
    "smstime1",
    "termination",
    "wendtime",
    "wlast",
    "wst1",
    "wst2",
    "wst3",
    "wst4",
    "wstarttime",
}


@dataclass(frozen=True)
class Proposal:
    var: str
    qid: str
    existing: tuple[str, str, str, str]
    proposed: tuple[str, str, str, str]
    confidence: str
    source: str
    note: str
    compare_status: str
    mismatch_reason: str
    matched_question: str
    question_excerpt: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(text: str) -> str:
    return unicodedata.normalize("NFKC", text or "").replace("＿", "_")


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def find_numeric_sheet(wb: openpyxl.Workbook) -> str:
    matches = [name for name in wb.sheetnames if name.startswith("數值題")]
    if not matches:
        raise ValueError("No numeric sheet found")
    return matches[0]


def read_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if value not in (None, ""):
            headers[str(value).strip()] = idx
    return headers


def make_question_blocks(docx_path: Path) -> dict[str, str]:
    doc = Document(str(docx_path))
    paras = [normalize_text(p.text.strip()) for p in doc.paragraphs if p.text.strip()]
    starts: list[tuple[int, str]] = []
    for idx, paragraph in enumerate(paras):
        compact = paragraph.replace(" ", "")
        match = re.match(r"^([A-Z]{1,4}[0-9]+(?:_[0-9]+)?[A-Z]?)", compact)
        if not match:
            match = re.match(r"^([0-9]{1,3})(?![\.)、])", compact)
        if match:
            starts.append((idx, match.group(1)))

    blocks: dict[str, str] = {}
    for index, (start, qid) in enumerate(starts):
        end = starts[index + 1][0] if index + 1 < len(starts) else len(paras)
        blocks.setdefault(qid, "\n".join(paras[start:end]))
    return blocks


def qid_candidates(var_name: str) -> tuple[str, ...]:
    clean = var_name[1:] if var_name.startswith("v") else var_name
    clean = re.sub(r"o\d+$", "", clean)
    clean = re.sub(r"m\d+$", "", clean)
    candidates = [clean]

    suffix_removed = re.sub(r"(city|town|U|D|R)$", "", clean)
    if suffix_removed != clean:
        candidates.append(suffix_removed)

    matrix = re.match(r"^([A-Z]*[0-9]+)s[0-9]+$", clean)
    if matrix:
        candidates.append(matrix.group(1))

    ck_removed = re.sub(r"CK$", "", clean)
    if ck_removed != clean:
        candidates.append(ck_removed)

    return tuple(dict.fromkeys(candidates))


def first_matched_block(var_name: str, blocks: dict[str, str]) -> tuple[str, str]:
    for qid in qid_candidates(var_name):
        if qid in blocks:
            return qid, blocks[qid]
    return "", ""


def extract_range_from_block(block: str, width: int) -> tuple[str, str, str]:
    if not block:
        return "", "manual", "no matching question block"

    range_match = re.search(r"範圍設定([^】\n]+)", block)
    if range_match:
        text = range_match.group(1)
        numbers = re.findall(r"\d+(?:\.\d+)?", text)
        upper = "999999999" if width >= 9 else "999"
        if "大於等於" in text and numbers:
            return compact_single_value_range(f"{numbers[0]},{upper}"), "high", "explicit lower-bound range"
        if "小於" in text and numbers:
            upper_value = float(numbers[0]) - 0.00001
            return compact_single_value_range(f"-100000,{upper_value:g}"), "medium", "explicit upper-bound range"
        if len(numbers) >= 2:
            return compact_single_value_range(f"{numbers[0]},{numbers[1]}"), "high", "explicit range tag"

    option_codes = [int(code) for code in re.findall(r"\((\d{1,2})\)", block)]
    if option_codes:
        return compact_single_value_range(f"{min(option_codes)},{max(option_codes)}"), "medium", "option-code span"

    return "", "manual", "no explicit range or option span"


def compact_single_value_range(value: str) -> str:
    parts = [part.strip() for part in value.split(",")]
    if len(parts) == 2 and parts[0] == parts[1]:
        return parts[0]
    return value


def normalize_date_bound(value: str) -> tuple[str, str]:
    parts = [part.strip() for part in value.split(",")]
    changed = False
    fixed: list[str] = []
    for part in parts:
        if part.isdigit() and len(part) == 13 and part.startswith("20"):
            fixed.append(part + "0")
            changed = True
        else:
            fixed.append(part)
    return ",".join(fixed), "date bound padded to 14 digits" if changed else ""


def normalize_user_datetime(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if len(digits) == 8:
        return digits + "000000"
    if len(digits) == 14:
        return digits
    raise ValueError(f"Date value must be YYYYMMDD, YYYY-MM-DD, or YYYYMMDDHHMMSS: {value}")


def configured_date_range(date_start: str | None, date_end: str | None) -> str:
    if not date_start and not date_end:
        return ""
    if not date_start or not date_end:
        raise ValueError("--date-start and --date-end must be provided together")
    return f"{normalize_user_datetime(date_start)},{normalize_user_datetime(date_end)}"


def is_datetime_width(width: int) -> bool:
    return width == 14


def infer_special_codes(var_name: str, r1: str, source: str, block: str) -> tuple[str, str, str]:
    if not r1:
        return "", "", ""
    if (
        "999999999" in r1
        and (
            re.search(r"(?:U|D|_[12]|_[12]CK)$", var_name)
            or "差額" in block
            or ("增加" in block and "減少" in block)
        )
    ):
        return "9999999996", "", "standard not-applicable code for increase/decrease amount"
    return "", "", ""


def max_digit_count(range_text: str) -> int:
    numbers = re.findall(r"-?\d+(?:\.\d+)?", range_text)
    max_digits = 1
    for number in numbers:
        integer = number.split(".", 1)[0].lstrip("-")
        max_digits = max(max_digits, len(integer))
    return max_digits


def special_code_for_width(width: int) -> str:
    if width <= 1:
        return "96"
    return "9" * (width - 1) + "6"


def is_probably_optional(var_name: str, block: str, source: str) -> bool:
    if source in {"date-format repair", "configured date range", "existing date range", "administrative code"}:
        return False
    clean = var_name[1:] if var_name.startswith("v") else var_name
    if re.search(r"(?:U|D|R|CK)$", clean) or "_" in clean:
        return True
    if re.search(r"(顯示此題|進入此題|續答|跳答)", block):
        return True
    return False


def compare_status_and_reason(existing: tuple[str, str, str, str], proposed: tuple[str, str, str, str], source: str) -> tuple[str, str]:
    if proposed == ("", "", "", ""):
        return "manual", "no automatic proposal"
    if existing == proposed:
        return "match", ""
    if existing[0] == proposed[0] and existing[1:] != proposed[1:]:
        return "range_match_special_diff", "r1 matches; special code columns differ"
    if source == "date-format repair":
        return "mismatch", "date bound format repaired; verify r2 special code"
    if source == "multiple-response binary variable":
        return "mismatch", "multiple-response special-code convention differs by project"
    if existing[0] and proposed[0] and existing[0] != proposed[0]:
        if existing[0].replace(",1", "") == proposed[0].replace(",1", ""):
            return "mismatch", "single value versus closed range representation"
        if source == "option-code span":
            return "mismatch", "Word option span includes codes that Excel excludes or stores as special code"
        if source.startswith("explicit"):
            return "mismatch", "Word explicit range differs from current Excel rule"
    return "mismatch", "automatic proposal differs from existing value"


def load_numeric_rows(
    workbook_path: Path,
    docx_path: Path,
    date_start: str | None = None,
    date_end: str | None = None,
) -> tuple[list[Proposal], dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    numeric_sheet = find_numeric_sheet(wb)
    ws = wb[numeric_sheet]
    headers = read_headers(ws)
    blocks = make_question_blocks(docx_path)
    multi_vars = read_multi_vars(wb)
    records = read_numeric_records(wb, numeric_sheet, headers)
    date_range = configured_date_range(date_start, date_end)

    proposals: list[Proposal] = []
    for record in records:
        var_name = record["var"]
        width = record["width"]
        existing = record["existing"]
        matched_qid, block = first_matched_block(var_name, blocks)

        r1, confidence, source = extract_range_from_block(block, width)
        note = ""

        if var_name in multi_vars:
            r1, confidence, source, note = "0,1", "high", "multiple-response binary variable", "from 複選題變項清單"
        elif var_name.lower() in SYSTEM_VAR_NAMES:
            r1, confidence, source, note = "", "manual", "system variable", "excluded from automatic range inference"
        elif "city" in var_name.lower():
            r1, confidence, source, note = "1,29", "high", "administrative city code", "city variables use project-wide 1,29"
        elif var_name.lower().endswith("town"):
            r1, confidence, source, note = "", "manual", "administrative town code", "needs project-specific code list"
        elif is_datetime_width(width):
            if date_range:
                r1, confidence, source, note = date_range, "high", "configured date range", "from app date range setting"
            else:
                fixed, date_note = normalize_date_bound(existing[0])
                if fixed:
                    source = "date-format repair" if date_note else "existing date range"
                    note = date_note or "from existing Excel date range"
                    r1, confidence = fixed, "high"
                else:
                    r1, confidence, source, note = "", "manual", "date range", "needs app date range setting"
        elif var_name.upper().startswith("VEND") or var_name.upper().startswith("END"):
            fixed, date_note = normalize_date_bound(existing[0])
            if fixed and date_note:
                r1, confidence, source, note = fixed, "high", "date-format repair", date_note
            else:
                r1, confidence, source, note = "", "manual", "end timestamp", "needs project fieldwork date rule"
        elif not matched_qid:
            r1, confidence, source, note = "", "manual", "no block", "needs parent-question mapping or manual review"

        r2, r3, r4 = "", "", ""
        if confidence in {"high", "medium"}:
            if source == "multiple-response binary variable" and not re.search(r"m0\d+$", var_name, flags=re.IGNORECASE):
                r2 = "96"
            elif source in {"date-format repair", "configured date range", "existing date range"}:
                r2, r3, r4 = existing[1], existing[2], existing[3]
            else:
                r2, r3, special_note = infer_special_codes(var_name, r1, source, block)
                if special_note:
                    note = special_note if not note else f"{note}; {special_note}"
                elif r1 and is_probably_optional(var_name, block, source):
                    r2 = special_code_for_width(width)
                    note = "optional question; inferred special code by variable width" if not note else f"{note}; optional question; inferred special code by variable width"

        proposed = (r1, r2, r3, r4)
        compare_status, mismatch_reason = compare_status_and_reason(existing, proposed, source)

        proposals.append(
            Proposal(
                var=var_name,
                qid="|".join(qid_candidates(var_name)),
                existing=existing,
                proposed=proposed,
                confidence=confidence,
                source=source,
                note=note,
                compare_status=compare_status,
                mismatch_reason=mismatch_reason,
                matched_question=matched_qid,
                question_excerpt=normalize_spaces(block[:260]),
            )
        )

    summary = {
        "workbook": str(workbook_path),
        "docx": str(docx_path),
        "numeric_sheet": numeric_sheet,
        "question_blocks": len(blocks),
        "rows": len(proposals),
        "confidence_counts": {},
        "exact_counts": {"r1": 0, "r1_r4": 0},
    }
    for proposal in proposals:
        summary["confidence_counts"][proposal.confidence] = summary["confidence_counts"].get(proposal.confidence, 0) + 1
        if proposal.existing[0] == proposal.proposed[0]:
            summary["exact_counts"]["r1"] += 1
        if proposal.existing == proposal.proposed:
            summary["exact_counts"]["r1_r4"] += 1
    return proposals, summary


def read_numeric_records(wb: openpyxl.Workbook, numeric_sheet: str, headers: dict[str, int]) -> list[dict[str, Any]]:
    ws = wb[numeric_sheet]
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[headers["var"]]:
            continue
        var_name = cell_text(row[headers["var"]])
        seen.add(var_name)
        width = int(cell_text(row[headers["寬度"]])) if cell_text(row[headers["寬度"]]).isdigit() else 2
        records.append(
            {
                "var": var_name,
                "width": width,
                "existing": tuple(cell_text(row[headers[col]]) for col in ("r1", "r2", "r3", "r4")),
            }
        )

    if "all" in wb.sheetnames:
        all_ws = wb["all"]
        all_headers = read_headers(all_ws)
        for row in all_ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[all_headers["變項名稱"]]:
                continue
            var_name = cell_text(row[all_headers["變項名稱"]])
            kind = cell_text(row[all_headers["變項屬性"]])
            if var_name in seen or kind != "數值" or var_name.lower() in SYSTEM_VAR_NAMES:
                continue
            width = int(cell_text(row[all_headers["寬度"]])) if cell_text(row[all_headers["寬度"]]).isdigit() else 2
            records.append({"var": var_name, "width": width, "existing": ("", "", "", "")})
            seen.add(var_name)
    return records


def read_multi_vars(wb: openpyxl.Workbook) -> set[str]:
    if "複選題變項清單" not in wb.sheetnames:
        return set()
    ws = wb["複選題變項清單"]
    return {cell_text(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}


def write_csv(path: Path, proposals: list[Proposal]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "var",
                "qid_candidates",
                "matched_question",
                "existing_r1",
                "existing_r2",
                "existing_r3",
                "existing_r4",
                "proposed_r1",
                "proposed_r2",
                "proposed_r3",
                "proposed_r4",
                "confidence",
                "source",
                "note",
                "compare_status",
                "mismatch_reason",
                "question_excerpt",
            ]
        )
        for p in proposals:
            writer.writerow(
                [
                    p.var,
                    p.qid,
                    p.matched_question,
                    *p.existing,
                    *p.proposed,
                    p.confidence,
                    p.source,
                    p.note,
                    p.compare_status,
                    p.mismatch_reason,
                    p.question_excerpt,
                ]
            )


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate proposed r1-r4 values for numeric survey checks.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--date-start", default=None, help="Default date range start for width-14 numeric datetime variables.")
    parser.add_argument("--date-end", default=None, help="Default date range end for width-14 numeric datetime variables.")
    args = parser.parse_args()

    proposals, summary = load_numeric_rows(args.workbook, args.docx, args.date_start, args.date_end)
    write_csv(args.output_csv, proposals)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(
        json.dumps(
            {
                "summary": summary,
                "manual_review": [
                    {
                        "var": p.var,
                        "existing": p.existing,
                        "proposed": p.proposed,
                        "source": p.source,
                        "note": p.note,
                        "compare_status": p.compare_status,
                        "mismatch_reason": p.mismatch_reason,
                        "matched_question": p.matched_question,
                    }
                    for p in proposals
                    if p.confidence == "manual"
                ],
                "mismatches": [
                    {
                        "var": p.var,
                        "existing": p.existing,
                        "proposed": p.proposed,
                        "confidence": p.confidence,
                        "source": p.source,
                        "note": p.note,
                        "compare_status": p.compare_status,
                        "mismatch_reason": p.mismatch_reason,
                    }
                    for p in proposals
                    if p.compare_status == "mismatch"
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"rows: {summary['rows']}")
    print(f"confidence: {summary['confidence_counts']}")
    print(f"exact r1: {summary['exact_counts']['r1']}")
    print(f"exact r1-r4: {summary['exact_counts']['r1_r4']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
