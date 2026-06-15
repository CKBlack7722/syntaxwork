from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from docx import Document


SYSTEM_PREFIXES = ("END",)
SYSTEM_NAMES = {
    "ct",
    "ic",
    "r",
    "rt",
    "sendtime",
    "smst",
    "smstime1",
    "termination",
    "wendtime",
    "wlast",
    "wst1",
    "wst2",
    "wst3",
    "wst4",
    "wstarttime",
    "browser1",
    "browser2",
    "browser3",
    "browser4",
    "cellphone",
    "device1",
    "device2",
    "device3",
    "device4",
    "email",
    "id",
    "ip1",
    "ip2",
    "ip3",
    "ip4",
    "name",
}
RESERVED_WORDS = {"and", "or", "not", "in", "any", "range", "sys"}


@dataclass(frozen=True)
class WordQuestion:
    qid: str
    excerpt: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(text: str) -> str:
    return unicodedata.normalize("NFKC", text or "").replace("＿", "_")


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def read_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if value not in (None, ""):
            headers[str(value).strip()] = idx
    return headers


def extract_word_questions(docx_path: Path) -> dict[str, WordQuestion]:
    doc = Document(str(docx_path))
    paras = [normalize_text(p.text.strip()) for p in doc.paragraphs if p.text.strip()]
    starts: list[tuple[int, str]] = []
    for idx, paragraph in enumerate(paras):
        compact = paragraph.replace(" ", "")
        match = re.match(r"^([A-Z]{1,4}[0-9]+(?:_[0-9]+)?[A-Z]?[0-9]?)(?![0-9])", compact)
        if not match:
            match = re.match(r"^([0-9]{1,3})(?![\.)、])", compact)
        if match:
            starts.append((idx, match.group(1)))

    questions: dict[str, WordQuestion] = {}
    for index, (start, qid) in enumerate(starts):
        end = starts[index + 1][0] if index + 1 < len(starts) else len(paras)
        block = "\n".join(paras[start:end])
        questions.setdefault(qid, WordQuestion(qid=qid, excerpt=normalize_spaces(block[:320])))
    return questions


def base_qid_from_var(var_name: str) -> str:
    clean = var_name.strip()
    if clean.startswith("v"):
        clean = clean[1:]
    clean = re.sub(r"o\d+$", "", clean)
    clean = re.sub(r"m\d+$", "", clean)
    clean = re.sub(r"(city|town)$", "", clean)
    matrix = re.match(r"^([A-Z]*[0-9]+)s[0-9A-Z]+$", clean)
    if matrix:
        clean = matrix.group(1)
    return clean


def qid_candidates_from_var(var_name: str) -> set[str]:
    base = base_qid_from_var(var_name)
    candidates = {base}
    candidates.add(re.sub(r"(U|D|R)$", "", base))
    candidates.add(re.sub(r"CK$", "", base))
    candidates.add(re.sub(r"[0-9]+R$", "", base))
    return {c for c in candidates if c}


def normalize_sheet_question(value: Any) -> str:
    text = cell_text(value)
    if not text:
        return ""
    if text.isdigit():
        return text
    return text[1:] if text.startswith("v") else text


def add_source(target: dict[str, set[str]], qid: str, source: str) -> None:
    if qid:
        target.setdefault(qid, set()).add(source)


def read_excel_question_sources(workbook_path: Path) -> dict[str, set[str]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sources: dict[str, set[str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == "all":
            headers = read_headers(ws)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                var_name = cell_text(row[0])
                if var_name.lower() in SYSTEM_NAMES:
                    continue
                for qid in qid_candidates_from_var(var_name):
                    add_source(sources, qid, "all")
        elif sheet_name.startswith("數值題") or sheet_name.startswith("開放欄位"):
            headers = read_headers(ws)
            var_col = headers.get("var")
            if var_col is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[var_col]:
                    continue
                for qid in qid_candidates_from_var(cell_text(row[var_col])):
                    add_source(sources, qid, sheet_name)
        elif sheet_name.startswith("複選題_") or sheet_name == "複選題內_互斥":
            headers = read_headers(ws)
            q_col = headers.get("題號")
            if q_col is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[q_col]:
                    add_source(sources, normalize_sheet_question(row[q_col]), sheet_name)
        elif sheet_name == "複選題變項清單":
            headers = read_headers(ws)
            group_col = headers.get("分組")
            if group_col is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[group_col]:
                    add_source(sources, normalize_sheet_question(row[group_col]), sheet_name)
        elif sheet_name == "邏輯組":
            headers = read_headers(ws)
            for col_name in ("條件", "應答", "不應答", "限制"):
                col = headers.get(col_name)
                if col is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[col]:
                        continue
                    for token in re.findall(r"\bv?[A-Za-z]+[A-Za-z0-9_]*\b", cell_text(row[col])):
                        if token.lower() in RESERVED_WORDS:
                            continue
                        for qid in qid_candidates_from_var(token):
                            add_source(sources, qid, f"{sheet_name}:{col_name}")
    return sources


def is_probably_non_question(qid: str) -> bool:
    if not qid:
        return True
    if qid.upper().startswith(SYSTEM_PREFIXES):
        return True
    if qid in {"A", "B", "C"}:
        return True
    if qid.lower() in SYSTEM_NAMES or qid.lower() in RESERVED_WORDS:
        return True
    if qid.isdigit() and int(qid) >= 100:
        return True
    if re.match(r"^[A-Z]+[0-9]{3,}$", qid):
        return True
    return False


def compare_project(name: str, workbook_path: Path, docx_path: Path) -> dict[str, Any]:
    word_questions = extract_word_questions(docx_path)
    excel_sources = read_excel_question_sources(workbook_path)

    word_ids = {qid for qid in word_questions if not is_probably_non_question(qid)}
    excel_ids = {qid for qid in excel_sources if not is_probably_non_question(qid)}

    missing_in_excel = sorted(word_ids - excel_ids, key=sort_key)
    extra_in_excel = sorted(excel_ids - word_ids, key=sort_key)
    covered = sorted(word_ids & excel_ids, key=sort_key)

    return {
        "project": name,
        "workbook": str(workbook_path),
        "docx": str(docx_path),
        "summary": {
            "word_questions": len(word_ids),
            "excel_question_ids": len(excel_ids),
            "covered": len(covered),
            "missing_in_excel": len(missing_in_excel),
            "extra_in_excel": len(extra_in_excel),
        },
        "covered": [{"qid": qid, "excel_sources": sorted(excel_sources[qid])} for qid in covered],
        "missing_in_excel": [
            {"qid": qid, "word_excerpt": word_questions[qid].excerpt}
            for qid in missing_in_excel
        ],
        "extra_in_excel": [
            {"qid": qid, "excel_sources": sorted(excel_sources[qid])}
            for qid in extra_in_excel
        ],
    }


def sort_key(value: str) -> tuple[Any, ...]:
    match = re.match(r"^([A-Z]*)(\d+)(.*)$", value)
    if not match:
        return (value, 0, "")
    prefix, number, suffix = match.groups()
    return (prefix, int(number), suffix)


def write_csv(path: Path, reports: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["project", "status", "qid", "excel_sources", "word_excerpt"])
        for report in reports:
            project = report["project"]
            for item in report["missing_in_excel"]:
                writer.writerow([project, "word_missing_in_excel", item["qid"], "", item["word_excerpt"]])
            for item in report["extra_in_excel"]:
                writer.writerow([project, "excel_not_found_in_word", item["qid"], "; ".join(item["excel_sources"]), ""])
            for item in report["covered"]:
                writer.writerow([project, "covered", item["qid"], "; ".join(item["excel_sources"]), ""])


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare Word questionnaire question ids with Excel workbook coverage.")
    parser.add_argument("--project", action="append", nargs=3, metavar=("NAME", "WORKBOOK", "DOCX"), required=True)
    parser.add_argument("--output-json", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    args = parser.parse_args()

    reports = [compare_project(name, Path(workbook), Path(docx)) for name, workbook, docx in args.project]
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(json.dumps(reports, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(args.output_csv, reports)

    for report in reports:
        summary = report["summary"]
        print(
            f"{report['project']}: word={summary['word_questions']} excel={summary['excel_question_ids']} "
            f"covered={summary['covered']} missing={summary['missing_in_excel']} extra={summary['extra_in_excel']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
