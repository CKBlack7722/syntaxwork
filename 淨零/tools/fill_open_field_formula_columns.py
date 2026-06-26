from __future__ import annotations

import argparse
import json
import re
import shutil
from pathlib import Path
from typing import Any

import openpyxl


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(value): idx
        for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if cell_text(value)
    }


def infer_qid_from_open_var(var_name: str) -> str:
    clean = cell_text(var_name)
    if clean.lower().endswith("_oth"):
        return re.sub(r"_oth$", "", clean, flags=re.IGNORECASE)
    return re.sub(r"o[0-9A-Za-z]+$", "", clean, flags=re.IGNORECASE)


def infer_option_from_open_var(var_name: str, qid: str) -> str:
    clean = cell_text(var_name)
    if clean.lower().endswith("_oth"):
        return ""
    suffix = clean[len(qid) :]
    match = re.search(r"o([0-9A-Za-z]+)$", suffix, flags=re.IGNORECASE)
    if not match:
        return ""
    value = match.group(1)
    return str(int(value)) if value.isdigit() else value


def read_multi_qids(wb: openpyxl.Workbook) -> set[str]:
    if "複選題變項清單" not in wb.sheetnames:
        return set()
    ws = wb["複選題變項清單"]
    h = headers(ws)
    if "變項名稱" not in h:
        return set()
    qids: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        var = cell_text(row[h["變項名稱"] - 1])
        match = re.match(r"^(v.+?)m[0-9]+$", var, flags=re.IGNORECASE)
        if match:
            qids.add(match.group(1))
    return qids


def fill(workbook_path: Path, apply: bool) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path)
    if "開放欄位" not in wb.sheetnames:
        raise ValueError("workbook has no 開放欄位 sheet")
    ws = wb["開放欄位"]
    h = headers(ws)
    required = ["var", "是否複選"]
    missing = [name for name in required if name not in h]
    if missing:
        raise ValueError(f"missing headers: {missing}")

    multi_qids = read_multi_qids(wb)
    changed: list[dict[str, Any]] = []
    formula_filled = 0
    multi_changed = 0
    for row_idx in range(2, ws.max_row + 1):
        var_name = cell_text(ws.cell(row_idx, h["var"]).value)
        if not var_name:
            continue
        qid = infer_qid_from_open_var(var_name)
        option = infer_option_from_open_var(var_name, qid)
        is_multi = "1" if qid in multi_qids else "0"
        existing_multi = cell_text(ws.cell(row_idx, h["是否複選"]).value)
        if existing_multi != is_multi:
            changed.append({"row": row_idx, "var": var_name, "field": "是否複選", "old": existing_multi, "new": is_multi})
            if apply:
                ws.cell(row_idx, h["是否複選"]).value = int(is_multi)
            multi_changed += 1
        should_fill_formula = bool(re.search(r"o\d+|_oth", var_name)) and option
        formulas = {
            "題號": f'=IF(ISERROR(SEARCH("_oth",C{row_idx},1)),MID($C{row_idx},1,SEARCH("o",C{row_idx},1)-1),MID($C{row_idx},1,SEARCH("_oth",C{row_idx},1)-1))',
            "題號變項名稱": f'=IF(ISNUMBER(SEARCH("oth",$C{row_idx},1)),MID($C{row_idx},1,SEARCH("_",$C{row_idx},1)-1),IF(AND($C{row_idx}<>"vZ0city_oth",$F{row_idx}=0),VALUE(MID($C{row_idx},SEARCH("o",$C{row_idx},1)+1,LEN($C{row_idx})-SEARCH("o",$C{row_idx},1))),CONCAT(MID($C{row_idx},1,SEARCH("o",$C{row_idx},1)-1),"m",I{row_idx})))',
            "選項數值": f'=MID($C{row_idx},LEN($G{row_idx})+2,2)',
            "var2_new": f'=IF(F{row_idx}=0,G{row_idx},H{row_idx})',
            "range_new": f'=IF(F{row_idx}=0,H{row_idx},1)',
            "n": f'=IF(LEN(K{row_idx})=1,LEN(K{row_idx})+1,LEN(K{row_idx}))',
        }
        if should_fill_formula:
            for field, formula in formulas.items():
                if field not in h:
                    continue
                cell = ws.cell(row_idx, h[field])
                if cell.value in (None, "") or str(cell.value).strip() == "#VALUE!":
                    changed.append({"row": row_idx, "var": var_name, "field": field, "old": cell_text(cell.value), "new": formula})
                    if apply:
                        cell.value = formula
                    formula_filled += 1
    backup = ""
    if apply and changed:
        backup_path = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_open_field_formula_fill.xlsx"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        if not backup_path.exists():
            shutil.copy2(workbook_path, backup_path)
        wb.save(workbook_path)
        backup = str(backup_path)
    return {
        "multi_qid_count": len(multi_qids),
        "multi_changed": multi_changed,
        "formula_filled": formula_filled,
        "changed": changed,
        "backup": backup,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    report = fill(args.workbook, args.apply)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: v for k, v in report.items() if k != "changed"}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
