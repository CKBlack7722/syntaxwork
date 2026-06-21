from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

import openpyxl


REVIEW_SHEET = "邏輯組_人工確認"


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def all_vars(workbook: Path) -> set[str]:
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    return {text(row[0]) for row in wb["all"].iter_rows(min_row=2, values_only=True) if text(row[0])}


def multi_vars(names: set[str], qid: str) -> list[str]:
    return sorted((name for name in names if re.fullmatch(rf"v{re.escape(qid)}m\d+", name, flags=re.I)), key=lambda value: int(re.search(r"m(\d+)$", value, re.I).group(1)))


def sum_expr(names: list[str]) -> str:
    return ",".join(names)


def proposal(kind: str, qid: str, condition: str, required: str, forbidden: str, source: str, note: str) -> dict[str, str]:
    inverse_required = forbidden
    inverse_forbidden = required
    return {
        "status": "proposal",
        "rule_type": kind,
        "qid": qid,
        "condition": f"spss:{condition}",
        "required": required,
        "forbidden": forbidden,
        "inverse_condition": f"spss:not ({condition})",
        "inverse_required": inverse_required,
        "inverse_forbidden": inverse_forbidden,
        "source": source,
        "note": note,
    }


def build_proposals(workbook: Path) -> list[dict[str, str]]:
    names = all_vars(workbook)
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    ws = wb[REVIEW_SHEET]
    h = {text(cell.value): cell.column for cell in ws[1] if text(cell.value)}
    result: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        qid = text(ws.cell(row_idx, h["題號"]).value)
        source = text(ws.cell(row_idx, h["問卷原始文字"]).value)
        note = text(ws.cell(row_idx, h["人工備註"]).value)
        merged = source + "\n" + note
        if qid == "E3_1" and "加總超過" in merged:
            vars_ = [f"v{name}" for name in ("E5", "E6", "E8", "E9", "E11", "E12") if f"v{name}" in names]
            pieces = ",".join(f"(not any({name},9797,9898,99996))*{name}" for name in vars_)
            cond = f"sum({pieces})>2400"
            result.append(proposal("sum_excluding_codes", qid, cond, "vE3_1", "", source, "反向不應答條件需以 not (...) 產生"))
        elif qid == "B7" and "B7a" in merged and "只選1個" in merged:
            vars_ = multi_vars(names, "B7a")
            args = sum_expr(vars_)
            cond = f"sum({args})>1 & max({args})<96"
            result.append(proposal("multi_count", qid, cond, "vB7", "", source, "反向不應答：not (此條件)"))
        elif qid == "I4" and "vB8m03" in merged:
            result.append(proposal("combined_option_zero", qid, "(vB8m03=1 | vB8m06=1) & vI1=0 & vI3=0", "vI4", "", source, "反向不應答：not (此條件)"))
        elif "K2答(90)" in merged and "Q5答(02)" in merged:
            vars_ = multi_vars(names, "K2")
            args = sum_expr(vars_)
            cond = f"vK2m90=1 | (min({args})=97 & max({args})=97) | (min({args})=98 & max({args})=98) | vQ5=2"
            target = "vP3_5" if "vP3_5" in names else ""
            result.append(proposal("multi_special_code_skip", "P3_5", cond, "", target, source, "K2 的 97/98 為整題特殊碼，須所有展開變項一致"))
        elif qid == "ZE2_1" and "ZE2答(01)或(02)" in merged:
            result.append(proposal("multi_option_need", qid, "vZE2m01=1 | vZE2m02=1", "vZE2_1", "", source, "反向不應答：not (此條件)"))
    unique: dict[tuple[str, str, str], dict[str, str]] = {}
    for item in result:
        unique[(item["rule_type"], item["qid"], item["condition"])] = item
    return list(unique.values())


def main() -> None:
    parser = argparse.ArgumentParser(description="Create reviewable advanced logic proposals from annotated review rows.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()
    rows = build_proposals(args.workbook)
    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.output_csv.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=["status", "rule_type", "qid", "condition", "required", "forbidden", "inverse_condition", "inverse_required", "inverse_forbidden", "source", "note"])
        writer.writeheader()
        writer.writerows(rows)
    args.report.write_text(json.dumps({"proposals": rows, "count": len(rows)}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"count": len(rows)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
