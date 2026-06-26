from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl
from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph


MAIN_SHEET_PREFIX = "複選題"
MUTEX_SHEET = "複選題內_互斥"
VAR_LIST_SHEET = "複選題變項清單"
STANDARD_SPECIAL_CODES = {"96", "97", "98"}


@dataclass
class MultiProposal:
    group: str
    qid: str
    status: str
    reason: str
    var_count: int
    doc_codes: str
    var_codes: str
    missing_in_doc: str
    extra_in_doc: str
    width: str
    start_code: str
    end_code: str
    option_count: str
    unavailable_count: str
    no_response_all: str
    mutex_code: str
    loop_count: str
    no_response_1: str
    no_response_2: str
    no_response_3: str
    no_response_4: str
    no_response_5: str
    excerpt: str


@dataclass
class MutexProposal:
    group: str
    qid: str
    status: str
    reason: str
    width: str
    mutex_start: str
    mutex_end: str
    kind: str
    unavailable_count: str
    noncontinuous_mutex: str
    mutex_code: str
    excerpt: str


def norm(value: str) -> str:
    return unicodedata.normalize("NFKC", value or "").strip()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(value): idx
        for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if cell_text(value)
    }


def question_id_from_text(text: str) -> str:
    compact = norm(text).replace(" ", "").upper()
    compact = re.sub(r"^(?:【[^】]+】)+", "", compact)
    match = re.match(r"^([A-Z]{1,5}[0-9][A-Z0-9_]*(?:_[0-9]+)?[A-Z]?)", compact)
    return match.group(1).upper() if match else ""


def table_text(table: Table) -> str:
    rows = []
    for row in table.rows:
        rows.append("\n".join(norm(cell.text) for cell in row.cells if norm(cell.text)))
    return "\n".join(rows)


def doc_items(docx_path: Path) -> list[tuple[str, str]]:
    doc = Document(docx_path)
    items: list[tuple[str, str]] = []
    for child in doc.element.body.iterchildren():
        if child.tag == qn("w:p"):
            text = norm(Paragraph(child, doc).text)
            if text:
                items.append(("P", text))
        elif child.tag == qn("w:tbl"):
            items.append(("T", table_text(Table(child, doc))))
    return items


def doc_blocks(docx_path: Path) -> dict[str, str]:
    items = doc_items(docx_path)
    starts: list[tuple[int, str]] = []
    for idx, (kind, text) in enumerate(items):
        if kind != "P":
            continue
        qid = question_id_from_text(text)
        if qid:
            starts.append((idx, qid))
    blocks: dict[str, str] = {}
    for i, (start, qid) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else len(items)
        blocks.setdefault(qid, "\n".join(text for _, text in items[start:end]))
    return blocks


def qid_from_group(group: str) -> str:
    return group[1:].upper() if group.lower().startswith("v") else group.upper()


def excel_qid(group: str) -> str:
    return group[1:] if group.lower().startswith("v") else group


def infer_group_from_var(var_name: str) -> str:
    match = re.match(r"^(v.+?)m[0-9]+$", cell_text(var_name), flags=re.I)
    return match.group(1) if match else ""


def option_code(var_name: str) -> str:
    match = re.search(r"m0*([0-9]+)$", var_name, flags=re.I)
    return str(int(match.group(1))) if match else ""


def option_codes_from_block(block: str) -> list[str]:
    codes: list[str] = []
    option_matches = re.findall(r"\u25a1\s*[\(\uff08]\s*(\d{1,3})\s*[\)\uff09]", block)
    if not option_matches:
        text_without_brackets = re.sub(r"\u3010[^\u3011]*\u3011", " ", block)
        option_matches = re.findall(
            r"(?:^|\s)[\(\uff08]\s*(\d{1,3})\s*[\)\uff09]",
            text_without_brackets,
        )
    for code in option_matches:
        code = str(int(code))
        if code not in codes:
            codes.append(code)
    return codes


def split_mutex_targets(start: str, end: str, available_codes: list[str], mutex_code: str) -> tuple[str, str]:
    start_i, end_i = int(start), int(end)
    candidates = sorted(
        {int(code) for code in available_codes if code != mutex_code and start_i <= int(code) <= end_i},
    )
    if not candidates:
        return end, ""
    contiguous_end = candidates[0]
    expected = candidates[0]
    rest: list[str] = []
    for code in candidates:
        if code == expected:
            contiguous_end = code
            expected += 1
        else:
            rest.append(str(code))
    return str(contiguous_end), ",".join(rest)


def internal_mutex_rules_from_block(block: str) -> list[tuple[str, str, str]]:
    rules: list[tuple[str, str, str]] = []
    for line in block.splitlines():
        if "\u4e92\u65a5" not in line or "\u8207\u9078\u9805" not in line:
            continue
        current = re.search(r"(?:\u25a1\s*)?[\(\uff08]\s*(\d{1,3})\s*[\)\uff09]", line)
        target_range = re.search(
            r"\u8207\u9078\u9805\s*[\(\uff08]\s*(\d{1,3})\s*[\)\uff09]\s*[-\uff5e~]\s*[\(\uff08]\s*(\d{1,3})\s*[\)\uff09]",
            line,
        )
        if not current:
            continue
        mutex_code = str(int(current.group(1)))
        if target_range:
            start, end = str(int(target_range.group(1))), str(int(target_range.group(2)))
        else:
            continue
        rule = (mutex_code, start, end)
        if rule not in rules:
            rules.append(rule)
    return rules


def no_response_codes(codes: list[str]) -> list[str]:
    return [code for code in codes if int(code) >= 90]


def normal_codes(codes: list[str]) -> list[str]:
    return [code for code in codes if int(code) < 90]


def read_var_widths(wb: openpyxl.Workbook) -> dict[str, int]:
    widths: dict[str, int] = {}
    ws = wb["all"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            try:
                widths[cell_text(row[0])] = int(cell_text(row[2]) or "2")
            except ValueError:
                widths[cell_text(row[0])] = 2
    return widths


def read_groups(wb: openpyxl.Workbook) -> dict[str, list[str]]:
    ws = wb[VAR_LIST_SHEET]
    groups: dict[str, list[str]] = {}
    for var, group in ws.iter_rows(min_row=2, values_only=True):
        if not var:
            continue
        group_text = cell_text(group)
        inferred = infer_group_from_var(cell_text(var))
        if not group_text or group_text.startswith("=") or group_text.lower() == "v":
            group_text = inferred
        if group_text:
            groups.setdefault(group_text, []).append(cell_text(var))
    return groups


def get_sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    matches = [
        name
        for name in wb.sheetnames
        if name.startswith(prefix) and name not in {MUTEX_SHEET, VAR_LIST_SHEET}
    ]
    if not matches:
        raise ValueError(f"missing sheet prefix {prefix}")
    return wb[matches[0]]


def build_proposals(workbook_path: Path, docx_path: Path) -> tuple[list[MultiProposal], list[MutexProposal]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    groups = read_groups(wb)
    widths = read_var_widths(wb)
    blocks = doc_blocks(docx_path)
    multi: list[MultiProposal] = []
    mutex: list[MutexProposal] = []
    for group, vars_for_group in groups.items():
        qid = qid_from_group(group)
        block = blocks.get(qid, "")
        var_codes = [code for code in (option_code(var) for var in vars_for_group) if code]
        raw_doc_codes = option_codes_from_block(block)
        doc_codes = [code for code in raw_doc_codes if code not in STANDARD_SPECIAL_CODES]
        missing = [code for code in var_codes if code not in doc_codes]
        extra = [code for code in doc_codes if code not in var_codes]
        normal = normal_codes(var_codes)
        no_resp = no_response_codes(sorted(set(var_codes + raw_doc_codes), key=int))
        mutex_rules = internal_mutex_rules_from_block(block)
        mutex_codes = [rule[0] for rule in mutex_rules if rule[0] in var_codes]
        width = str(widths.get(vars_for_group[0], 2)) if vars_for_group else "2"
        status = "match" if block and not missing and not extra else "review"
        reason = ""
        if not block:
            reason = "question block not found"
        elif missing:
            reason = "variable codes not found in questionnaire"
        elif extra:
            reason = "questionnaire has codes not in variable list"
        no_resp_padded = (no_resp + ["", "", "", "", ""])[:5]
        start_code = min(normal, key=int) if normal else ""
        end_code = max(normal, key=int) if normal else ""
        unavailable_count = str(len(no_resp)) if no_resp else "0"
        multi.append(
            MultiProposal(
                group=group,
                qid=qid,
                status=status,
                reason=reason,
                var_count=len(vars_for_group),
                doc_codes=",".join(raw_doc_codes),
                var_codes=",".join(var_codes),
                missing_in_doc=",".join(missing),
                extra_in_doc=",".join(extra),
                width=width,
                start_code=start_code,
                end_code=end_code,
                option_count=str(len(vars_for_group)),
                unavailable_count=unavailable_count,
                no_response_all=",".join(no_resp),
                mutex_code=",".join(mutex_codes),
                loop_count=str(max(len(vars_for_group) - 1, 0)),
                no_response_1=no_resp_padded[0],
                no_response_2=no_resp_padded[1],
                no_response_3=no_resp_padded[2],
                no_response_4=no_resp_padded[3],
                no_response_5=no_resp_padded[4],
                excerpt=re.sub(r"\s+", " ", block[:350]),
            )
        )
        for code, mutex_start, mutex_end in mutex_rules:
            if code in var_codes:
                continuous_end, noncontinuous = split_mutex_targets(mutex_start, mutex_end, var_codes, code)
                mutex.append(
                    MutexProposal(
                        group=group,
                        qid=qid,
                        status="match",
                        reason="",
                        width=width,
                        mutex_start=mutex_start,
                        mutex_end=continuous_end,
                        kind="range",
                        unavailable_count=str(len(no_resp)) if no_resp else "0",
                        noncontinuous_mutex=noncontinuous,
                        mutex_code=code,
                        excerpt=re.sub(r"\s+", " ", block[:350]),
                    )
                )
    return multi, mutex


def write_csv(path: Path, rows: list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = list(asdict(rows[0]).keys()) if rows else ["group"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def apply_to_workbook(workbook_path: Path, multi: list[MultiProposal], mutex: list[MutexProposal]) -> Path:
    backup = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_multi_fill.xlsx"
    backup.parent.mkdir(parents=True, exist_ok=True)
    if not backup.exists():
        shutil.copy2(workbook_path, backup)
    wb = openpyxl.load_workbook(workbook_path)
    main_ws = get_sheet_by_prefix(wb, MAIN_SHEET_PREFIX)
    mutex_ws = wb[MUTEX_SHEET]
    mh = headers(main_ws)
    xh = headers(mutex_ws)
    for idx, proposal in enumerate(multi, start=2):
        if proposal.status != "match":
            continue
        values = {
            "題號": excel_qid(proposal.group),
            "寬度": proposal.width,
            "選項編號_起": proposal.start_code,
            "選項編號_迄": proposal.end_code,
            "選項數量": proposal.option_count,
            "不可選擇數量": proposal.unavailable_count,
            "全部無反應選項": proposal.no_response_all,
            "互斥選項": proposal.mutex_code,
            "迴圈進行次數": proposal.loop_count,
            "無反應選項1": proposal.no_response_1,
            "無反應選項2": proposal.no_response_2,
            "無反應選項3": proposal.no_response_3,
            "無反應選項4": proposal.no_response_4,
            "無反應選項5": proposal.no_response_5,
        }
        for header, value in values.items():
            if header in mh:
                main_ws.cell(idx, mh[header]).value = value or None
    for idx, proposal in enumerate(mutex, start=2):
        values = {
            "題號": excel_qid(proposal.group),
            "寬度": proposal.width,
            "互斥選項編號_起": proposal.mutex_start,
            "互斥選項編號_迄": proposal.mutex_end,
            "類型": proposal.kind,
            "不可選擇數量": proposal.unavailable_count,
            "非連續互斥選項": proposal.noncontinuous_mutex,
            "互斥選項": proposal.mutex_code,
        }
        for header, value in values.items():
            if header in xh:
                mutex_ws.cell(idx, xh[header]).value = value or None
    wb.save(workbook_path)
    return backup


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--output-multi", required=True, type=Path)
    parser.add_argument("--output-mutex", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    multi, mutex = build_proposals(args.workbook, args.docx)
    write_csv(args.output_multi, multi)
    write_csv(args.output_mutex, mutex)
    report = {
        "multi_count": len(multi),
        "mutex_count": len(mutex),
        "multi_status": {},
        "mutex_status": {},
        "review": [asdict(row) for row in multi if row.status != "match"],
    }
    for row in multi:
        report["multi_status"][row.status] = report["multi_status"].get(row.status, 0) + 1
    for row in mutex:
        report["mutex_status"][row.status] = report["mutex_status"].get(row.status, 0) + 1
    if args.apply:
        backup = apply_to_workbook(args.workbook, multi, mutex)
        report["applied"] = True
        report["backup"] = str(backup)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
