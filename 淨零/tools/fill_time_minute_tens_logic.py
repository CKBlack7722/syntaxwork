from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path
from typing import Any

import openpyxl


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(value): idx
        for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if cell_text(value)
    }


def candidate_time_vars(wb: openpyxl.Workbook) -> list[str]:
    if "數值題" not in wb.sheetnames:
        return []
    ws = wb["數值題"]
    h = headers(ws)
    required = {"var", "寬度", "r1"}
    if not required.issubset(h):
        return []
    result: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        var_name = cell_text(ws.cell(row_idx, h["var"]).value)
        width = cell_text(ws.cell(row_idx, h["寬度"]).value)
        r1 = cell_text(ws.cell(row_idx, h["r1"]).value)
        if var_name and width == "5" and r1 == "1,2359":
            result.append(var_name)
    return result


def fill(workbook_path: Path, apply: bool) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path)
    if "邏輯組" not in wb.sheetnames:
        raise ValueError("workbook has no 邏輯組 sheet")
    ws = wb["邏輯組"]
    h = headers(ws)
    if "限制" not in h:
        raise ValueError("邏輯組 sheet has no 限制 column")
    existing_limits = {
        cell_text(ws.cell(row_idx, h["限制"]).value)
        for row_idx in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row_idx, h["限制"]).value)
    }
    candidates = candidate_time_vars(wb)
    additions: list[dict[str, Any]] = []
    blank_rows = [
        row_idx
        for row_idx in range(2, ws.max_row + 1)
        if not any(cell_text(ws.cell(row_idx, col).value) for col in (h["條件"], h["應答"], h["不應答"], h["限制"]))
    ]
    numeric_m_values = [
        int(cell_text(ws.cell(row_idx, h["m"]).value))
        for row_idx in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row_idx, h["m"]).value).isdigit()
    ]
    next_m = max(numeric_m_values) + 1 if numeric_m_values else 1
    for var_name in candidates:
        limit = f"{var_name} minute_tens in 0,1,2,3,4,5"
        if limit in existing_limits:
            continue
        if not blank_rows:
            row_idx = ws.max_row + 1
            blank_rows.append(row_idx)
            if apply:
                ws.cell(row_idx, h["m"]).value = next_m
                if "p" in h:
                    ws.cell(row_idx, h["p"]).value = f"=A{row_idx}"
            next_m += 1
        row_idx = blank_rows.pop(0)
        additions.append({"var": var_name, "limit": limit, "status": "apply", "row": row_idx})
        if apply:
            ws.cell(row_idx, h["限制"]).value = limit
    backup = ""
    if apply and any(item["status"] == "apply" for item in additions):
        backup_path = workbook_path.parent / "generated" / f"{workbook_path.stem}.before_time_minute_tens_logic.xlsx"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        if not backup_path.exists():
            shutil.copy2(workbook_path, backup_path)
        wb.save(workbook_path)
        backup = str(backup_path)
    return {"candidates": candidates, "additions": additions, "backup": backup}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    report = fill(args.workbook, args.apply)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
