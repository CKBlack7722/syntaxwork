from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import openpyxl


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(text: str) -> str:
    return unicodedata.normalize("NFKC", text or "").replace("（", "(").replace("）", ")")


def read_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {str(value).strip(): idx for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) if value not in (None, "")}


def numeric_sheet_name(wb: openpyxl.Workbook) -> str:
    for name in wb.sheetnames:
        if name.startswith("數值題"):
            return name
    raise ValueError("No numeric sheet found")


def docx_text(docx_path: Path) -> str:
    parts: list[str] = []
    with ZipFile(docx_path) as archive:
        for name in archive.namelist():
            if not (name.startswith("word/") and name.endswith(".xml")):
                continue
            try:
                root = ET.fromstring(archive.read(name))
            except ET.ParseError:
                continue
            for element in root.iter():
                if element.tag.endswith("}t") and element.text:
                    parts.append(element.text)
    return normalize_text("".join(parts))


def question_blocks(text: str) -> dict[str, str]:
    qid_pattern = re.compile(
        r"(?<![A-Z0-9(第])([A-Z][A-Z0-9_]{0,12}[0-9][A-Z0-9_]*|[0-9]{1,3}(?:_[0-9]+)?)(?:\s+|(?=請問|一般|為了|承上題|您上))(?=[\u4e00-\u9fffA-Z])",
        flags=re.IGNORECASE,
    )
    starts: list[tuple[int, str]] = []
    for match in qid_pattern.finditer(text):
        starts.append((match.start(1), match.group(1).upper()))
    blocks: dict[str, str] = {}
    for index, (start, qid) in enumerate(starts):
        end = starts[index + 1][0] if index + 1 < len(starts) else len(text)
        blocks.setdefault(qid, text[start:end])
    return blocks


def qid_candidates(var_name: str) -> list[str]:
    clean = var_name[1:] if var_name.lower().startswith("v") else var_name
    clean = re.sub(r"o\d+$", "", clean, flags=re.IGNORECASE)
    candidates = [clean.upper()]
    if clean.upper().startswith("CK"):
        candidates.append(clean[2:].upper())
    if clean.upper().endswith("R"):
        candidates.append(clean[:-1].upper())
    return list(dict.fromkeys(candidates))


def find_block(var_name: str, blocks: dict[str, str]) -> str:
    for qid in qid_candidates(var_name):
        if qid in blocks:
            return blocks[qid]
    return ""


def parse_r_values(row: dict[str, str]) -> list[str]:
    return [row.get(col, "").strip() for col in ("r1", "r2", "r3", "r4") if row.get(col, "").strip()]


def range_atom(var_name: str, value: str) -> str:
    parts = [part.strip() for part in value.split(",") if part.strip()]
    if len(parts) == 1:
        return f"range({var_name},{parts[0]},{parts[0]})"
    if len(parts) == 2:
        return f"range({var_name},{parts[0]},{parts[1]})"
    return f"any({var_name},{','.join(parts)})"


def proposed_in_or_not(values: list[str]) -> str:
    if not values:
        return ""
    if all("," not in value for value in values):
        return "any"
    return "range"


def proposed_range_expression(var_name: str, values: list[str], extra_check: str = "", decimal_check: str = "") -> str:
    mode = proposed_in_or_not(values)
    if mode == "any":
        base = f"any({var_name},{','.join(values)})"
    elif mode == "range":
        args: list[str] = []
        for value in values:
            parts = [part.strip() for part in value.split(",") if part.strip()]
            if len(parts) == 1:
                args.extend([parts[0], parts[0]])
            elif len(parts) == 2:
                args.extend(parts)
            else:
                return f"any({var_name},{','.join(parts)})"
        base = f"range({var_name},{','.join(args)})"
    else:
        base = ""
    if decimal_check and base:
        base = f"({base} & {decimal_check})"
    if extra_check and base:
        return f"{base}{extra_check}"
    if extra_check:
        return extra_check.lstrip("|&")
    return base


def alternative_range_expression(var_name: str, values: list[str]) -> str:
    atoms = [range_atom(var_name, value) for value in values]
    if not atoms:
        return ""
    elif len(atoms) == 1:
        return atoms[0]
    else:
        return "(" + " | ".join(atoms) + ")"


def combine_extra_checks(existing_extra: str, decimal_check: str) -> str:
    existing = (existing_extra or "").strip()
    decimal = (decimal_check or "").strip()
    if decimal and not existing:
        return f"&{decimal}"
    return existing


def decimal_check_for_block(var_name: str, block: str) -> tuple[str, str]:
    normalized = normalize_text(block)
    if not normalized:
        return "", ""
    match = re.search(r"可輸入至?小數點\s*([0-9一二兩三四五六七八九十]+)\s*位", normalized)
    if match:
        raw = match.group(1)
        digits = {"一": 1, "二": 2, "兩": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
        places = int(raw) if raw.isdigit() else digits.get(raw, 0)
        if places > 0:
            return f"mod({var_name}*{10 ** places},1)=0", f"allow up to {places} decimal places"
    if re.search(r"只能.*(?:\.5|0\.5|半)", normalized):
        return f"mod({var_name}*2,1)=0", "allow integer or .5 only"
    return "", ""


def load_rows(workbook_path: Path, docx_path: Path) -> tuple[list[dict[str, str]], dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[numeric_sheet_name(wb)]
    h = read_headers(ws)
    blocks = question_blocks(docx_text(docx_path))
    results: list[dict[str, str]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not cell_text(row[h["var"]]):
            continue
        record = {name: cell_text(row[idx]) if idx < len(row) else "" for name, idx in h.items()}
        var = record["var"]
        values = parse_r_values(record)
        block = find_block(var, blocks)
        decimal_check, decimal_note = decimal_check_for_block(var, block)
        existing_decimal = record.get("可輸入小數點", record.get("額外條件", ""))
        proposed_extra = combine_extra_checks(existing_decimal, decimal_check)
        proposed_range = proposed_range_expression(var, values, existing_decimal, decimal_check)
        proposed_alternative = alternative_range_expression(var, values)
        result = {
            "row": str(row_index),
            "var": var,
            "existing_in_or_not": record.get("in_or_not", ""),
            "proposed_in_or_not": proposed_in_or_not(values),
            "existing_range": record.get("range", ""),
            "proposed_range": proposed_range,
            "alternative_range": proposed_alternative,
            "existing_decimal_check": existing_decimal,
            "proposed_decimal_check": proposed_extra,
            "decimal_note": decimal_note,
            "status": "match",
            "mismatch_reason": "",
            "question_excerpt": block[:360],
        }
        mismatches = []
        if result["existing_in_or_not"].strip().lower() != result["proposed_in_or_not"].strip().lower():
            mismatches.append("in_or_not differs")
        if normalize_space(result["existing_range"]) != normalize_space(result["proposed_range"]):
            mismatches.append("range differs")
        if decimal_check and normalize_space(existing_decimal) != normalize_space(decimal_check):
            mismatches.append("decimal check missing/differs")
        if mismatches:
            result["status"] = "review"
            result["mismatch_reason"] = "; ".join(mismatches)
        results.append(result)
    counts: dict[str, int] = {}
    for row in results:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    return results, {"rows": len(results), "status": counts}


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", "", value or "").lower()


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "row",
        "var",
        "status",
        "mismatch_reason",
        "existing_in_or_not",
        "proposed_in_or_not",
        "existing_range",
        "proposed_range",
        "alternative_range",
        "existing_decimal_check",
        "proposed_decimal_check",
        "decimal_note",
        "question_excerpt",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def main() -> int:
    parser = argparse.ArgumentParser(description="Propose numeric Excel fields used by SPSS mail-merge formulas.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    rows, summary = load_rows(args.workbook, args.docx)
    write_csv(args.output_csv, rows)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(
        json.dumps({"summary": summary, "review": [row for row in rows if row["status"] != "match"]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(summary)
    return 0


if __name__ == "__main__":
    sys.exit(main())
