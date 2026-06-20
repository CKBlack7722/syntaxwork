from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import openpyxl


OPEN_SHEET_PREFIX = "開放欄位"
CITY_SUFFIX = "city"


@dataclass(frozen=True)
class OpenRecord:
    m: str
    p: str
    var: str
    kind: str
    width: str
    is_multi: str
    qid: str
    qvar: str
    option_value: str
    var2_new: str
    range_new: str
    n: str


@dataclass(frozen=True)
class OpenProposal:
    var: str
    qid: str
    qvar: str
    option_value: str
    is_multi: str
    confidence: str
    source: str
    note: str
    matched_existing_var: str
    status: str
    mismatch_reason: str
    question_excerpt: str


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(text: str) -> str:
    return unicodedata.normalize("NFKC", text or "").replace("（", "(").replace("）", ")")


def open_sheet_name(wb: openpyxl.Workbook) -> str:
    matches = [name for name in wb.sheetnames if name.startswith(OPEN_SHEET_PREFIX)]
    if not matches:
        raise ValueError("No open-field sheet found")
    return matches[0]


def read_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(value).strip(): idx for idx, value in enumerate(values) if value not in (None, "")}


def infer_open_formula_values(var_name: str, is_multi: str) -> dict[str, str]:
    if "city" in var_name.lower():
        qid = re.sub(r"_oth$", "", var_name, flags=re.IGNORECASE)
        return {"qid": qid, "qvar": qid, "option_value": "29", "var2_new": qid, "range_new": "29", "n": "3"}
    marker_index = var_name.rfind("o")
    option_value = var_name[marker_index + 1 :] if marker_index > 0 else ""
    if marker_index <= 0 or not option_value.isdigit():
        return {"qid": var_name, "qvar": "", "option_value": "", "var2_new": "", "range_new": "", "n": ""}
    qid = var_name[:marker_index]
    qvar = f"{qid}m{int(option_value)}" if is_multi == "1" else qid
    var2_new = qvar
    range_new = "1" if is_multi == "1" else str(int(option_value))
    option_value = str(int(option_value))
    n = str(len(option_value) + 1) if len(option_value) == 1 else str(len(option_value))
    return {
        "qid": qid,
        "qvar": qvar,
        "option_value": option_value,
        "var2_new": var2_new,
        "range_new": range_new,
        "n": n,
    }


def read_open_records(wb: openpyxl.Workbook) -> list[OpenRecord]:
    ws = wb[open_sheet_name(wb)]
    headers = read_headers(ws)
    required = ["m", "p", "var", "變項屬性", "寬度", "是否複選", "題號", "題號變項名稱", "選項數值", "var2_new", "range_new", "n"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Open-field sheet missing headers: {missing}")

    records: list[OpenRecord] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not cell_text(row[headers["var"]]):
            continue
        var_name = cell_text(row[headers["var"]])
        is_multi = cell_text(row[headers["是否複選"]])
        inferred = infer_open_formula_values(var_name, is_multi)
        records.append(
            OpenRecord(
                m=cell_text(row[headers["m"]]),
                p=cell_text(row[headers["p"]]),
                var=var_name,
                kind=cell_text(row[headers["變項屬性"]]),
                width=cell_text(row[headers["寬度"]]),
                is_multi=is_multi,
                qid=cell_text(row[headers["題號"]]) or inferred["qid"],
                qvar=cell_text(row[headers["題號變項名稱"]]) or inferred["qvar"],
                option_value=cell_text(row[headers["選項數值"]]) or inferred["option_value"],
                var2_new=cell_text(row[headers["var2_new"]]) or inferred["var2_new"],
                range_new=cell_text(row[headers["range_new"]]) or inferred["range_new"],
                n=cell_text(row[headers["n"]]) or inferred["n"],
            )
        )
    return records


def docx_text(docx_path: Path) -> str:
    parts: list[str] = []
    with ZipFile(docx_path) as archive:
        for name in archive.namelist():
            if not (name.startswith("word/") and name.endswith(".xml")):
                continue
            try:
                root = ET.fromstring(archive.read(name))
            except ET.ParseError:
                continue
            for element in root.iter():
                if element.tag.endswith("}t") and element.text:
                    parts.append(element.text)
    return normalize_text("".join(parts))


def normalize_qid(raw: str) -> str:
    return normalize_text(raw).replace(" ", "").upper()


def question_blocks(text: str) -> dict[str, str]:
    qid_pattern = re.compile(
        r"(?<![A-Z0-9(第])([A-Z][A-Z0-9_]{0,12}[0-9][A-Z0-9_]*|[0-9]{1,3}(?:_[0-9]+)?)(?:\s+|(?=請問|一般|為了|承上題|您上))(?=[\u4e00-\u9fffA-Z])"
    )
    starts: list[tuple[int, str]] = []
    for match in qid_pattern.finditer(text):
        qid = normalize_qid(match.group(1))
        if qid in {"7", "11", "100", "200", "500"}:
            continue
        starts.append((match.start(1), qid))

    blocks: dict[str, str] = {}
    for index, (start, qid) in enumerate(starts):
        end = starts[index + 1][0] if index + 1 < len(starts) else len(text)
        block = text[start:end]
        if len(block) >= 12:
            blocks.setdefault(qid, block)
    return blocks


def qid_to_var_prefix(qid: str) -> str:
    return "v" + qid


def clean_option_code(code: str) -> str:
    normalized = normalize_text(code).strip()
    if normalized.isdigit():
        return normalized.zfill(2) if len(normalized) > 1 and normalized.startswith("0") else str(int(normalized))
    return normalized


def option_var_suffix(code: str) -> str:
    normalized = normalize_text(code).strip()
    if normalized.isdigit() and len(normalized) >= 2 and normalized.startswith("0"):
        return normalized
    if normalized.isdigit():
        return str(int(normalized))
    return normalized.lower()


def extract_option_candidates(qid: str, block: str, multi_qids: set[str]) -> list[OpenProposal]:
    proposals: list[OpenProposal] = []
    option_pattern = re.compile(r"\((\d{1,2}|[A-Za-z]{1,2})\)\s*")
    matches = list(option_pattern.finditer(block))
    for index, match in enumerate(matches):
        code_raw = match.group(1)
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(block)
        option_text = block[start:end]
        if not (
            re.search(r"請說明|請填|請註明|說明百分比", option_text)
            or re.search(r"(?:其他|其它)[^()]{0,30}_{2,}", option_text)
        ):
            continue

        code_for_var = option_var_suffix(code_raw)
        option_value = clean_option_code(code_raw)
        is_multi = "1" if qid in multi_qids else "0"
        var = f"{qid_to_var_prefix(qid)}o{code_for_var}"
        qvar = f"{qid_to_var_prefix(qid)}m{option_value}" if is_multi == "1" else qid_to_var_prefix(qid)
        proposals.append(
            OpenProposal(
                var=var,
                qid=qid,
                qvar=qvar,
                option_value=option_value,
                is_multi=is_multi,
                confidence="high",
                source="option text requires explanation",
                note="",
                matched_existing_var="",
                status="",
                mismatch_reason="",
                question_excerpt=block[:360],
            )
        )
    return proposals


def extract_structured_open_candidates(qid: str, block: str) -> list[OpenProposal]:
    proposals: list[OpenProposal] = []
    if "工作是什麼" not in block:
        return proposals
    option_pattern = re.compile(r"\((\d{1,2})\)\s*")
    matches = list(option_pattern.finditer(block))
    if len(matches) < 2:
        return proposals
    for index, match in enumerate(matches):
        code = int(match.group(1))
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(block)
        option_text = block[start:end]
        if "___" not in option_text or re.search(r"其他|其它|請說明|說明百分比", option_text):
            continue
        var = f"{qid_to_var_prefix(qid)}s{code}"
        proposals.append(
            OpenProposal(
                var=var,
                qid=qid,
                qvar=var,
                option_value=str(code),
                is_multi="0",
                confidence="medium",
                source="structured open subfield",
                note="open text subfield inferred from underscored option text",
                matched_existing_var="",
                status="",
                mismatch_reason="",
                question_excerpt=block[:360],
            )
        )
    return proposals


def extract_unnumbered_other_candidate(qid: str, block: str) -> list[OpenProposal]:
    if not re.search(r"其他[_\s]*", block):
        return []
    if re.search(r"\(\d{1,2}\).*其他", block):
        return []
    option_count = len(re.findall(r"【(?:續答|跳答)", block))
    if option_count <= 0:
        return []
    option_value = str(option_count)
    return [
        OpenProposal(
            var=f"{qid_to_var_prefix(qid)}o{option_value}",
            qid=qid,
            qvar=qid_to_var_prefix(qid),
            option_value=option_value,
            is_multi="0",
            confidence="medium",
            source="unnumbered other option",
            note="option code inferred by counting answer choices with skip/continue markers",
            matched_existing_var="",
            status="",
            mismatch_reason="",
            question_excerpt=block[:360],
        )
    ]


def extract_city_candidates(blocks: dict[str, str]) -> list[OpenProposal]:
    proposals: list[OpenProposal] = []
    for qid, block in blocks.items():
        if "目前居住" not in block and "哪個縣市" not in block:
            continue
        var_base = qid_to_var_prefix(qid) + CITY_SUFFIX
        proposals.append(
            OpenProposal(
                var=f"{var_base}_oth",
                qid=qid,
                qvar=var_base,
                option_value="29",
                is_multi="0",
                confidence="medium",
                source="city other field",
                note="city/town open field inferred from location wording; verify project code convention",
                matched_existing_var="",
                status="",
                mismatch_reason="",
                question_excerpt=block[:360],
            )
        )
    return proposals


def read_multi_qids(wb: openpyxl.Workbook) -> set[str]:
    result: set[str] = set()
    sheet_name = "複選題變項清單"
    if sheet_name not in wb.sheetnames:
        return result
    ws = wb[sheet_name]
    headers = read_headers(ws)
    var_header = "變項名稱" if "變項名稱" in headers else None
    if not var_header:
        return result
    for row in ws.iter_rows(min_row=2, values_only=True):
        var = cell_text(row[headers[var_header]])
        match = re.match(r"v?([A-Z]*[0-9]+(?:_[0-9]+)?)[mo]\d+", var, flags=re.IGNORECASE)
        if match:
            result.add(normalize_qid(match.group(1)))
    return result


def merge_and_compare(existing: list[OpenRecord], proposed: list[OpenProposal]) -> list[OpenProposal]:
    existing_by_var = {record.var.lower(): record for record in existing}
    used: set[str] = set()
    result: list[OpenProposal] = []

    for proposal in proposed:
        proposal = remap_confirmation_open_var(proposal, existing_by_var)
        record = existing_by_var.get(proposal.var.lower())
        if not record:
            result.append(
                proposal.__class__(
                    **{**proposal.__dict__, "status": "extra_proposed", "mismatch_reason": "Word suggests an open field not listed in Excel"}
                )
            )
            continue
        used.add(record.var.lower())
        mismatches = []
        if record.kind != "字串":
            mismatches.append("變項屬性 is not 字串")
        if record.width != "150":
            mismatches.append("寬度 is not 150")
        if record.is_multi != proposal.is_multi:
            mismatches.append("是否複選 differs")
        if proposal.source in {"city other field", "structured open subfield"}:
            pass
        elif record.var2_new.lower() != proposal.qvar.lower():
            mismatches.append("var2_new differs")
        if proposal.source in {"city other field", "structured open subfield"}:
            pass
        elif record.option_value != proposal.option_value:
            mismatches.append("選項數值 differs")
        status = "match" if not mismatches else "mismatch"
        result.append(
            proposal.__class__(
                **{
                    **proposal.__dict__,
                    "matched_existing_var": record.var,
                    "status": status,
                    "mismatch_reason": "; ".join(mismatches),
                }
            )
        )

    for record in existing:
        if record.var.lower() in used:
            continue
        result.append(
            OpenProposal(
                var=record.var,
                qid=record.qid,
                qvar=record.qvar,
                option_value=record.option_value,
                is_multi=record.is_multi,
                confidence="manual",
                source="existing Excel row",
                note="No automatic Word match; needs parent-question mapping or manual review",
                matched_existing_var=record.var,
                status="missing_in_word",
                mismatch_reason="Existing Excel open field was not inferred from Word",
                question_excerpt="",
            )
        )
    return result


def remap_confirmation_open_var(proposal: OpenProposal, existing_by_var: dict[str, OpenRecord]) -> OpenProposal:
    if proposal.var.lower() in existing_by_var:
        return proposal
    match = re.match(r"^(v)(.+?)(o[0-9A-Za-z]+)$", proposal.var, flags=re.IGNORECASE)
    if not match:
        return proposal
    prefix, base, suffix = match.groups()
    candidates = [f"{prefix}{base}R{suffix}"]
    if base.upper().startswith("Q"):
        candidates.append(f"{prefix}CK{base[1:]}{suffix}")
    for candidate in candidates:
        record = existing_by_var.get(candidate.lower())
        if not record:
            continue
        qvar = record.var2_new if record.var2_new and record.var2_new != "#VALUE!" else re.sub(r"o[0-9A-Za-z]+$", "", candidate, flags=re.IGNORECASE)
        return OpenProposal(
            var=record.var,
            qid=record.qid,
            qvar=qvar,
            option_value=proposal.option_value,
            is_multi=proposal.is_multi,
            confidence="medium",
            source="confirmation follow-up open field",
            note=f"remapped from {proposal.var} because Excel stores the open field on confirmation variable {record.var}",
            matched_existing_var="",
            status="",
            mismatch_reason="",
            question_excerpt=proposal.question_excerpt,
        )
    return proposal


def generate(workbook_path: Path, docx_path: Path) -> tuple[list[OpenProposal], dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    existing = read_open_records(wb)
    text = docx_text(docx_path)
    blocks = question_blocks(text)
    multi_qids = read_multi_qids(wb)

    proposed: list[OpenProposal] = []
    for qid, block in blocks.items():
        proposed.extend(extract_option_candidates(qid, block, multi_qids))
        proposed.extend(extract_structured_open_candidates(qid, block))
        proposed.extend(extract_unnumbered_other_candidate(qid, block))
    proposed.extend(extract_city_candidates(blocks))

    deduped: dict[str, OpenProposal] = {}
    for proposal in proposed:
        deduped.setdefault(proposal.var.lower(), proposal)
    compared = merge_and_compare(existing, list(deduped.values()))
    counts: dict[str, int] = {}
    for item in compared:
        counts[item.status] = counts.get(item.status, 0) + 1
    return compared, {
        "existing_rows": len(existing),
        "proposed_rows": len(deduped),
        "status": counts,
        "question_blocks": len(blocks),
    }


def write_csv(path: Path, rows: list[OpenProposal]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "var",
                "qid",
                "qvar",
                "option_value",
                "is_multi",
                "confidence",
                "source",
                "status",
                "mismatch_reason",
                "note",
                "matched_existing_var",
                "question_excerpt",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.var,
                    row.qid,
                    row.qvar,
                    row.option_value,
                    row.is_multi,
                    row.confidence,
                    row.source,
                    row.status,
                    row.mismatch_reason,
                    row.note,
                    row.matched_existing_var,
                    row.question_excerpt,
                ]
            )


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate proposed open-field rows from Word questionnaire and compare with Excel.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--docx", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    rows, summary = generate(args.workbook, args.docx)
    write_csv(args.output_csv, rows)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(
        json.dumps(
            {
                "summary": summary,
                "review": [row.__dict__ for row in rows if row.status != "match"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"rows: {len(rows)}")
    print(f"summary: {summary}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
