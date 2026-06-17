from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class VarSpec:
    name: str
    kind: str
    width: int

    @property
    def is_numeric(self) -> bool:
        return self.kind == "數值"


@dataclass(frozen=True)
class OpenRule:
    row: int
    m: str
    p: str
    s: str
    s_value: str
    var: str
    width: int
    is_multi: str
    parent_var: str
    range_value: str
    parent_width: int


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sheet_by_prefix(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return wb[name]
    raise ValueError(f"missing sheet prefix {prefix}")


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        cell_text(value): idx
        for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)
        if cell_text(value)
    }


def resolved_mp(value: Any, row_values: dict[str, Any]) -> str:
    text = cell_text(value)
    if text.startswith("="):
        return cell_text(row_values.get("m"))
    return text


def resolved_optional_id(value: Any, row_values: dict[str, Any], fallback_key: str = "") -> str:
    text = cell_text(value)
    if text.startswith("="):
        return cell_text(row_values.get(fallback_key)) if fallback_key else ""
    return text


def load_specs(workbook_path: Path) -> dict[str, VarSpec]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["all"]
    specs: dict[str, VarSpec] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = cell_text(row[0] if len(row) > 0 else "")
        if not name:
            continue
        try:
            width = int(cell_text(row[2] if len(row) > 2 else "") or "2")
        except ValueError:
            width = 2
        specs[name] = VarSpec(name, cell_text(row[1] if len(row) > 1 else ""), width)
    return specs


def infer_open_qid(var_name: str) -> str:
    if var_name.lower().endswith("_oth"):
        return re.sub(r"_oth$", "", var_name, flags=re.IGNORECASE)
    return re.sub(r"o[0-9A-Za-z]+$", "", var_name, flags=re.IGNORECASE)


def infer_open_option(var_name: str, qid: str) -> str:
    if var_name.lower().endswith("_oth"):
        return ""
    match = re.search(r"o([0-9A-Za-z]+)$", var_name[len(qid) :], flags=re.IGNORECASE)
    if not match:
        return ""
    value = match.group(1)
    return str(int(value)) if value.isdigit() else value


def inferred_values(var_name: str, is_multi: str) -> tuple[str, str]:
    match = re.fullmatch(r"v71_2s[1-4]", var_name, flags=re.IGNORECASE)
    if match:
        return "v71_1", "1,10"
    if var_name.lower().endswith("city_oth"):
        return re.sub(r"_oth$", "", var_name, flags=re.IGNORECASE), "29"
    qid = infer_open_qid(var_name)
    option = infer_open_option(var_name, qid)
    if is_multi == "1":
        return (f"{qid}m{option}" if option else qid, "1")
    return qid, option


def load_rules(workbook_path: Path) -> tuple[list[OpenRule], list[dict[str, Any]]]:
    specs = load_specs(workbook_path)
    wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    ws = sheet_by_prefix(wb, "開放欄位")
    h = headers(ws)
    required = ["m", "p", "var", "變項屬性", "寬度", "是否複選", "var2_new", "range_new"]
    missing = [name for name in required if name not in h]
    if missing:
        raise ValueError(f"open sheet missing headers: {missing}")
    rules: list[OpenRule] = []
    skipped: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = {name: row[idx - 1] for name, idx in h.items()}
        var = cell_text(values.get("var"))
        if not var:
            continue
        if cell_text(values.get("變項屬性")) != "字串":
            continue
        m = cell_text(values.get("m"))
        p = resolved_mp(values.get("p"), values)
        s = resolved_optional_id(values.get("s"), values, "m") if "s" in h else ""
        s_value = resolved_optional_id(values.get("s="), values) if "s=" in h else ""
        if not m or not p:
            skipped.append({"row": row_idx, "var": var, "reason": "blank m or p"})
            continue
        is_multi = cell_text(values.get("是否複選")) or "0"
        inferred_parent, inferred_range = inferred_values(var, is_multi)
        parent_var = cell_text(values.get("var2_new"))
        range_value = cell_text(values.get("range_new"))
        if not parent_var or parent_var.startswith("="):
            parent_var = inferred_parent
        if not range_value or range_value.startswith("="):
            range_value = inferred_range
        if not parent_var or not range_value:
            skipped.append({"row": row_idx, "var": var, "reason": "cannot infer parent/range"})
            continue
        try:
            width = int(cell_text(values.get("寬度")) or "150")
        except ValueError:
            width = 150
        parent_width = specs.get(parent_var, VarSpec(parent_var, "數值", 2)).width
        rules.append(OpenRule(row_idx, m, p, s, s_value, var, width, is_multi, parent_var, range_value, parent_width))
    return rules, skipped


def parent_expr(rule: OpenRule) -> str:
    if "," in rule.range_value:
        return f"range({rule.parent_var},{rule.range_value})"
    return f"{rule.parent_var}={rule.range_value}"


def parent_not_expr(rule: OpenRule) -> str:
    if "," in rule.range_value:
        return f"not range({rule.parent_var},{rule.range_value})"
    return f"{rule.parent_var}~={rule.range_value}"


def render_message(rule: OpenRule, text: str) -> str:
    return f'compute p{rule.p}="{rule.var}{text}".'


def render_s(rule: OpenRule) -> str:
    return f"compute s{rule.s}={rule.s_value}." if rule.s and rule.s_value else ""


def render_m(rule: OpenRule) -> str:
    show_width = min(rule.width, 150)
    return (
        f'Compute m{rule.m}=concat("{rule.parent_var}=",string({rule.parent_var},n{rule.parent_width}),'
        f'";{rule.var}=",char.substr({rule.var},1,{show_width})).'
    )


def render_rule(rule: OpenRule) -> str:
    return "\n".join(
        [
            f"*{rule.var} 開放欄位檢核 是否為複選題={rule.is_multi}.",
            f'do if {parent_expr(rule)} & {rule.var}="".',
            render_m(rule),
            render_message(rule, "開放欄位應答而未答"),
            *([render_s(rule)] if render_s(rule) else []),
            f'else if {parent_not_expr(rule)} & {rule.var}~="".',
            render_m(rule),
            render_message(rule, "開放欄位不該答而答"),
            *([render_s(rule)] if render_s(rule) else []),
            f'else if {parent_expr(rule)} & {rule.var}~="" & range(keyin,keyindate1, Keyindate2).',
            render_m(rule),
            render_message(rule, "開放欄位內容列出確認"),
            *([render_s(rule)] if render_s(rule) else []),
            "end if.",
            "Exec.",
            "",
        ]
    )


def render_spss(rules: list[OpenRule]) -> str:
    parts = [
        "* Encoding: UTF-8.",
        "**OPEN FIELD CHECKS.",
        "* SYNTAXWORK_BEGIN_OPEN.",
    ]
    parts.extend(render_rule(rule).rstrip() for rule in rules)
    parts.append("* SYNTAXWORK_END_OPEN.")
    return "\n\n".join(parts).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    rules, skipped = load_rules(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(render_spss(rules), encoding="utf-8-sig")
    report = {"rules": len(rules), "skipped": len(skipped), "skipped_rows": skipped, "output": str(args.output)}
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
