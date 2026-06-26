from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


MISSING_CODES = "96,996,999996,9999996,9999999996,99969696969696"


@dataclass(frozen=True)
class VarSpec:
    name: str
    kind: str
    width: int

    @property
    def is_numeric(self) -> bool:
        return self.kind == "數值"


@dataclass(frozen=True)
class LogicRule:
    m: str
    p: str
    s: str
    s_value: str
    condition_raw: str
    condition_spss: str
    required_vars: tuple[str, ...]
    forbidden_vars: tuple[str, ...]
    limits: tuple[str, ...]
    mutex_expr: str


@dataclass(frozen=True)
class LogicSheetNormalization:
    corrected: tuple[dict[str, object], ...]
    unresolved: tuple[dict[str, object], ...]


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def optional_row_value(row: tuple[object, ...], headers: dict[str, int], name: str, fallback_name: str = "") -> str:
    if name not in headers or headers[name] >= len(row):
        return ""
    text = cell_text(row[headers[name]])
    if text.startswith("="):
        if fallback_name and fallback_name in headers and headers[fallback_name] < len(row):
            return cell_text(row[headers[fallback_name]])
        return ""
    return text


def load_var_specs(workbook_path: Path) -> dict[str, VarSpec]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["all"]
    specs: dict[str, VarSpec] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, kind, width = row[:3]
        if not name:
            continue
        try:
            width_int = int(width)
        except (TypeError, ValueError):
            width_int = 2
        specs[str(name).strip()] = VarSpec(str(name).strip(), str(kind).strip(), width_int)
    return specs


RESERVED_WORDS = {"and", "or", "not", "any", "range", "sys", "in", "mod", "trunc", "minute_tens"}


def normalize_var_name(name: str, specs: dict[str, VarSpec]) -> tuple[str, bool, bool]:
    clean = name.strip()
    if not clean or clean.lower() in RESERVED_WORDS or clean.isdigit():
        return clean, False, True
    if clean in specs:
        return clean, False, True
    with_v = f"v{clean}"
    if with_v in specs:
        return with_v, True, True
    return clean, False, False


def normalize_condition_text(condition: str, specs: dict[str, VarSpec]) -> tuple[str, list[dict[str, object]]]:
    unresolved: list[dict[str, object]] = []

    def replace_token(match: re.Match[str]) -> str:
        token = match.group(0)
        normalized, changed, found = normalize_var_name(token, specs)
        if not found:
            unresolved.append({"token": token, "context": "條件"})
        return normalized if changed else token

    normalized = re.sub(r"\b[A-Za-z_][A-Za-z0-9_]*\b", replace_token, condition)
    return normalized, unresolved


def normalize_var_list_text(value: str, specs: dict[str, VarSpec], context: str) -> tuple[str, list[dict[str, object]]]:
    normalized_vars: list[str] = []
    unresolved: list[dict[str, object]] = []
    for var_name in split_vars(value):
        normalized, _changed, found = normalize_var_name(var_name, specs)
        if found:
            normalized_vars.append(normalized)
        else:
            unresolved.append({"token": var_name, "context": context})
            normalized_vars.append(var_name)
    return ",".join(normalized_vars), unresolved


def normalize_logic_sheet_variables(workbook_path: Path, specs: dict[str, VarSpec]) -> LogicSheetNormalization:
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["邏輯組"]
    corrected: list[dict[str, object]] = []
    unresolved: list[dict[str, object]] = []

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, context in [(3, "條件"), (4, "應答"), (5, "不應答"), (6, "限制")]:
            cell = ws.cell(row_idx, col_idx)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            original = str(cell.value).strip()
            if col_idx in (3, 6):
                normalized, cell_unresolved = normalize_condition_text(original, specs)
            else:
                normalized, cell_unresolved = normalize_var_list_text(original, specs, context)

            for item in cell_unresolved:
                unresolved.append({"row": row_idx, "column": context, **item})

            if normalized != original:
                cell.value = normalized
                corrected.append(
                    {
                        "row": row_idx,
                        "column": context,
                        "before": original,
                        "after": normalized,
                    }
                )

    if corrected:
        wb.save(workbook_path)

    return LogicSheetNormalization(tuple(corrected), tuple(unresolved))


def split_vars(value: object) -> tuple[str, ...]:
    if value is None:
        return ()
    return tuple(v.strip() for v in str(value).split(",") if v and v.strip())


def expand_values(text: str) -> list[str]:
    values: list[str] = []
    for part in [p.strip() for p in text.split(",") if p.strip()]:
        if "~" in part:
            start, end = [x.strip() for x in part.split("~", 1)]
            if start.lstrip("-").isdigit() and end.lstrip("-").isdigit():
                step = 1 if int(start) <= int(end) else -1
                values.extend(str(x) for x in range(int(start), int(end) + step, step))
            else:
                values.append(part)
        else:
            values.append(part)
    return values


def convert_condition(raw: str) -> str:
    condition = raw.strip()
    if condition.lower().startswith("spss:"):
        return condition[5:].strip()

    def replace_in(match: re.Match[str]) -> str:
        var_name = match.group("var")
        values = ",".join(expand_values(match.group("values")))
        return f"any({var_name},{values})"

    condition = re.sub(
        r"\b(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s+in\s+(?P<values>[0-9,\-~\s]+)",
        replace_in,
        condition,
        flags=re.IGNORECASE,
    )
    condition = compact_not_equals(condition)
    condition = compact_negative_any(condition)
    condition = compact_positive_any(condition)
    return condition


def compact_not_equals(condition: str) -> str:
    """Collapse `v~=1 & v~=2` chains to SPSS-friendly `not any(v,1,2)`."""
    term_pattern = re.compile(r"\b(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s*~=\s*(?P<value>-?\d+)\b")

    def compact_segment(segment: str) -> str:
        parts = [part.strip() for part in segment.split("&")]
        groups: list[tuple[str, list[str], list[int]]] = []
        for idx, part in enumerate(parts):
            match = term_pattern.fullmatch(part)
            if not match:
                continue
            var_name = match.group("var")
            value = match.group("value")
            for group_var, values, indexes in groups:
                if group_var == var_name:
                    values.append(value)
                    indexes.append(idx)
                    break
            else:
                groups.append((var_name, [value], [idx]))

        replacements: dict[int, str] = {}
        skip_indexes: set[int] = set()
        for var_name, values, indexes in groups:
            if len(indexes) < 2:
                continue
            replacements[indexes[0]] = f"not any({var_name},{','.join(values)})"
            skip_indexes.update(indexes[1:])

        if not replacements:
            return segment.strip()
        compacted: list[str] = []
        for idx, part in enumerate(parts):
            if idx in skip_indexes:
                continue
            compacted.append(replacements.get(idx, part))
        return " & ".join(compacted)

    # Compact each OR branch independently; `a~=1 & b | a~=2` must not cross `|`.
    branches = [compact_segment(branch) for branch in condition.split("|")]
    return " | ".join(branches)


def peel_parens(text: str) -> tuple[str, str, str]:
    stripped = text.strip()
    prefix = re.match(r"^\(+", stripped)
    suffix = re.search(r"\)+$", stripped)
    prefix_text = prefix.group(0) if prefix else ""
    suffix_text = suffix.group(0) if suffix else ""
    start = len(prefix_text)
    end = len(stripped) - len(suffix_text) if suffix_text else len(stripped)
    return prefix_text, stripped[start:end].strip(), suffix_text


def match_with_outer_parens(
    text: str, patterns: tuple[re.Pattern[str], ...]
) -> tuple[re.Match[str], str, str] | None:
    prefix = ""
    suffix = ""
    core = text.strip()
    for _ in range(8):
        for pattern in patterns:
            match = pattern.fullmatch(core)
            if match:
                return match, prefix, suffix
        changed = False
        if core.startswith("("):
            prefix += "("
            core = core[1:].strip()
            changed = True
        elif core.endswith(")"):
            suffix += ")"
            core = core[:-1].strip()
            changed = True
        if not changed:
            return None
    return None


def compact_negative_any(condition: str) -> str:
    """Merge same-variable negative checks inside each AND branch."""
    term_patterns = (
        re.compile(r"^(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s*~=\s*(?P<values>-?\d+)$", re.IGNORECASE),
        re.compile(r"^not\s+any\(\s*(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s*,\s*(?P<values>[-0-9,\s]+)\)$", re.IGNORECASE),
    )

    def parse_negative(part: str) -> tuple[str, list[str], str, str] | None:
        wrapped_match = match_with_outer_parens(part, term_patterns)
        if wrapped_match:
            match, prefix, suffix = wrapped_match
            values = [value.strip() for value in match.group("values").split(",") if value.strip()]
            return match.group("var"), values, prefix, suffix
        return None

    def compact_segment(segment: str) -> str:
        parts = [part.strip() for part in segment.split("&")]
        groups: dict[str, list[int]] = {}
        parsed: dict[int, tuple[str, list[str], str, str]] = {}
        for idx, part in enumerate(parts):
            parsed_part = parse_negative(part)
            if not parsed_part:
                continue
            var_name = parsed_part[0]
            parsed[idx] = parsed_part
            groups.setdefault(var_name, []).append(idx)

        replacements: dict[int, str] = {}
        skip_indexes: set[int] = set()
        for var_name, indexes in groups.items():
            if len(indexes) < 2:
                continue
            values: list[str] = []
            prefix = parsed[indexes[0]][2]
            suffix = parsed[indexes[0]][3]
            for idx in indexes:
                _, part_values, part_prefix, part_suffix = parsed[idx]
                for value in part_values:
                    if value not in values:
                        values.append(value)
                if idx != indexes[0]:
                    prefix = prefix or part_prefix
                    suffix = suffix + part_suffix
            replacements[indexes[0]] = f"{prefix}not any({var_name},{','.join(values)}){suffix}"
            skip_indexes.update(indexes[1:])

        if not replacements:
            return segment.strip()
        compacted: list[str] = []
        for idx, part in enumerate(parts):
            if idx in skip_indexes:
                continue
            compacted.append(replacements.get(idx, part))
        return " & ".join(compacted)

    return " | ".join(compact_segment(branch) for branch in condition.split("|"))


def compact_positive_any(condition: str) -> str:
    """Merge same-variable `any(var,value)` checks inside each OR branch."""
    term_pattern = re.compile(
        r"^any\(\s*(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s*,\s*(?P<values>[-0-9,\s]+)\)$",
        re.IGNORECASE,
    )

    def parse_positive(part: str) -> tuple[str, list[str], str, str] | None:
        wrapped_match = match_with_outer_parens(part, (term_pattern,))
        if not wrapped_match:
            return None
        match, prefix, suffix = wrapped_match
        values = [value.strip() for value in match.group("values").split(",") if value.strip()]
        return match.group("var"), values, prefix, suffix

    def compact_segment(segment: str) -> str:
        parts = [part.strip() for part in segment.split("|")]
        groups: dict[str, list[int]] = {}
        parsed: dict[int, tuple[str, list[str], str, str]] = {}
        for idx, part in enumerate(parts):
            parsed_part = parse_positive(part)
            if not parsed_part:
                continue
            var_name = parsed_part[0]
            parsed[idx] = parsed_part
            groups.setdefault(var_name, []).append(idx)

        replacements: dict[int, str] = {}
        skip_indexes: set[int] = set()
        for var_name, indexes in groups.items():
            if len(indexes) < 2:
                continue
            values: list[str] = []
            prefix = parsed[indexes[0]][2]
            suffix = parsed[indexes[0]][3]
            for idx in indexes:
                _, part_values, part_prefix, part_suffix = parsed[idx]
                for value in part_values:
                    if value not in values:
                        values.append(value)
                if idx != indexes[0]:
                    prefix = prefix or part_prefix
                    suffix = suffix + part_suffix
            replacements[indexes[0]] = f"{prefix}any({var_name},{','.join(values)}){suffix}"
            skip_indexes.update(indexes[1:])

        if not replacements:
            return segment.strip()
        compacted: list[str] = []
        for idx, part in enumerate(parts):
            if idx in skip_indexes:
                continue
            compacted.append(replacements.get(idx, part))
        return " | ".join(compacted)

    return " & ".join(compact_segment(branch) for branch in condition.split("&"))


def condition_vars(condition: str) -> tuple[str, ...]:
    names = re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", condition)
    result: list[str] = []
    for name in names:
        if name.lower() in RESERVED_WORDS:
            continue
        if name not in result:
            result.append(name)
    return tuple(result)


def load_logic_rules(workbook_path: Path) -> list[LogicRule]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["邏輯組"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    h = {cell_text(value): idx for idx, value in enumerate(header_row) if cell_text(value)}
    rules: list[LogicRule] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        m, p, condition, required, forbidden, limit = (row[:6] + (None,) * 6)[:6]
        mutex = optional_row_value(row, h, "互斥")
        if not m or not any(v not in (None, "") for v in [condition, required, forbidden, limit, mutex]):
            continue
        p_value = p if p not in (None, "") else m
        condition_raw = str(condition).strip() if condition not in (None, "") else ""
        rules.append(
            LogicRule(
                m=str(int(m)) if isinstance(m, float) else str(m).strip(),
                p=str(int(p_value)) if isinstance(p_value, float) else str(p_value).strip(),
                s=optional_row_value(row, h, "s", "m"),
                s_value=optional_row_value(row, h, "s="),
                condition_raw=condition_raw,
                condition_spss=convert_condition(condition_raw) if condition_raw else "",
                required_vars=split_vars(required),
                forbidden_vars=split_vars(forbidden),
                limits=tuple(part.strip() for part in str(limit).split(";") if part.strip()) if limit not in (None, "") else (),
                mutex_expr=convert_condition(mutex) if mutex else "",
            )
        )
    return rules


def spec_for(var_name: str, specs: dict[str, VarSpec]) -> VarSpec:
    return specs.get(var_name, VarSpec(var_name, "數值", 2))


def display_expr(var_name: str, specs: dict[str, VarSpec]) -> str:
    spec = spec_for(var_name, specs)
    if spec.is_numeric:
        return f'string({var_name},n{spec.width})'
    return f"rtrim(ltrim({var_name}))"


def concat_parts(vars_to_show: list[str], specs: dict[str, VarSpec]) -> list[str]:
    parts: list[str] = []
    for index, var_name in enumerate(vars_to_show):
        label = f'"{var_name}="'
        if index > 0:
            parts.append(f'",{var_name}="')
        else:
            parts.append(label)
        parts.append(display_expr(var_name, specs))
    return parts


def concat_expr_single_line(vars_to_show: list[str], specs: dict[str, VarSpec]) -> str:
    return f"concat({','.join(concat_parts(vars_to_show, specs))})"


def wrap_args(parts: list[str], indent: str = "  ", max_len: int = 160) -> list[str]:
    lines: list[str] = []
    current = indent
    for index, part in enumerate(parts):
        token = part + ("," if index < len(parts) - 1 else "")
        if current.strip() and len(current) + len(token) + 1 > max_len:
            lines.append(current.rstrip())
            current = indent + token
        else:
            current += ("" if current == indent else " ") + token
    if current.strip():
        lines.append(current.rstrip())
    return lines


def render_compute_m(m_id: str, vars_to_show: list[str], specs: dict[str, VarSpec]) -> list[str]:
    parts = concat_parts(vars_to_show, specs)
    single_line = f"compute m{m_id}=concat({','.join(parts)})."
    if len(single_line) <= 160:
        return [single_line]

    lines = [f"compute m{m_id}=concat("]
    lines.extend(wrap_args(parts))
    lines.append(").")
    return lines


def split_outside_quotes(text: str, separators: set[str]) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    in_quote = False
    for char in text:
        if char == '"':
            in_quote = not in_quote
        if not in_quote and char in separators:
            piece = "".join(current).strip()
            if piece:
                parts.append(piece)
            parts.append(char)
            current = []
        else:
            current.append(char)
    piece = "".join(current).strip()
    if piece:
        parts.append(piece)
    return parts


def wrap_expression(expr: str, first_prefix: str, final_suffix: str = ".", max_len: int = 160) -> list[str]:
    single = f"{first_prefix}{expr}{final_suffix}"
    if len(single) <= max_len:
        return [single]
    tokens = split_outside_quotes(expr, {"&", "|", ","})
    lines: list[str] = []
    current = first_prefix
    pending_op = ""
    for token in tokens:
        if token in {"&", "|", ","}:
            pending_op = token
            continue
        piece = f"{pending_op} {token}" if pending_op else token
        if current.strip() and len(current) + len(piece) + 1 > max_len:
            lines.append(current.rstrip())
            current = f"{pending_op} {token}" if pending_op else token
        else:
            current += ("" if current == first_prefix else " ") + piece
        pending_op = ""
    if current.strip():
        lines.append(current.rstrip() + final_suffix)
    return lines


def string_chunks(text: str, chunk_size: int = 100) -> list[str]:
    return [text[index : index + chunk_size] for index in range(0, len(text), chunk_size)] or [""]


def render_compute_p(p_id: str, message: str) -> list[str]:
    escaped = message.replace('"', "'")
    single = f'compute p{p_id}="{escaped}".'
    if len(single) <= 160:
        return [single]
    chunks = string_chunks(escaped)
    lines = [f"compute p{p_id}=concat("]
    for index, chunk in enumerate(chunks):
        suffix = "," if index < len(chunks) - 1 else ""
        lines.append(f'  "{chunk}"{suffix}')
    lines.append(").")
    return lines


def wrap_comment(text: str, max_len: int = 160) -> list[str]:
    if len(text) <= max_len:
        return [text]
    body = text[1:].rstrip(".").strip() if text.startswith("*") else text.rstrip(".").strip()
    parts = [part for part in body.split(",") if part]
    lines: list[str] = []
    current = "* "
    for index, part in enumerate(parts):
        token = part.strip() + ("," if index < len(parts) - 1 else "")
        if current.strip() != "*" and len(current) + len(token) + 1 > max_len:
            lines.append(current.rstrip().rstrip(",") + ".")
            current = "* " + token
        else:
            current += ("" if current == "* " else "") + token
    if current.strip() != "*":
        lines.append(current.rstrip().rstrip(",") + ".")
    return lines


def answer_check(var_name: str, expect_answer: bool, specs: dict[str, VarSpec]) -> str:
    spec = spec_for(var_name, specs)
    if spec.is_numeric:
        flag = "1" if expect_answer else "0"
        return f"any({flag},{var_name}_96)"
    op = "=" if expect_answer else "~="
    return f'{var_name}{op}"96"'


def grouped_answer_check(vars_to_check: tuple[str, ...], expect_answer: bool, specs: dict[str, VarSpec]) -> str:
    checks = [answer_check(v, expect_answer, specs) for v in vars_to_check]
    if len(checks) == 1:
        return checks[0]
    return "(" + " | ".join(checks) + ")"


def unique_preserve_order(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        if value not in result:
            result.append(value)
    return result


def render_rule(rule: LogicRule, specs: dict[str, VarSpec]) -> str:
    blocks: list[str] = []
    if rule.required_vars or rule.forbidden_vars:
        blocks.append(render_show_hide_rule(rule, specs))
    for limit in rule.limits:
        blocks.append(render_limit_rule(rule, limit, specs))
    if rule.mutex_expr:
        blocks.append(render_mutex_rule(rule, specs))
    return "\n".join(blocks)


def render_mutex_rule(rule: LogicRule, specs: dict[str, VarSpec]) -> str:
    condition = f"({rule.condition_spss})" if rule.condition_spss else ""
    predicate = f"{condition} & ({rule.mutex_expr})" if condition else f"({rule.mutex_expr})"
    vars_to_show = unique_preserve_order(list(condition_vars(rule.condition_spss)) + list(condition_vars(rule.mutex_expr)))
    message = f"{rule.condition_raw}與互斥條件不應同時成立"
    lines = wrap_comment("* logic check mutex.")
    lines.extend(wrap_expression(predicate, "do if "))
    lines.extend(render_compute_m(rule.m, vars_to_show, specs))
    lines.extend(render_compute_p(rule.p, message))
    lines.extend(
        [
            *([f"compute s{rule.s}={rule.s_value}."] if rule.s and rule.s_value else []),
            "end if.",
            "",
        ]
    )
    return "\n".join(lines)


def render_show_hide_rule(rule: LogicRule, specs: dict[str, VarSpec]) -> str:
    if rule.required_vars:
        target_vars = rule.required_vars
        expect_answer = True
        action = "show"
        message = f"{rule.condition_raw}，應答{','.join(target_vars)}而未答"
    else:
        target_vars = rule.forbidden_vars
        expect_answer = False
        action = "hide"
        message = f"{rule.condition_raw}，不應答{','.join(target_vars)}而答"

    vars_to_show = unique_preserve_order(list(condition_vars(rule.condition_spss)) + list(target_vars))
    check = grouped_answer_check(target_vars, expect_answer, specs)
    target_label = ",".join(target_vars)
    condition = f"({rule.condition_spss})" if rule.condition_spss else ""
    lines = wrap_comment(f"* logic check {action} {target_label}.")
    lines.extend(wrap_expression(f"{condition} & {check}" if condition else check, "do if "))
    lines.extend(render_compute_m(rule.m, vars_to_show, specs))
    lines.extend(render_compute_p(rule.p, message))
    lines.extend(
        [
            *([f"compute s{rule.s}={rule.s_value}."] if rule.s and rule.s_value else []),
            "end if.",
            "",
        ]
    )
    return "\n".join(lines)


def parse_limit(limit: str) -> tuple[str, str, str]:
    match = re.match(
        r"^(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s+minute_tens\s+(?P<op>not\s+in|in)\s+(?P<rhs>.+)$",
        limit,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group("var"), f"minute_tens {' '.join(match.group('op').lower().split())}", match.group("rhs").strip()

    match = re.match(
        r"^(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s+(?P<op>not\s+in|in)\s+(?P<rhs>.+)$",
        limit,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group("var"), " ".join(match.group("op").lower().split()), match.group("rhs").strip()

    match = re.match(
        r"^(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s*(?P<op>~=|>=|<=|=|>|<)\s*(?P<rhs>.+)$",
        limit,
    )
    if not match:
        raise ValueError(f"Unsupported limit syntax: {limit}")
    return match.group("var"), match.group("op"), match.group("rhs").strip()


def rhs_vars(rhs: str) -> tuple[str, ...]:
    vars_found: list[str] = []
    for token in re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", rhs):
        if token.lower() not in RESERVED_WORDS and token not in vars_found:
            vars_found.append(token)
    return tuple(vars_found)


def invert_limit_to_violation(var_name: str, op: str, rhs: str) -> str:
    if op == "minute_tens not in":
        return f"any(mod(trunc({var_name}/10),10),{','.join(expand_values(rhs))})"
    if op == "minute_tens in":
        return f"not any(mod(trunc({var_name}/10),10),{','.join(expand_values(rhs))})"
    if op == "not in":
        return f"any({var_name},{','.join(expand_values(rhs))})"
    if op == "in":
        return f"not any({var_name},{','.join(expand_values(rhs))})"
    inverse_ops = {
        "<=": ">",
        "<": ">=",
        ">=": "<",
        ">": "<=",
        "=": "~=",
        "~=": "=",
    }
    return f"{var_name}{inverse_ops[op]}{rhs}"


def limit_message(limit: str, condition_raw: str) -> str:
    prefix = f"{condition_raw}，" if condition_raw else ""
    var_name, op, rhs = parse_limit(limit)
    if op == "minute_tens not in":
        return f"{prefix}{var_name}第3碼不可為{rhs}"
    if op == "minute_tens in":
        return f"{prefix}{var_name}第3碼應為{rhs}"
    if op == "not in":
        return f"{prefix}{var_name}不可為{rhs}"
    if op == "in":
        return f"{prefix}{var_name}應為{rhs}"
    return f"{prefix}{var_name}應{op}{rhs}"


def render_limit_rule(rule: LogicRule, limit: str, specs: dict[str, VarSpec]) -> str:
    var_name, op, rhs = parse_limit(limit)
    violation = invert_limit_to_violation(var_name, op, rhs)
    condition = f"({rule.condition_spss})" if rule.condition_spss else ""
    vars_to_show = unique_preserve_order(list(condition_vars(rule.condition_spss)) + [var_name] + list(rhs_vars(rhs)))
    lines = wrap_comment(f"* logic check limit {var_name}.")
    lines.extend(wrap_expression(f"{condition} & {violation}" if condition else violation, "do if "))
    lines.extend(render_compute_m(rule.m, vars_to_show, specs))
    lines.extend(render_compute_p(rule.p, limit_message(limit, rule.condition_raw)))
    lines.extend(
        [
            *([f"compute s{rule.s}={rule.s_value}."] if rule.s and rule.s_value else []),
            "end if.",
            "",
        ]
    )
    return "\n".join(lines)


def render_spss(rules: list[LogicRule], specs: dict[str, VarSpec]) -> str:
    blocks = [
        "* Encoding: UTF-8.",
        "**四、邏輯檢核.",
        "* SYNTAXWORK_BEGIN_LOGIC.",
    ]
    blocks.extend(render_rule(rule, specs) for rule in rules)
    blocks.append("* SYNTAXWORK_END_LOGIC.")
    return "\n".join(blocks).rstrip() + "\n"


def extract_existing_logic(sps_path: Path) -> list[str] | None:
    text = sps_path.read_text(encoding="utf-8-sig")
    marker_match = re.search(
        r"\*\s*SYNTAXWORK_BEGIN_LOGIC\.\s*(.*?)\*\s*SYNTAXWORK_END_LOGIC\.",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if marker_match:
        section = marker_match.group(1)
        return [
            line.strip()
            for line in section.splitlines()
            if line.strip().lower().startswith("do if ") and "!i" not in line
        ]

    try:
        start = text.index("**4.邏輯檢核.")
    except ValueError:
        return None
    end_candidates = [
        pos
        for marker in ("*檢核項目清單.", "\n**5.", "\n**ss", "\n***************************************************************.")
        if (pos := text.find(marker, start + 1)) != -1
    ]
    end = min(end_candidates) if end_candidates else len(text)
    return [
        line.strip()
        for line in text[start:end].splitlines()
        if line.strip().lower().startswith("do if ") and "!i" not in line
    ]


def normalize_condition(condition: str) -> str:
    result = condition.strip().rstrip(".")
    result = re.sub(r"\s+", "", result)
    result = result.replace("~=", "!=")
    result = result.replace("= ", "=")
    result = re.sub(r"any\(([^,()]+),([0-9,]+)\)", lambda m: "any(" + m.group(1) + "," + ",".join(expand_values(m.group(2))) + ")", result)
    return result.lower()


def semantic_key_from_generated(rule: LogicRule, specs: dict[str, VarSpec]) -> dict[str, object]:
    target_vars = rule.required_vars or rule.forbidden_vars
    if target_vars:
        mode = "required" if rule.required_vars else "forbidden"
        check = grouped_answer_check(target_vars, bool(rule.required_vars), specs)
        predicate = f"{rule.condition_spss} & {check}"
    elif rule.mutex_expr:
        mode = "mutex"
        check = rule.mutex_expr
        predicate = f"{rule.condition_spss} & {rule.mutex_expr}" if rule.condition_spss else rule.mutex_expr
    else:
        mode = "limit"
        check = ""
        predicate = rule.condition_spss
    return {
        "m": rule.m,
        "condition": normalize_condition(rule.condition_spss),
        "mode": mode,
        "targets": list(target_vars),
        "limits": list(rule.limits),
        "mutex": normalize_condition(rule.mutex_expr),
        "check": normalize_condition(check),
        "predicate": normalize_condition(predicate),
    }


def existing_keys(lines: list[str]) -> list[str]:
    keys: list[str] = []
    for line in lines:
        expr = re.sub(r"^do if\s+", "", line, flags=re.I).rstrip(".")
        keys.append(normalize_condition(expr))
    return keys


def write_compare_report(
    report_path: Path,
    rules: list[LogicRule],
    specs: dict[str, VarSpec],
    existing_do_ifs: list[str] | None,
    normalization: LogicSheetNormalization,
) -> None:
    generated = [semantic_key_from_generated(rule, specs) for rule in rules]
    existing_normalized = existing_keys(existing_do_ifs) if existing_do_ifs is not None else []
    review = []
    for item in generated:
        review.append(
            {
                "m": item["m"],
                "condition": item["condition"],
                "mode": item["mode"],
                "targets": item["targets"],
                "expected_check": item["check"],
                "generated_predicate": item["predicate"],
                "matched_existing_do_if": item["predicate"] in existing_normalized,
            }
        )

    report = {
        "note": "Compare predicates only. Ignore m/p id differences and displayed mXXX field differences.",
        "generated_rule_count": len(generated),
        "existing_logic_do_if_count": len(existing_do_ifs) if existing_do_ifs is not None else None,
        "matched_count": sum(1 for item in review if item["matched_existing_do_if"]),
        "unmatched_count": sum(1 for item in review if not item["matched_existing_do_if"]),
        "generated_rules": generated,
        "logic_sheet_normalization": {
            "corrected_count": len(normalization.corrected),
            "unresolved_count": len(normalization.unresolved),
            "corrected": list(normalization.corrected),
            "unresolved": list(normalization.unresolved),
        },
        "existing_do_ifs": existing_do_ifs,
        "existing_normalized_do_ifs": existing_normalized,
        "review": review,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default="檢核程式套印-淨零.xlsx")
    parser.add_argument("--existing-sps", default=None)
    parser.add_argument("--output", default="generated/generated_logic_checks.sps")
    parser.add_argument("--report", default="generated/generated_logic_compare.json")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    existing_sps_path = Path(args.existing_sps) if args.existing_sps else None
    output_path = Path(args.output)
    report_path = Path(args.report)

    specs = load_var_specs(workbook_path)
    normalization = normalize_logic_sheet_variables(workbook_path, specs)
    rules = load_logic_rules(workbook_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(render_spss(rules, specs), encoding="utf-8-sig")
    existing_do_ifs = extract_existing_logic(existing_sps_path) if existing_sps_path and existing_sps_path.exists() else None
    write_compare_report(report_path, rules, specs, existing_do_ifs, normalization)

    print(f"wrote {output_path}")
    print(f"wrote {report_path}")
    print(f"rules: {len(rules)}")
    print(f"logic sheet corrections: {len(normalization.corrected)}")
    print(f"logic sheet unresolved: {len(normalization.unresolved)}")


if __name__ == "__main__":
    main()
